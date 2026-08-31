"""Construction du classeur Excel du rapport d'inspection.

Les graphiques sont des graphiques Excel natifs, pas des images : le
destinataire peut les redimensionner, changer la plage, imprimer proprement,
et surtout remonter aux chiffres qui les alimentent. Les donnees sources sont
laissees visibles sur leur propre feuille — masquer les chiffres d'un document
d'audit lui oterait sa valeur de preuve.

Palette alignee sur le back-office : ardoise pour l'activite normale, rouge
pour l'echec, ambre pour l'avertissement.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .services import CRITIQUE, ELEVE, MODERE

ARDOISE, ROUGE, AMBRE, VERT, GRIS = "374151", "DC2626", "D97706", "16A34A", "6B7280"

_TITRE = Font(bold=True, size=14, color="111827")
_ENTETE = Font(bold=True, color="FFFFFF")
_FOND_ENTETE = PatternFill("solid", fgColor=ARDOISE)
_BORDURE = Border(*[Side(style="thin", color="E5E7EB")] * 4)

_COULEUR_SEVERITE = {
    CRITIQUE: ("FEE2E2", "991B1B", "Critique"),
    ELEVE:    ("FEF3C7", "92400E", "Élevé"),
    MODERE:   ("F3F4F6", "374151", "Modéré"),
}


def _entetes(ws, ligne, valeurs, largeurs=None):
    for i, v in enumerate(valeurs, start=1):
        c = ws.cell(row=ligne, column=i, value=v)
        c.font, c.fill = _ENTETE, _FOND_ENTETE
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if largeurs:
            ws.column_dimensions[get_column_letter(i)].width = largeurs[i - 1]
    ws.row_dimensions[ligne].height = 22


def _feuille_synthese(wb, r):
    ws = wb.active
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 62

    ws["A1"] = "Rapport d'inspection des connexions"
    ws["A1"].font = _TITRE
    ws["A2"] = (f"Période du {r.debut:%d/%m/%Y %H:%M} au {r.fin:%d/%m/%Y %H:%M} "
                f"({r.jours} jour{'s' if r.jours > 1 else ''})")
    ws["A2"].font = Font(italic=True, color="6B7280")

    lignes = [
        ("Événements enregistrés", r.total, ""),
        ("Connexions réussies", r.connexions, ""),
        ("Déconnexions", r.deconnexions, ""),
        ("Tentatives refusées", r.echecs, "Échecs d'authentification"),
        ("Taux d'échec", f"{r.taux_echec} %", "Part des tentatives refusées"),
        ("Comptes actifs", r.comptes_actifs, "Identifiants s'étant connectés au moins une fois"),
        ("Comptes visés par un échec", r.comptes_vises,
         "Identifiants distincts ayant subi au moins un refus"),
        ("dont identifiants inexistants", r.echecs_compte_inconnu,
         "Refus portant sur un identifiant qui n'existe pas : signe d'un sondage"),
        ("Adresses à l'origine d'échecs", r.sources_echec, ""),
        ("Alertes", len(r.alertes), f"dont {len(r.alertes_critiques)} critique(s)"),
    ]
    _entetes(ws, 4, ["Indicateur", "Valeur", "Lecture"])
    for i, (label, valeur, aide) in enumerate(lignes, start=5):
        ws.cell(row=i, column=1, value=label).border = _BORDURE
        c = ws.cell(row=i, column=2, value=valeur)
        c.alignment = Alignment(horizontal="right")
        c.border = _BORDURE
        c.font = Font(bold=True)
        a = ws.cell(row=i, column=3, value=aide)
        a.font = Font(color="6B7280", size=9)
        a.border = _BORDURE

    depart = 5 + len(lignes) + 2
    if r.alertes_critiques:
        ws.cell(row=depart, column=1,
                value=f"{len(r.alertes_critiques)} alerte(s) critique(s) — voir la feuille « Alertes »"
                ).font = Font(bold=True, color="991B1B")
    else:
        ws.cell(row=depart, column=1,
                value="Aucune alerte critique sur la période."
                ).font = Font(bold=True, color="166534")
    ws.freeze_panes = "A5"
    return ws


def _feuille_alertes(wb, r):
    ws = wb.create_sheet("Alertes")
    ws.sheet_view.showGridLines = False
    _entetes(ws, 1, ["Sévérité", "Motif", "Cible", "Constat", "Conduite à tenir"],
             [12, 38, 26, 52, 66])
    if not r.alertes:
        ws.cell(row=2, column=1, value="Aucune anomalie détectée sur la période.").font = Font(italic=True)
        ws.freeze_panes = "A2"
        return ws
    for i, a in enumerate(r.alertes, start=2):
        fond, texte, libelle = _COULEUR_SEVERITE[a.severite]
        c = ws.cell(row=i, column=1, value=libelle)
        c.fill = PatternFill("solid", fgColor=fond)
        c.font = Font(bold=True, color=texte)
        for col, valeur in enumerate([a.motif, a.cible, a.detail, a.conduite], start=2):
            cel = ws.cell(row=i, column=col, value=valeur)
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            cel.border = _BORDURE
        ws.row_dimensions[i].height = 46
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(r.alertes) + 1}"
    return ws


def _feuille_donnees(wb, r):
    """Series alimentant les graphiques, laissees lisibles."""
    ws = wb.create_sheet("Données")
    ws.sheet_view.showGridLines = False

    _entetes(ws, 1, ["Jour", "Connexions", "Échecs"], [14, 14, 12])
    for i, l in enumerate(r.par_jour, start=2):
        ws.cell(row=i, column=1, value=l['jour'].strftime("%d/%m"))
        ws.cell(row=i, column=2, value=l['connexions'])
        ws.cell(row=i, column=3, value=l['echecs'])
    fin_jours = 1 + len(r.par_jour)

    col = 5
    _entetes(ws, 1, [""] * 4, None)
    ws.cell(row=1, column=col, value="Heure").font = _ENTETE
    ws.cell(row=1, column=col).fill = _FOND_ENTETE
    ws.cell(row=1, column=col + 1, value="Connexions").font = _ENTETE
    ws.cell(row=1, column=col + 1).fill = _FOND_ENTETE
    ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(col + 1)].width = 14
    for i, l in enumerate(r.par_heure, start=2):
        ws.cell(row=i, column=col, value=f"{l['heure']:02d}h")
        ws.cell(row=i, column=col + 1, value=l['connexions'])

    col = 8
    for titre, serie, largeur in (("Compte visé", r.top_comptes_vises, 30),
                                  ("Adresse source", r.top_sources, 22)):
        ws.cell(row=1, column=col, value=titre).font = _ENTETE
        ws.cell(row=1, column=col).fill = _FOND_ENTETE
        ws.cell(row=1, column=col + 1, value="Échecs").font = _ENTETE
        ws.cell(row=1, column=col + 1).fill = _FOND_ENTETE
        ws.column_dimensions[get_column_letter(col)].width = largeur
        ws.column_dimensions[get_column_letter(col + 1)].width = 10
        for i, l in enumerate(serie, start=2):
            ws.cell(row=i, column=col, value=str(l['valeur']))
            ws.cell(row=i, column=col + 1, value=l['nombre'])
        col += 3
    return ws, fin_jours


def _feuille_graphiques(wb, r, ws_donnees, fin_jours):
    ws = wb.create_sheet("Graphiques")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Lecture graphique de la période"
    ws["A1"].font = _TITRE

    d = ws_donnees.title

    # Activite quotidienne : la superposition connexions / echecs fait ressortir
    # une pointe d'echecs sans activite legitime.
    g1 = LineChart()
    g1.title = "Activité quotidienne"
    g1.y_axis.title, g1.x_axis.title = "Événements", "Jour"
    g1.height, g1.width = 8, 20
    g1.add_data(Reference(ws_donnees, min_col=2, max_col=3, min_row=1, max_row=fin_jours),
                titles_from_data=True)
    g1.set_categories(Reference(ws_donnees, min_col=1, min_row=2, max_row=fin_jours))
    g1.series[0].graphicalProperties.line.solidFill = ARDOISE
    g1.series[1].graphicalProperties.line.solidFill = ROUGE
    g1.series[1].graphicalProperties.line.width = 28000
    ws.add_chart(g1, "A3")

    # Repartition horaire : une activite nocturne reguliere se voit ici.
    g2 = BarChart()
    g2.type, g2.title = "col", "Connexions par heure de la journée"
    g2.y_axis.title, g2.x_axis.title = "Connexions", "Heure"
    g2.height, g2.width = 8, 20
    g2.legend = None
    g2.add_data(Reference(ws_donnees, min_col=6, min_row=1, max_row=25), titles_from_data=True)
    g2.set_categories(Reference(ws_donnees, min_col=5, min_row=2, max_row=25))
    g2.series[0].graphicalProperties.solidFill = ARDOISE
    ws.add_chart(g2, "A21")

    if r.top_comptes_vises:
        g3 = BarChart()
        g3.type, g3.title = "bar", "Comptes les plus visés par des échecs"
        g3.height, g3.width = 8, 20
        g3.legend = None
        n = 1 + len(r.top_comptes_vises)
        g3.add_data(Reference(ws_donnees, min_col=9, min_row=1, max_row=n), titles_from_data=True)
        g3.set_categories(Reference(ws_donnees, min_col=8, min_row=2, max_row=n))
        g3.series[0].graphicalProperties.solidFill = ROUGE
        ws.add_chart(g3, "A39")

    if r.top_sources:
        g4 = BarChart()
        g4.type, g4.title = "bar", "Adresses à l'origine du plus d'échecs"
        g4.height, g4.width = 8, 20
        g4.legend = None
        n = 1 + len(r.top_sources)
        g4.add_data(Reference(ws_donnees, min_col=12, min_row=1, max_row=n), titles_from_data=True)
        g4.set_categories(Reference(ws_donnees, min_col=11, min_row=2, max_row=n))
        g4.series[0].graphicalProperties.solidFill = AMBRE
        ws.add_chart(g4, "A57")
    return ws


def _feuille_journal(wb, r, evenements):
    ws = wb.create_sheet("Journal")
    ws.sheet_view.showGridLines = False
    _entetes(ws, 1, ["Date et heure", "Événement", "Identifiant", "Compte connu",
                     "Profil", "Centre", "Adresse IP"],
             [20, 20, 30, 14, 12, 34, 18])
    libelles = {"connexion": "Connexion", "deconnexion": "Déconnexion", "echec": "Échec"}
    for i, e in enumerate(evenements, start=2):
        ws.cell(row=i, column=1, value=e.date_evenement.strftime("%d/%m/%Y %H:%M:%S"))
        c = ws.cell(row=i, column=2, value=libelles.get(e.type_evenement, e.type_evenement))
        if e.type_evenement == "echec":
            c.font = Font(bold=True, color="991B1B")
        ws.cell(row=i, column=3, value=e.username or "—")
        ws.cell(row=i, column=4, value="oui" if e.utilisateur_id else "non")
        ws.cell(row=i, column=5, value="Apprenant" if e.est_apprenant else "Agent")
        ws.cell(row=i, column=6, value=e.centre.nom_centre if e.centre_id else "—")
        ws.cell(row=i, column=7, value=e.adresse_ip or "—")
    ws.freeze_panes = "A2"
    if evenements:
        ws.auto_filter.ref = f"A1:G{len(evenements) + 1}"
    return ws


def construire_classeur(rapport, evenements):
    """Retourne le classeur serialise, pret a etre joint a un courriel."""
    wb = Workbook()
    _feuille_synthese(wb, rapport)
    _feuille_alertes(wb, rapport)
    ws_donnees, fin_jours = _feuille_donnees(wb, rapport)
    _feuille_graphiques(wb, rapport, ws_donnees, fin_jours)
    _feuille_journal(wb, rapport, evenements)
    wb.move_sheet("Données", offset=2)      # les series brutes apres les graphiques
    flux = BytesIO()
    wb.save(flux)
    return flux.getvalue()
