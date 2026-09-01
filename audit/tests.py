"""Tests de l'application d'audit.

Chaque detection a son test : ce sont elles qui donnent sa valeur au rapport,
et une detection qui cesse silencieusement de se declencher est pire que pas
de detection du tout, puisqu'elle laisse croire que rien ne se passe.
"""
import datetime as dt
from io import BytesIO

from django.core import mail
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import HistoriqueConnexion, Utilisateur
from audit import services
from audit.classeur import construire_classeur


class BaseAudit(TestCase):

    def setUp(self):
        self.maintenant = timezone.now()
        self.agent = Utilisateur.objects.create_user(
            username='agent.un', password='x', nom='Un', prenom='Agent',
            email='agent.un@example.invalid', user_type='membre')

    def evt(self, username, type_evt, quand, ip='203.0.113.1', user=None, apprenant=False):
        return HistoriqueConnexion.objects.create(
            utilisateur=user, username=username, est_apprenant=apprenant,
            type_evenement=type_evt, date_evenement=quand, adresse_ip=ip)

    def motifs(self, rapport):
        return {a.motif for a in rapport.alertes}


class DetectionTests(BaseAudit):

    def test_sondage_de_comptes_depuis_une_ip(self):
        for i in range(5):
            self.evt(f'inconnu{i}', 'echec', self.maintenant - dt.timedelta(minutes=i), '203.0.113.7')
        r = services.construire_rapport(jours=7)
        alerte = next(a for a in r.alertes if 'Sondage' in a.motif)
        self.assertEqual(alerte.severite, services.CRITIQUE)
        self.assertEqual(alerte.cible, '203.0.113.7')

    def test_sondage_sous_le_seuil_ne_declenche_pas(self):
        self.evt('inconnu1', 'echec', self.maintenant, '203.0.113.7')
        self.assertNotIn("Sondage de comptes depuis une meme adresse",
                         self.motifs(services.construire_rapport(jours=7)))

    def test_compromission_probable_apres_serie_d_echecs(self):
        base = self.maintenant - dt.timedelta(days=1)
        for i in range(6):
            self.evt('agent.un', 'echec', base + dt.timedelta(minutes=i), user=self.agent)
        self.evt('agent.un', 'connexion', base + dt.timedelta(minutes=30), user=self.agent)
        r = services.construire_rapport(jours=7)
        alerte = next(a for a in r.alertes if 'Connexion reussie' in a.motif)
        self.assertEqual(alerte.severite, services.CRITIQUE)

    def test_reussite_longtemps_apres_les_echecs_ne_declenche_pas(self):
        """Au-dela d'une heure, le lien de causalite n'est plus etabli."""
        base = self.maintenant - dt.timedelta(days=2)
        for i in range(6):
            self.evt('agent.un', 'echec', base + dt.timedelta(minutes=i), user=self.agent)
        self.evt('agent.un', 'connexion', base + dt.timedelta(hours=5), user=self.agent)
        self.assertNotIn("Connexion reussie apres une serie d'echecs",
                         self.motifs(services.construire_rapport(jours=7)))

    def test_acharnement_sur_un_compte(self):
        for i in range(6):
            self.evt('agent.un', 'echec', self.maintenant - dt.timedelta(minutes=i), user=self.agent)
        self.assertIn("Echecs repetes sur un meme compte",
                      self.motifs(services.construire_rapport(jours=7)))

    def test_compte_utilise_depuis_plusieurs_adresses(self):
        for i, ip in enumerate(['10.0.0.1', '10.0.0.2', '10.0.0.3']):
            self.evt('agent.un', 'connexion', self.maintenant - dt.timedelta(hours=i),
                     ip=ip, user=self.agent)
        self.assertIn("Compte utilise depuis plusieurs adresses",
                      self.motifs(services.construire_rapport(jours=7)))

    def test_connexions_hors_heures_ouvrees(self):
        for i in range(3):
            nuit = (self.maintenant - dt.timedelta(days=i + 1)).replace(hour=3, minute=0)
            self.evt('agent.un', 'connexion', nuit, user=self.agent)
        self.assertIn("Connexions d'agent hors heures ouvrees",
                      self.motifs(services.construire_rapport(jours=7)))

    def test_apprenant_hors_heures_n_est_pas_signale(self):
        """Un apprenant consulte son dossier a toute heure : c'est attendu."""
        eleve = Utilisateur.objects.create_user(
            username='eleve.un', password='x', nom='Un', prenom='Eleve',
            email='e.un@example.invalid', user_type='eleve')
        for i in range(4):
            nuit = (self.maintenant - dt.timedelta(days=i + 1)).replace(hour=2)
            self.evt('eleve.un', 'connexion', nuit, user=eleve, apprenant=True)
        self.assertNotIn("Connexions d'agent hors heures ouvrees",
                         self.motifs(services.construire_rapport(jours=7)))

    def test_periode_vide_ne_produit_aucune_alerte(self):
        r = services.construire_rapport(jours=7)
        self.assertEqual(r.total, 0)
        self.assertEqual(r.alertes, [])
        self.assertEqual(r.taux_echec, 0.0)

    def test_evenements_hors_periode_sont_exclus(self):
        self.evt('agent.un', 'echec', self.maintenant - dt.timedelta(days=40), user=self.agent)
        self.assertEqual(services.construire_rapport(jours=7).total, 0)


class IndicateurTests(BaseAudit):

    def test_taux_echec_et_comptes_vises(self):
        for i in range(3):
            self.evt('agent.un', 'connexion', self.maintenant - dt.timedelta(hours=i), user=self.agent)
        self.evt('inconnu', 'echec', self.maintenant)
        r = services.construire_rapport(jours=7)
        self.assertEqual((r.connexions, r.echecs), (3, 1))
        self.assertEqual(r.taux_echec, 25.0)
        self.assertEqual(r.comptes_vises, 1)
        self.assertEqual(r.echecs_compte_inconnu, 1)

    def test_serie_quotidienne_couvre_les_jours_sans_evenement(self):
        """Un graphique a trous laisse croire a une absence de donnees."""
        r = services.construire_rapport(jours=7)
        self.assertEqual(len(r.par_jour), 8)
        self.assertEqual(len(r.par_heure), 24)


class ClasseurTests(BaseAudit):

    def test_structure_du_classeur(self):
        self.evt('inconnu', 'echec', self.maintenant, '203.0.113.7')
        r = services.construire_rapport(jours=7)
        wb = load_workbook(BytesIO(construire_classeur(r, list(HistoriqueConnexion.objects.all()))))
        self.assertEqual(wb.sheetnames,
                         ['Synthèse', 'Alertes', 'Graphiques', 'Journal', 'Données'])
        self.assertEqual(len(wb['Graphiques']._charts), 4)

    def test_classeur_sans_donnees_reste_valide(self):
        r = services.construire_rapport(jours=7)
        wb = load_workbook(BytesIO(construire_classeur(r, [])))
        self.assertIn('Synthèse', wb.sheetnames)
        # Sans compte vise ni source, seuls les deux graphiques temporels subsistent.
        self.assertEqual(len(wb['Graphiques']._charts), 2)


@override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
class CommandeTests(BaseAudit):

    def test_envoi_avec_piece_jointe(self):
        self.evt('inconnu', 'echec', self.maintenant, '203.0.113.7')
        call_command('envoyer_rapport_audit', '--jours', '7', '--a', 'audit@example.invalid')
        self.assertEqual(len(mail.outbox), 1)
        message = mail.outbox[0]
        self.assertEqual(message.to, ['audit@example.invalid'])
        self.assertIn('text/html', [t for _, t in message.alternatives])
        nom, _, type_mime = message.attachments[0]
        self.assertTrue(nom.endswith('.xlsx'))
        self.assertIn('spreadsheetml', type_mime)

    def test_sans_destinataire_la_commande_refuse(self):
        with override_settings(AUDIT_DESTINATAIRES=[]):
            with self.assertRaises(CommandError):
                call_command('envoyer_rapport_audit')

    def test_sans_envoi_ne_produit_aucun_courriel(self):
        call_command('envoyer_rapport_audit', '--sans-envoi')
        self.assertEqual(len(mail.outbox), 0)

    def test_periode_invalide_refusee(self):
        with self.assertRaises(CommandError):
            call_command('envoyer_rapport_audit', '--jours', '0', '--sans-envoi')


class DestinataireTests(TestCase):
    """Diffusion parametree depuis l'ecran plutot que depuis .env."""

    def setUp(self):
        from audit.models import DestinataireRapport
        self.Modele = DestinataireRapport

    def test_seules_les_adresses_actives_sont_retenues(self):
        self.Modele.objects.create(email='actif@example.invalid', actif=True)
        self.Modele.objects.create(email='suspendu@example.invalid', actif=False)
        self.assertEqual(self.Modele.adresses_actives(), ['actif@example.invalid'])

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                       AUDIT_DESTINATAIRES=[])
    def test_la_commande_envoie_aux_destinataires_enregistres(self):
        self.Modele.objects.create(email='inspection@example.invalid', actif=True)
        mail.outbox = []
        call_command('envoyer_rapport_audit', '--jours', '7')
        self.assertEqual(mail.outbox[0].to, ['inspection@example.invalid'])

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                       AUDIT_DESTINATAIRES=['fichier@example.invalid'])
    def test_les_adresses_du_fichier_et_de_l_ecran_se_cumulent(self):
        """Une installation deja configuree par .env ne perd pas ses destinataires."""
        self.Modele.objects.create(email='ecran@example.invalid', actif=True)
        mail.outbox = []
        call_command('envoyer_rapport_audit', '--jours', '7')
        self.assertEqual(sorted(mail.outbox[0].to),
                         ['ecran@example.invalid', 'fichier@example.invalid'])

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                       AUDIT_DESTINATAIRES=['fichier@example.invalid'])
    def test_l_option_a_remplace_tout(self):
        """Envoi ponctuel cible : il ne doit pas partir a toute la liste."""
        self.Modele.objects.create(email='ecran@example.invalid', actif=True)
        mail.outbox = []
        call_command('envoyer_rapport_audit', '--jours', '7', '--a', 'ponctuel@example.invalid')
        self.assertEqual(mail.outbox[0].to, ['ponctuel@example.invalid'])

    @override_settings(AUDIT_DESTINATAIRES=[])
    def test_aucun_destinataire_actif_la_commande_refuse(self):
        self.Modele.objects.create(email='suspendu@example.invalid', actif=False)
        with self.assertRaises(CommandError):
            call_command('envoyer_rapport_audit')


class ReglageDiffusionTests(TestCase):
    """Planification de l'envoi automatique : echeances et non-double-envoi."""

    def setUp(self):
        from audit.models import ReglageDiffusion
        self.Reglage = ReglageDiffusion
        self.r = ReglageDiffusion.charge()

    def test_desactive_n_est_jamais_du(self):
        self.r.frequence = self.Reglage.DESACTIVE
        self.assertFalse(self.r.est_du(timezone.now()))

    def test_quotidien_du_puis_plus_du_apres_diffusion(self):
        maintenant = timezone.localtime(timezone.now())
        self.r.frequence = self.Reglage.QUOTIDIEN
        self.r.heure = maintenant.hour
        self.r.derniere_diffusion = None
        self.assertTrue(self.r.est_du(timezone.now()))
        self.r.derniere_diffusion = timezone.now()
        self.assertFalse(self.r.est_du(timezone.now()))

    def test_creneau_avant_l_heure_est_reporte_a_la_veille(self):
        maintenant = timezone.localtime(timezone.now())
        self.r.frequence = self.Reglage.QUOTIDIEN
        # une heure future aujourd'hui : le dernier creneau echu est hier.
        self.r.heure = (maintenant.hour + 1) % 24
        creneau = self.r.creneau_courant(timezone.now())
        self.assertIsNotNone(creneau)
        self.assertLessEqual(creneau, timezone.now())

    def test_prochaine_est_dans_le_futur(self):
        self.r.frequence = self.Reglage.HEBDOMADAIRE
        self.assertGreater(self.r.prochaine(timezone.now()), timezone.now())

    def test_periode_jours_suit_la_frequence(self):
        self.r.frequence = self.Reglage.QUOTIDIEN
        self.assertEqual(self.r.periode_jours, 1)
        self.r.frequence = self.Reglage.HEBDOMADAIRE
        self.assertEqual(self.r.periode_jours, 7)


class CommandeAutoTests(TestCase):
    """Mode --auto : n'envoie qu'a echeance, et jamais deux fois la meme."""

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                       AUDIT_DESTINATAIRES=['inspecteur@example.invalid'])
    def test_auto_desactive_n_envoie_rien(self):
        from audit.models import ReglageDiffusion
        r = ReglageDiffusion.charge(); r.frequence = r.DESACTIVE; r.save()
        mail.outbox = []
        call_command('envoyer_rapport_audit', '--auto')
        self.assertEqual(len(mail.outbox), 0)

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
                       AUDIT_DESTINATAIRES=['inspecteur@example.invalid'])
    def test_auto_envoie_a_echeance_puis_ne_double_pas(self):
        from audit.models import ReglageDiffusion
        maintenant = timezone.localtime(timezone.now())
        r = ReglageDiffusion.charge()
        r.frequence = r.QUOTIDIEN; r.heure = maintenant.hour
        r.derniere_diffusion = None; r.save()
        mail.outbox = []
        call_command('envoyer_rapport_audit', '--auto')
        self.assertEqual(len(mail.outbox), 1)
        r.refresh_from_db()
        self.assertIsNotNone(r.derniere_diffusion)
        # deuxieme passage immediat : aucune nouvelle diffusion.
        call_command('envoyer_rapport_audit', '--auto')
        self.assertEqual(len(mail.outbox), 1)
