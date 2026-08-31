from django.shortcuts import get_object_or_404, redirect, render
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils.http import url_has_allowed_host_and_scheme
from django.http import Http404, HttpResponse
from django.conf import settings
import qrcode
import csv
import io
import os
import uuid
from django.db.models import Q, Sum, Count
from accounts.models import Utilisateur, Formateur, MembreAdministration, HistoriqueConnexion
from .forms import AgentForm
from courses.models import TypeFrais, TrancheFrais
from courses.forms import TypeFraisForm, TrancheFraisFormSet
from django.utils import timezone
from .permissions import require_permission
from .ui import gabarit
from .models import (
    Direction_reg, Filiere, CentreFormation, Module, 
    Frais, Cours, Inscription, Paiement, CentreEtFiliere,PieceJointeInscription,
    DocumentEleve,AnneeScolaire,Dette
)
from .forms import (
    DirectionRegForm, FiliereForm, CentreFormationForm, ModuleForm,
    FraisForm, CoursForm, InscriptionForm, PaiementForm, PaiementAdminForm, CentreEtFiliereForm,
    PieceJointeFormSet,FraisFormSet,AnneeScolaireForm, EleveForm
)
from accounts.models import Eleve,Formateur
from .admin_filters import FormationFilter,FiliereFilter,SubscriptionFilter
from django.db.models import Sum
from datetime import datetime
from django.urls import reverse
from .views import _base_qs, _get_scope, _pdf_header_lines, _draw_pdf_watermark


def _filiere_modules_map():
    """{filiere_id: [{id, nom_module}, ...]} pour filtrer côté client le select Module selon le métier choisi."""
    mapping = {}
    for filiere in Filiere.objects.filter(is_active=True).prefetch_related('modules'):
        mapping[filiere.id] = [
            {'id': m.id, 'nom': m.nom_module} for m in filiere.modules.all()
        ]
    return mapping


def _agent_in_scope(agent, centres_qs):
    """True si cet agent (Formateur ou MembreAdministration) appartient à l'un des centres du périmètre donné."""
    centre_ids = list(centres_qs.values_list("id", flat=True))
    membre = getattr(agent, 'membreadministration', None)
    if membre and membre.structure_id in centre_ids:
        return True
    formateur = getattr(agent, 'formateur', None)
    if formateur and formateur.centre_id in centre_ids:
        return True
    return False


# DASHBOARD
@require_permission('courses.voir_statistiques')
def admin_dashboard(request):
    """Tableau de bord de direction.

    Il repond a trois questions, dans cet ordre : qu'y a-t-il a traiter
    aujourd'hui, ou en est le recouvrement, et de quoi l'offre est-elle faite.
    Les volumes bruts viennent apres : un effectif ne se regarde qu'une fois
    par mois, une pile de dossiers en attente tous les matins.
    """
    from datetime import timedelta
    from django.db.models import Count
    from accounts.models import HistoriqueConnexion
    from actualites.models import Actualite

    inscriptions, dettes, paiements = _base_qs(request.user)[:3]
    maintenant = timezone.now()
    depuis_7j = maintenant - timedelta(days=7)

    du = dettes.aggregate(s=Sum('montant_total'))['s'] or 0
    encaisse = paiements.aggregate(s=Sum('montant_paiement'))['s'] or 0

    a_traiter = [
        {
            'libelle': "Inscriptions en attente de validation",
            'valeur': inscriptions.filter(statut='en_cours').count(),
            'url': reverse('bsb_admin:subscription_list') + '?statut=en_cours',
            'ton': 'alerte',
            'aide': "Dossiers deposes par des apprenants, sans decision du centre.",
        },
        {
            'libelle': "Tentatives de connexion refusees (7 jours)",
            'valeur': HistoriqueConnexion.objects.filter(
                type_evenement='echec', date_evenement__gte=depuis_7j).count(),
            'url': reverse('bsb_admin:historique_connexion_list') + '?type_evenement=echec',
            'ton': 'garde',
            'aide': "Un pic sans explication merite une inspection.",
        },
        {
            'libelle': "Actualites en brouillon",
            'valeur': Actualite.objects.filter(statut='brouillon').count(),
            'url': reverse('bsb_actualites:actualite_list'),
            'ton': 'attente',
            'aide': "Redigees mais jamais publiees.",
        },
    ]

    recouvrement = {
        'du': du,
        'encaisse': encaisse,
        'reste': max(du - encaisse, 0),
        'taux': round(100 * encaisse / du, 1) if du else 0,
        'versements': paiements.count(),
    }

    volumes = [
        {'libelle': 'Apprenants inscrits', 'valeur': inscriptions.values('eleve').distinct().count()},
        {'libelle': 'Formations actives', 'valeur': CentreEtFiliere.objects.filter(is_active=True).count()},
        {'libelle': 'Centres de formation', 'valeur': CentreFormation.objects.count()},
        {'libelle': 'Metiers', 'valeur': Filiere.objects.count()},
        {'libelle': 'Formateurs', 'valeur': Formateur.objects.count()},
        {'libelle': 'Directions inter-regionales', 'valeur': Direction_reg.objects.count()},
    ]

    # Repartition des dossiers : lue d'un coup d'oeil, elle dit si la chaine
    # de validation avance ou si les dossiers s'accumulent quelque part.
    par_statut = list(
        inscriptions.order_by().values('statut').annotate(n=Count('id')).order_by('-n')
    )
    libelles_statut = dict(Inscription.STATUT_CHOICE)
    total_inscriptions = sum(l['n'] for l in par_statut) or 1
    repartition = [{
        'libelle': libelles_statut.get(l['statut'], l['statut']),
        'nombre': l['n'],
        'part': round(100 * l['n'] / total_inscriptions),
    } for l in par_statut]

    formations_recentes = (
        CentreEtFiliere.objects
        .filter(is_active=True)
        .select_related('centre', 'filiere', 'annee_prog')
        .annotate(total_frais=Sum('frais__montant'))
        .order_by('-date_lancement')[:8]
    )

    return render(request, "admin/admin_dashboard/dashboard.html", {
        'a_traiter': a_traiter,
        'recouvrement': recouvrement,
        'volumes': volumes,
        'repartition': repartition,
        'total_inscriptions': sum(l['n'] for l in par_statut),
        'formations_recentes': formations_recentes,
    })


#  DIRECTION CRUD
@require_permission('courses.gerer_directions')
def direction_list(request):
    directions = Direction_reg.objects.all().order_by('-date_modification')

    search = request.GET.get('q', '').strip()
    if search:
        directions = directions.filter(nom_direction__icontains=search)

    paginator = Paginator(directions, 10)
    page = request.GET.get('page')
    directions = paginator.get_page(page)
    return render(request, 'admin/direction/direction_list.html', {'directions': directions, 'search': search})

@require_permission('courses.gerer_directions')
def direction_create(request):
    if request.method == 'POST':
        form = DirectionRegForm(request.POST)
        if form.is_valid():
            direction = form.save()
            messages.success(request, f'Direction "{direction.nom_direction}" créée avec succès!')
            return redirect('bsb_admin:direction_list')
    else:
        form = DirectionRegForm()
    return render(request, 'admin/direction/direction_form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.gerer_directions')
def direction_import_template(request):
    from .bulk_import_registry import SPEC_DIRECTION
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_DIRECTION)


@require_permission('courses.gerer_directions')
def direction_import(request):
    from .bulk_import_registry import SPEC_DIRECTION
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_DIRECTION)


@require_permission('courses.gerer_directions')
def direction_update(request, id):
    direction = get_object_or_404(Direction_reg, id=id)
    if request.method == 'POST':
        form = DirectionRegForm(request.POST, instance=direction)
        if form.is_valid():
            direction = form.save()
            messages.success(request, f'Direction "{direction.nom_direction}" modifiée avec succès!')
            return redirect('bsb_admin:direction_list')
    else:
        form = DirectionRegForm(instance=direction)
    return render(request, 'admin/direction/direction_form.html', {'form': form, 'action': 'Modifier', 'object': direction})

@require_permission('courses.gerer_directions')
def direction_delete(request, id):
    direction = get_object_or_404(Direction_reg, id=id)
    if request.method == 'POST':
        nom = direction.nom_direction
        direction.delete()
        messages.success(request, f'Direction "{nom}" supprimée avec succès!')
        return redirect('bsb_admin:direction_list')
    return render(request, 'admin/direction/direction_confirm_delete.html', {'object': direction})


# FIELD CRUD
@require_permission('courses.gerer_metiers')
def field_list(request):
    fields = Filiere.objects.all().order_by('-date_modification')
    f=FiliereFilter(request.GET,queryset=fields)
    paginator = Paginator(f.qs, 10)
    page = request.GET.get('page')
    fields = paginator.get_page(page)
    return render(request, 'admin/field/list.html', {'fields': fields,'filter':f})

@require_permission('courses.gerer_metiers')
def field_create(request):
    if request.method == 'POST':
        form = FiliereForm(request.POST)
        if form.is_valid():
            field = form.save()
            next_url=request.GET.get('next') or request.POST.get('next')
            if next_url and url_has_allowed_host_and_scheme(
                next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
            ):
                return redirect(next_url)
            messages.success(request, f'Filière "{field.nom_filiere}" créée avec succès!')
            return redirect('bsb_admin:field_list')
    else:
        form = FiliereForm()
    return render(request, 'admin/field/form.html', {'form': form, 'action': 'Créer'})

@require_permission('courses.gerer_metiers')
def field_update(request, id):
    field = get_object_or_404(Filiere, id=id)
    if request.method == 'POST':
        form = FiliereForm(request.POST, request.FILES, instance=field)
        if form.is_valid():
            field = form.save()
            from .views import _process_metier_modules
            _process_metier_modules(request, field, form)
            messages.success(request, f'Filière "{field.nom_filiere}" modifiée avec succès!')
            return redirect('bsb_admin:field_list')
    else:
        form = FiliereForm(instance=field)
    return render(request, 'member/filiere/metier_form.html', {'form': form, 'action': 'Modifier', 'object': field})

@require_permission('courses.gerer_metiers')
def field_delete(request, id):
    field = get_object_or_404(Filiere, id=id)
    if request.method == 'POST':
        nom = field.nom_filiere
        field.delete()
        messages.success(request, f'Filière "{nom}" supprimée avec succès!')
        return redirect('bsb_admin:field_list')
    return render(request, 'member/filiere/confirm_delete.html', {'object': field})


# CENTER CRUD
@require_permission('courses.gerer_centres')
def center_list(request):
    q = request.GET.get('q', '').strip()
    direction_id = request.GET.get('direction', '').strip()
    niveau = request.GET.get('niveau', '').strip()

    qs = (
        CentreFormation.objects
        .filter(pk__in=_get_scope(request.user)[0])
        .select_related('direction', 'province')
        .order_by('nom_centre')
    )

    if q:
        qs = qs.filter(
            Q(nom_centre__icontains=q) |
            Q(adresse__icontains=q) |
            Q(province__nom_province__icontains=q)  # ← adapter si le champ s'appelle autrement
        ).distinct()

    if direction_id and direction_id.isdigit():
        qs = qs.filter(direction_id=int(direction_id))

    if niveau and niveau.isdigit():
        qs = qs.filter(niveau_centre=int(niveau))
    else:
        niveau = ''

    try:
        page_number = int(request.GET.get('page', 1))
        if page_number < 1:
            page_number = 1
    except (ValueError, TypeError):
        page_number = 1

    paginator = Paginator(qs, 10)
    centers = paginator.get_page(page_number)

    directions = Direction_reg.objects.all().order_by('nom_direction')

    niveaux = (
        CentreFormation.objects
        .exclude(niveau_centre__isnull=True)
        .values_list('niveau_centre', flat=True)
        .distinct()
        .order_by('niveau_centre')
    )

    return render(request, 'admin/center/list.html', {
        'centers': centers,
        'q': q,
        'direction_id': direction_id,
        'niveau': niveau,
        'directions': directions,
        'niveaux': niveaux,
    })

@require_permission('courses.gerer_centres')
def center_create(request):
    direction_queryset = _get_scope(request.user)[1]
    if request.method == 'POST':
        form = CentreFormationForm(request.POST, direction_queryset=direction_queryset)
        if form.is_valid():
            center = form.save()
            next_url=request.GET.get('next') or request.POST.get('next')
            if next_url and url_has_allowed_host_and_scheme(
                next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
            ):
                return redirect(next_url)
            messages.success(request, f'Centre "{center.nom_centre}" créé avec succès!')
            return redirect('bsb_admin:center_list')
    else:
        form = CentreFormationForm(direction_queryset=direction_queryset)
    return render(request, 'admin/center/form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.gerer_centres')
def center_import_template(request):
    from .bulk_import_registry import SPEC_CENTRE
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_CENTRE)


@require_permission('courses.gerer_centres')
def center_import(request):
    from .bulk_import_registry import SPEC_CENTRE
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_CENTRE)


@require_permission('courses.gerer_centres')
def center_update(request, id):
    centres_qs, direction_queryset, _ = _get_scope(request.user)
    center = get_object_or_404(CentreFormation, id=id, pk__in=centres_qs)
    if request.method == 'POST':
        form = CentreFormationForm(request.POST, instance=center, direction_queryset=direction_queryset)
        if form.is_valid():
            center = form.save()
            messages.success(request, f'Centre "{center.nom_centre}" modifié avec succès!')
            return redirect('bsb_admin:center_list')
    else:
        form = CentreFormationForm(instance=center, direction_queryset=direction_queryset)
    return render(request, 'admin/center/form.html', {'form': form, 'action': 'Modifier', 'object': center})

@require_permission('courses.gerer_centres')
def center_delete(request, id):
    center = get_object_or_404(CentreFormation, id=id, pk__in=_get_scope(request.user)[0])
    if request.method == 'POST':
        nom = center.nom_centre
        center.delete()
        messages.success(request, f'Centre "{nom}" supprimé avec succès!')
        return redirect('bsb_admin:center_list')
    return render(request, 'admin/center/confirm_delete.html', {'object': center})


# MODULE CRUD
@require_permission('courses.gerer_modules')
def module_list(request):
    modules = Module.objects.prefetch_related('filieres').all().order_by('-date_creation')

    q = request.GET.get('q', '').strip()
    if q:
        modules = modules.filter(
            Q(nom_module__icontains=q) |
            Q(filieres__nom_filiere__icontains=q)
        ).distinct()

    paginator = Paginator(modules, 10)
    page = request.GET.get('page')
    modules = paginator.get_page(page)
    return render(request, 'admin/module/list.html', {'modules': modules, 'q': q})

@require_permission('courses.gerer_modules')
def module_create(request):
    if request.method == 'POST':
        form = ModuleForm(request.POST)
        if form.is_valid():
            module = form.save()
            messages.success(request, f'Module "{module.nom_module}" créé avec succès!')
            return redirect('bsb_admin:module_list')
    else:
        initial = {}
        filiere_id = request.GET.get('filiere')
        if filiere_id:
            initial['filieres'] = [filiere_id]
        form = ModuleForm(initial=initial)
    return render(request, 'admin/module/form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.gerer_modules')
def module_import_template(request):
    from .bulk_import_registry import SPEC_MODULE
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_MODULE)


@require_permission('courses.gerer_modules')
def module_import(request):
    from .bulk_import_registry import SPEC_MODULE
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_MODULE)


@require_permission('courses.gerer_modules')
def module_update(request, id):
    module = get_object_or_404(Module, id=id)
    if request.method == 'POST':
        form = ModuleForm(request.POST, instance=module)
        if form.is_valid():
            module = form.save()
            messages.success(request, f'Module "{module.nom_module}" modifié avec succès!')
            return redirect('bsb_admin:module_list')
    else:
        form = ModuleForm(instance=module)
    return render(request, 'admin/module/form.html', {'form': form, 'action': 'Modifier', 'object': module})

@require_permission('courses.gerer_modules')
def module_delete(request, id):
    module = get_object_or_404(Module, id=id)
    if request.method == 'POST':
        nom = module.nom_module
        module.delete()
        messages.success(request, f'Module "{nom}" supprimé avec succès!')
        return redirect('bsb_admin:module_list')
    return render(request, 'admin/module/confirm_delete.html', {'object': module})


# FEE CRUD
@require_permission('courses.gerer_frais')
def fees_list(request):
    fees = Frais.objects.select_related(
        'formation',
        'formation__centre',
        'formation__filiere',
        'type_frais'
    ).all().order_by('-date_creation')

    q = request.GET.get('q', '').strip()
    if q:
        fees = fees.filter(
            Q(formation__centre__nom_centre__icontains=q) |
            Q(formation__filiere__nom_filiere__icontains=q) |
            Q(type_frais__libelle__icontains=q)
        )

    paginator = Paginator(fees, 10)
    page = request.GET.get('page')
    fees = paginator.get_page(page)

    return render(request, 'admin/fees/list.html', {'fees': fees, 'q': q})

@require_permission('courses.gerer_frais')
def fees_create(request):
    if request.method == 'POST':
        form = FraisForm(request.POST)
        if form.is_valid():
            fees = form.save()
            messages.success(request, f'Frais "{fees.type_frais}" créé avec succès!')
            return redirect('bsb_admin:fees_list')
    else:
        form = FraisForm()
    return render(request, 'admin/fees/form.html', {'form': form, 'action': 'Créer'})

@require_permission('courses.gerer_frais')
def fees_update(request, id):
    fees = get_object_or_404(Frais, id=id)
    if request.method == 'POST':
        form = FraisForm(request.POST, instance=fees)
        if form.is_valid():
            fees = form.save()
            messages.success(request, f'Frais "{fees.type_frais}" modifié avec succès!')
            return redirect('bsb_admin:fees_list')
    else:
        form = FraisForm(instance=fees)
    return render(request, 'admin/fees/form.html', {'form': form, 'action': 'Modifier', 'object': fees})

@require_permission('courses.gerer_frais')
def fees_delete(request, id):
    fees = get_object_or_404(Frais, id=id)
    if request.method == 'POST':
        libelle = fees.type_frais
        fees.delete()
        messages.success(request, f'Frais "{libelle}" supprimé avec succès!')
        return redirect('bsb_admin:fees_list')
    return render(request, 'admin/fees/confirm_delete.html', {'object': fees})


# COURSE CRUD
@require_permission('courses.gerer_modules')
def course_list(request):
    course = Cours.objects.select_related('module').prefetch_related('module__filieres').order_by('-date_creation')

    q = request.GET.get('q', '').strip()
    if q:
        course = course.filter(
            Q(libelle_cours__icontains=q) |
            Q(module__nom_module__icontains=q) |
            Q(module__filieres__nom_filiere__icontains=q)
        ).distinct()

    paginator = Paginator(course, 10)
    page = request.GET.get('page')
    course = paginator.get_page(page)
    return render(request, 'admin/course/list.html', {'course': course, 'q': q})

@require_permission('courses.gerer_modules')
def course_create(request):
    if request.method == 'POST':
        form = CoursForm(request.POST)
        if form.is_valid():
            course = form.save()
            messages.success(request, f'Cours "{course.libelle_cours}" créé avec succès!')
            return redirect('bsb_admin:course_list')
    else:
        form = CoursForm()
    return render(request, 'admin/course/form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.gerer_modules')
def course_import_template(request):
    from .bulk_import_registry import SPEC_COURS
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_COURS)


@require_permission('courses.gerer_modules')
def course_import(request):
    from .bulk_import_registry import SPEC_COURS
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_COURS)


@require_permission('courses.gerer_modules')
def course_update(request, id):
    course = get_object_or_404(Cours, id=id)
    if request.method == 'POST':
        form = CoursForm(request.POST, instance=course)
        if form.is_valid():
            course = form.save()
            messages.success(request, f'Cours "{course.libelle_cours}" modifié avec succès!')
            return redirect('bsb_admin:course_list')
    else:
        form = CoursForm(instance=course)
    return render(request, 'admin/course/form.html', {'form': form, 'action': 'Modifier', 'object': course})

@require_permission('courses.gerer_modules')
def course_delete(request, id):
    course = get_object_or_404(Cours, id=id)
    if request.method == 'POST':
        libelle = course.libelle_cours
        course.delete()
        messages.success(request, f'Cours "{libelle}" supprimé avec succès!')
        return redirect('bsb_admin:course_list')
    return render(request, 'admin/course/confirm_delete.html', {'object': course})


# SUBSCRIPTION CRUD Faudra peut etre filter par année aussi cça sera bien pour les staistiques
@require_permission('courses.voir_inscriptions')
def subscription_list(request):
    subscriptions = Inscription.objects.select_related('eleve', 'formation__centre')\
    .exclude(statut="en-cours")\
    .order_by('-date_inscription')
    centres_qs, _, scope = _get_scope(request.user)
    multi_centre = scope == "global" or scope == "direction"
    if scope != "global":
        centre_ids = list(centres_qs.values_list("id", flat=True))
        subscriptions = subscriptions.filter(formation__centre_id__in=centre_ids)

    f=SubscriptionFilter(request.GET,queryset=subscriptions)

    centre_id = request.GET.get('centre', '').strip()
    if not multi_centre:
        centre_id = ''
    filtre_actif = bool(
        request.GET.get('recherche') or request.GET.get('statut') or request.GET.get('formation')
    )

    # Filtre (recherche/statut/formation) actif, ou portée mono-centre :
    # liste plate, comme avant.
    if filtre_actif or not multi_centre:
        paginator = Paginator(f.qs, 10)
        subscriptions = paginator.get_page(request.GET.get('page'))
        return render(request, 'admin/subscription/list.html', {
            'mode': 'plat',
            'subscriptions': subscriptions,
            'filter': f,
        })

    # Accordeon centres -> inscriptions : les deux niveaux sont pagines
    # separement (?centre= et ?ipage=), sans recharger la liste des centres.
    compte_par_centre = {
        row['formation__centre_id']: row['n']
        for row in f.qs.values('formation__centre_id').annotate(n=Count('id'))
    }
    centres_annotes = list(centres_qs.order_by('nom_centre'))
    for c in centres_annotes:
        c.nb_inscriptions = compte_par_centre.get(c.id, 0)
    paginator = Paginator(centres_annotes, 10)
    centres_page = paginator.get_page(request.GET.get('page'))

    centre_ouvert = None
    subscriptions = None
    if centre_id:
        centre_ouvert = centres_qs.filter(pk=centre_id).first()
        if centre_ouvert:
            iqs = f.qs.filter(formation__centre_id=centre_ouvert.id)
            ipaginator = Paginator(iqs, 10)
            subscriptions = ipaginator.get_page(request.GET.get('ipage'))

    return render(request, 'admin/subscription/list.html', {
        'mode': 'accordeon',
        'centres_page': centres_page,
        'centre_ouvert': centre_ouvert,
        'subscriptions': subscriptions,
        'filter': f,
    })

@require_permission('courses.valider_inscription')
def subscription_create(request):
    if request.method == 'POST':
        form = InscriptionForm(request.POST)
        if form.is_valid():
            inscription = form.save()
            messages.success(request, 'Inscription créée avec succès!')
            return redirect('bsb_admin:subscription_list')
    else:
        form = InscriptionForm()
    return render(request, 'admin/subscription/form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.valider_inscription')
def subscription_update(request, id):
    subscription = get_object_or_404(Inscription, id=id)
    if subscription.statut not in ('en_cours', 'rejete'):
        messages.error(request, "Cette inscription est validée : elle ne peut plus être modifiée.")
        return redirect('bsb_admin:subscription_list')
    if request.method == 'POST':
        form = InscriptionForm(request.POST, instance=subscription)
        if form.is_valid():
            subscription = form.save()
            messages.success(request, 'Inscription modifiée avec succès!')
            return redirect('bsb_admin:subscription_list')
    else:
        form = InscriptionForm(instance=subscription)
    return render(request, 'admin/subscription/form.html', {'form': form, 'action': 'Modifier', 'object': subscription})

@require_permission('courses.valider_inscription')
def subscription_delete(request, id):
    subscription = get_object_or_404(Inscription, id=id)
    if subscription.statut not in ('en_cours', 'rejete'):
        messages.error(request, "Cette inscription est validée : elle ne peut plus être supprimée.")
        return redirect('bsb_admin:subscription_list')
    if request.method == 'POST':
        subscription.delete()
        messages.success(request, 'Inscription supprimée avec succès!')
        return redirect('bsb_admin:subscription_list')
    return render(request, 'admin/subscription/confirm_delete.html', {'object': subscription})

#Detail d'insciption
# Ici je veux la liste des pièces jointes que l'elève à renseigner pour une inscription
@require_permission('courses.voir_inscriptions')
def subscription_detail(request,id):
    subscription=get_object_or_404(Inscription.objects.select_related('eleve','formation__centre'),id=id)
    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global" and (not subscription.formation_id or not centres_qs.filter(pk=subscription.formation.centre_id).exists()):
        raise PermissionDenied("Vous n'avez pas accès à cette inscription.")
    eleve_docs=DocumentEleve.objects.select_related('piece_requise').filter(inscription=subscription)
    return render(request,'admin/subscription/inscription_detail.html',{'detail':subscription,'eleve_docs':eleve_docs})

#Permet de gére l'inscription en fait 
@require_permission('courses.valider_inscription')
def gerer_inscription(request,id):
     subscription=get_object_or_404(Inscription.objects.select_related('formation__centre','eleve'),id=id)
     centres_qs, _, scope = _get_scope(request.user)
     if scope != "global" and (not subscription.formation_id or not centres_qs.filter(pk=subscription.formation.centre_id).exists()):
         raise PermissionDenied("Vous n'avez pas accès à cette inscription.")
     if request.method == 'POST':
         action=request.POST.get('action')
         if action == 'valide':
             if subscription.formation and subscription.formation.type_formation == 'initiale':
                 conflit = Inscription.objects.filter(
                     eleve=subscription.eleve,
                     annee_scolaire=subscription.annee_scolaire,
                     formation__type_formation='initiale',
                     statut__in=['valide', 'valide_paye'],
                 ).exclude(pk=subscription.pk).select_related(
                     'formation__filiere', 'formation__centre'
                 ).first()
                 if conflit:
                     messages.error(
                         request,
                         "Cet apprenant a déjà une inscription validée en Formation Initiale "
                         f"({conflit.formation.filiere} - {conflit.formation.centre}) pour cette année de formation."
                     )
                     return redirect("bsb_admin:subscription_en_cours")
             subscription.statut='valide'
             subscription.date_validation=timezone.now()
             subscription.motif_rejet=None
             messages.success(request, "Inscription validée et dettes générées.")
             subscription.save()
     return redirect("bsb_admin:subscription_en_cours")

@require_permission('courses.rejeter_inscription')
def rejeter_inscription(request,id):
    subscription=get_object_or_404(Inscription.objects.select_related('formation__centre','eleve'),id=id)
    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global" and (not subscription.formation_id or not centres_qs.filter(pk=subscription.formation.centre_id).exists()):
        raise PermissionDenied("Vous n'avez pas accès à cette inscription.")

    if request.method == 'POST':
        motif=request.POST.get('motif')
        if not motif:
            messages.error(request,"Veuillez renseigner le motif du rejet du dossier ")
            return redirect("bsb_admin:subscription_en_cours")
        
        subscription.statut = "rejete"
        subscription.motif_rejet = motif
        subscription.date_validation = timezone.now()
        subscription.save()
        messages.warning(request, "Inscription rejetée")
        return redirect("bsb_admin:subscription_en_cours")

    return render(request, "admin/subscription/rejeter_inscription.html", {"subscription": subscription})

#Fonction de récupération des insciptions qui ont non validés 
@require_permission('courses.voir_inscriptions')
def inscription__en_cours_view(request):
    subscriptions=Inscription.objects.filter(statut="en_cours").select_related('eleve','formation').order_by('-date_inscription')
    inscrit_non_valide_qs = Inscription.objects.filter(statut="en_cours")
    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global":
        centre_ids = list(centres_qs.values_list("id", flat=True))
        subscriptions = subscriptions.filter(formation__centre_id__in=centre_ids)
        inscrit_non_valide_qs = inscrit_non_valide_qs.filter(formation__centre_id__in=centre_ids)

    recherche = request.GET.get('recherche', '').strip()
    if recherche:
        subscriptions = subscriptions.filter(
            Q(eleve__nom__icontains=recherche) |
            Q(eleve__prenom__icontains=recherche) |
            Q(eleve__matricule__icontains=recherche)
        )

    paginator=Paginator(subscriptions,10)
    page=request.GET.get('page')
    subscriptions = paginator.get_page(page)
    inscrit_non_valide=inscrit_non_valide_qs.count()
    return render(request,'admin/subscription/validate_inscription.html',{
        'subscriptions':subscriptions,
        'numbers':inscrit_non_valide,
        'recherche': recherche,
    })

# PAYMENT CRUD

@require_permission('courses.gerer_paiements')
def payment_list(request):
    # Une inscription = une ligne, on filtre celles qui ont au moins un paiement
    inscriptions = Inscription.objects.select_related(
        'eleve',
        'formation',
        'formation__centre',
        'formation__filiere',
        'annee_scolaire',
    ).prefetch_related(
        'dettes',
        'dettes__paiements',
        'dettes__frais_formation',
        'dettes__frais_formation__type_frais',
    ).filter(
        dettes__paiements__isnull=False
    ).distinct().order_by('-date_inscription')

    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global":
        centre_ids = list(centres_qs.values_list("id", flat=True))
        inscriptions = inscriptions.filter(formation__centre_id__in=centre_ids)

    q = request.GET.get('q', '').strip()
    if q:
        inscriptions = inscriptions.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__matricule__icontains=q) |
            Q(dettes__paiements__numero_quittance__icontains=q)
        ).distinct()

    paginator = Paginator(inscriptions, 10)
    page = request.GET.get('page')
    inscriptions_page = paginator.get_page(page)

    # Re-calcul après pagination (les objets sont déjà prefetch, pas de requêtes supplémentaires)
    for insc in inscriptions_page:
        insc.total_du   = sum(d.montant_total for d in insc.dettes.all())
        insc.total_paye = sum(
            p.montant_paiement
            for d in insc.dettes.all()
            for p in d.paiements.all()
            if not p.annule
        )
        insc.reste = insc.total_du - insc.total_paye

    return render(request, 'admin/payment/list.html', {'inscriptions': inscriptions_page, 'q': q})

@require_permission('courses.encaisser_paiement')
def payment_create(request):
    if request.method == 'POST':
        form = PaiementAdminForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.cree_par = request.user
            payment.groupe_id = uuid.uuid4()
            payment.save()
            messages.success(request, 'Paiement enregistré avec succès!')
            return redirect('bsb_admin:payment_list')
    else:
        form = PaiementAdminForm()
    return render(request, 'admin/payment/form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.encaisser_paiement')
def payment_update(request, id):
    payment = get_object_or_404(Paiement, id=id)
    if request.method == 'POST':
        form = PaiementAdminForm(request.POST, instance=payment)
        if form.is_valid():
            payment = form.save()
            messages.success(request, 'Paiement modifié avec succès!')
            return redirect('bsb_admin:payment_list')
    else:
        form = PaiementAdminForm(instance=payment)
    return render(request, 'admin/payment/form.html', {'form': form, 'action': 'Modifier', 'object': payment})


@require_permission('courses.gerer_paiements')
def payment_delete(request, id):
    payment = get_object_or_404(Paiement, id=id)
    if request.method == 'POST':
        payment.delete()
        messages.success(request, 'Paiement supprimé avec succès!')
        return redirect('bsb_admin:payment_list')
    return render(request, 'admin/payment/confirm_delete.html', {'object': payment})


# == HISTORIQUE DES CONNEXIONS =================================================
def _date_francaise(valeur):
    """Convertit une date saisie en jj/mm/aaaa.

    Le champ natif <input type="date"> affiche toujours la date selon la langue
    du navigateur, jamais selon celle du site : sur un poste configure en
    anglais, l'agent lisait mm/jj/aaaa. Le champ est donc un champ texte, et
    c'est ici qu'on interprete ce qu'il contient. Le format ISO reste accepte,
    pour que les liens deja partages continuent de fonctionner.
    """
    valeur = (valeur or '').strip()
    if not valeur:
        return None
    for format_essaye in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(valeur, format_essaye).date()
        except ValueError:
            continue
    return None


def _format_francais(valeur):
    """Reaffiche la valeur du filtre en jj/mm/aaaa, meme si l'URL portait de l'ISO."""
    date = _date_francaise(valeur)
    return date.strftime('%d/%m/%Y') if date else ''


def _historique_connexion_filtered(request):
    centres_qs, _, scope = _get_scope(request.user)
    qs = HistoriqueConnexion.objects.select_related('utilisateur', 'centre').order_by('-date_evenement')
    if scope != "global":
        # Les evenements sans centre (eleves, comptes sans role de centre)
        # restent visibles : seuls ceux d'un AUTRE centre sont exclus.
        qs = qs.filter(Q(centre__in=centres_qs) | Q(centre__isnull=True))

    centre_id = request.GET.get('centre', '').strip()
    if centre_id:
        qs = qs.filter(centre_id=centre_id)

    date_debut = _date_francaise(request.GET.get('date_debut', ''))
    if date_debut:
        qs = qs.filter(date_evenement__date__gte=date_debut)

    date_fin = _date_francaise(request.GET.get('date_fin', ''))
    if date_fin:
        qs = qs.filter(date_evenement__date__lte=date_fin)

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(username__icontains=q) | Q(nom_complet__icontains=q))

    type_utilisateur = request.GET.get('type_utilisateur', '').strip()
    if type_utilisateur == 'apprenant':
        qs = qs.filter(est_apprenant=True)
    elif type_utilisateur == 'autre':
        qs = qs.filter(est_apprenant=False)

    type_evenement = request.GET.get('type_evenement', '').strip()
    if type_evenement in ('connexion', 'deconnexion', 'echec'):
        qs = qs.filter(type_evenement=type_evenement)

    return qs, centres_qs


def _historique_connexion_indicateurs(qs):
    """Indicateurs de tete d'ecran, calcules sur le perimetre filtre.

    `comptes_vises` compte les identifiants distincts ayant subi un echec et
    `sources` les adresses IP distinctes en ayant produit : une seule adresse
    visant beaucoup de comptes est le motif caracteristique d'un sondage."""
    # order_by() vide : l'ordre du queryset entrerait sinon dans le GROUP BY et
    # produirait une ligne par date au lieu d'une ligne par type d'evenement.
    par_type = dict(qs.order_by().values_list('type_evenement').annotate(n=Count('id')))
    echecs = qs.filter(type_evenement='echec')
    return {
        'total': qs.count(),
        'connexions': par_type.get('connexion', 0),
        'echecs': par_type.get('echec', 0),
        'comptes_vises': echecs.values('username').distinct().count(),
        'sources': echecs.exclude(adresse_ip__isnull=True)
                         .values('adresse_ip').distinct().count(),
    }


@require_permission('accounts.voir_historique_connexion')
def historique_connexion_list(request):
    qs, centres_qs = _historique_connexion_filtered(request)
    paginator = Paginator(qs, 25)
    page = request.GET.get('page')
    filtres_actifs = any(request.GET.get(c, '').strip() for c in
                         ('q', 'centre', 'date_debut', 'date_fin',
                          'type_utilisateur', 'type_evenement'))
    return render(request, 'admin/historique_connexion/list.html', {
        'historique': paginator.get_page(page),
        'indicateurs': _historique_connexion_indicateurs(qs),
        'centres': centres_qs.order_by('nom_centre'),
        'centre_selectionne': request.GET.get('centre', ''),
        'date_debut': _format_francais(request.GET.get('date_debut', '')),
        'date_fin': _format_francais(request.GET.get('date_fin', '')),
        'type_utilisateur': request.GET.get('type_utilisateur', ''),
        'type_evenement': request.GET.get('type_evenement', ''),
        'q': request.GET.get('q', ''),
        'filtres_actifs': filtres_actifs,
        'onglets': _onglets_audit(request.user, actif='journal'),
    })


def _onglets_audit(utilisateur, actif):
    """Sous-onglets du domaine supervision. Le reglage d'envoi n'apparait
    qu'aux comptes habilites a le modifier."""
    onglets = [{'libelle': 'Journal des connexions',
                'url': reverse('bsb_admin:historique_connexion_list'),
                'actif': actif == 'journal'}]
    if utilisateur.has_perm('audit.gerer_destinataires_audit'):
        onglets.append({'libelle': "Réglage d'envoi",
                        'url': reverse('bsb_admin:destinataire_audit_list'),
                        'actif': actif == 'reglage'})
    return onglets


def _historique_connexion_resume_filtres(request, centres_qs):
    """Phrase récapitulant les filtres actifs — reprise telle quelle dans
    chaque fichier exporté, pour que le contenu du fichier reste traçable
    même une fois détaché de l'écran qui l'a généré."""
    parties = []
    centre_id = request.GET.get('centre', '').strip()
    if centre_id:
        centre = centres_qs.filter(pk=centre_id).first()
        parties.append(f"Centre : {centre.nom_centre if centre else centre_id}")
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin = request.GET.get('date_fin', '').strip()
    if date_debut or date_fin:
        parties.append(f"Période : du {date_debut or '…'} au {date_fin or '…'}")
    q = request.GET.get('q', '').strip()
    if q:
        parties.append(f"Recherche : « {q} »")
    type_utilisateur = request.GET.get('type_utilisateur', '').strip()
    if type_utilisateur == 'apprenant':
        parties.append("Type : Apprenant")
    elif type_utilisateur == 'autre':
        parties.append("Type : Autre")
    return " | ".join(parties) if parties else "Aucun filtre appliqué"


@require_permission('accounts.voir_historique_connexion')
def historique_connexion_export(request, format):
    qs, centres_qs = _historique_connexion_filtered(request)
    resume_filtres = _historique_connexion_resume_filtres(request, centres_qs)
    rows = [{
        'Date connexion/déconnexion': h.date_evenement.strftime('%d/%m/%Y %H:%M'),
        'Utilisateur': h.nom_complet or h.username,
        "Nom d'utilisateur": h.username,
        'Type': 'Apprenant' if h.est_apprenant else 'Autre',
        'Événement': h.get_type_evenement_display(),
        'Centre': h.centre.nom_centre if h.centre else '—',
        'Adresse IP': h.adresse_ip or '—',
    } for h in qs]

    headers = ['Date connexion/déconnexion', 'Utilisateur', "Nom d'utilisateur", 'Type', 'Événement', 'Centre', 'Adresse IP']

    if format not in ('xlsx', 'pdf'):
        raise Http404("Format d'export inconnu.")

    if format == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historique connexions"

        ws.append(["Historique des connexions — BSB"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.append([f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')}  —  Filtres : {resume_filtres}  —  {len(rows)} événement(s)"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws['A2'].font = Font(italic=True, size=9, color="6B7280")
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="C0392B")
        for cell_ in ws[header_row]:
            cell_.font = Font(bold=True, color="FFFFFF")
            cell_.fill = header_fill
            cell_.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(list(row.values()))
        for col in ws.columns:
            first_valid_cell = next((c for c in col if not isinstance(c, MergedCell)), None)
            if first_valid_cell:
                ws.column_dimensions[get_column_letter(first_valid_cell.column)].width = 24

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="historique_connexions.xlsx"'
        return response

    elif format == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm
        )
        styles = getSampleStyleSheet()
        elements = []

        # En-tête officiel (ministère/BSB + logo), identique aux autres exports PDF.
        favicon_path = os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
        header_line_style = ParagraphStyle(
            'header_line', parent=styles['Normal'], fontSize=6, leading=8,
            alignment=1, fontName='Helvetica-Bold',
        )
        header_left, header_right = _pdf_header_lines()
        header_table = Table(
            [[
                Paragraph('<br/>'.join(header_left), header_line_style),
                Image(favicon_path, width=1.6*cm, height=1.6*cm),
                Paragraph('<br/>'.join(header_right), header_line_style),
            ]],
            colWidths=[12*cm, 3*cm, 12*cm],
        )
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 12))

        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, spaceAfter=6, alignment=1)
        sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, spaceAfter=12, alignment=1, textColor=rl_colors.grey)
        elements.append(Paragraph("Historique des connexions — BSB", title_style))
        elements.append(Paragraph(
            f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')} | Filtres : {resume_filtres} | {len(rows)} événement(s)",
            sub_style
        ))

        data = [headers] + [list(row.values()) for row in rows[:1000]]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor("#C0392B")),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(t)

        def _watermark_page(canvas_obj, doc_obj):
            _draw_pdf_watermark(canvas_obj, doc_obj.pagesize[0], doc_obj.pagesize[1], favicon_path)

        doc.build(elements, onFirstPage=_watermark_page, onLaterPages=_watermark_page)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historique_connexions.pdf"'
        return response

    return redirect('bsb_admin:historique_connexion_list')


#CRUD PROGRAMMING

@require_permission('courses.gerer_programmations')
def programming_list(request):
    centres_qs, _, _ = _get_scope(request.user)
    programs=CentreEtFiliere.objects.filter(centre__in=centres_qs).select_related('centre','filiere').order_by('-date_creation')
    f=FormationFilter(request.GET,queryset=programs)
    paginator=Paginator(f.qs,10)
    page=request.GET.get('page')
    programs=paginator.get_page(page)
    return render(request,'admin/programming/list.html',{'programs':programs,'filter':f})

from .models import TypeFrais, PieceJointeInscription  # assure-toi des imports

@require_permission('courses.gerer_programmations')
def programming_create(request):
    if not AnneeScolaire.objects.exists():
        messages.warning(request, "Aucune année de formation disponible. Veuillez en créer une d'abord.")
        return redirect('bsb_admin:annee_create')

    centres_qs, _, _ = _get_scope(request.user)

    if request.method == 'POST':
        form = CentreEtFiliereForm(request.POST, request.FILES, centre_queryset=centres_qs)
        frais_formset = FraisFormSet(request.POST, prefix='frais')
        piece_jointe_formset = PieceJointeFormSet(request.POST, prefix='piece')

        if form.is_valid() and frais_formset.is_valid() and piece_jointe_formset.is_valid():
            programming = form.save()
            frais_formset.instance = programming
            frais_formset.save()
            piece_jointe_formset.instance = programming
            piece_jointe_formset.save()
            messages.success(request, 'Programmation ajoutée avec succès')
            return redirect('bsb_admin:programming_list')
    else:
        form = CentreEtFiliereForm(centre_queryset=centres_qs)
        frais_formset = FraisFormSet(prefix='frais')
        piece_jointe_formset = PieceJointeFormSet(prefix='piece')

    return render(request, 'admin/programming/form.html', {
        'form': form,
        'frais_formset': frais_formset,
        'piece_jointe_formset': piece_jointe_formset,
        'action': 'Créer',
        # ↓ Les deux lignes ajoutées
        'type_frais_options': TypeFrais.objects.all(),
        'type_piece_options': PieceJointeInscription._meta.get_field('type_piece').choices,
    })


@require_permission('courses.gerer_programmations')
def update_pregramming(request, id):
    centres_qs, _, _ = _get_scope(request.user)
    program = get_object_or_404(CentreEtFiliere, id=id, centre__in=centres_qs)

    if request.method == 'POST':
        form = CentreEtFiliereForm(request.POST, request.FILES, instance=program, centre_queryset=centres_qs)
        frais_formset = FraisFormSet(request.POST, instance=program, prefix='frais')
        piece_jointe_formset = PieceJointeFormSet(request.POST, instance=program, prefix='piece')

        if form.is_valid() and frais_formset.is_valid() and piece_jointe_formset.is_valid():
            program = form.save()
            frais_formset.instance = program
            frais_formset.save()
            piece_jointe_formset.instance = program
            piece_jointe_formset.save()
            messages.success(request, 'Modification réussie avec succès')
            return redirect('bsb_admin:programming_list')
    else:
        form = CentreEtFiliereForm(instance=program, centre_queryset=centres_qs)
        frais_formset = FraisFormSet(instance=program, prefix='frais')
        piece_jointe_formset = PieceJointeFormSet(instance=program, prefix='piece')

    return render(request, 'admin/programming/form.html', {
        'action': 'Modifier',
        'form': form,
        'frais_formset': frais_formset,
        'piece_jointe_formset': piece_jointe_formset,
        # ↓ Les deux lignes ajoutées
        'type_frais_options': TypeFrais.objects.all(),
        'type_piece_options': PieceJointeInscription._meta.get_field('type_piece').choices,
    })
@require_permission('courses.gerer_programmations')
def programming_delete(request, id):
    centres_qs, _, _ = _get_scope(request.user)
    program = get_object_or_404(CentreEtFiliere, id=id, centre__in=centres_qs)
    if request.method == 'POST':
        program.delete()
        messages.success(request, 'Programme supprimé avec succès!')
        return redirect('bsb_admin:programming_list')
    return render(request, 'admin/programming/confirm_delete.html', {'object': program})


@require_permission('courses.gerer_programmations')
def programming_import_template(request):
    from .bulk_import_registry import SPEC_PROGRAMMATION
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_PROGRAMMATION)


@require_permission('courses.gerer_programmations')
def programming_import(request):
    from .bulk_import_registry import SPEC_PROGRAMMATION
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_PROGRAMMATION)


@require_permission('courses.gerer_annees')
def annee_create(request):
    if request.method == 'POST':
        form=AnneeScolaireForm(request.POST)
        if form.is_valid():
            annee=form.save()
            messages.success(request, 'Année enregistré avec succès!')
            return redirect('bsb_admin:annee_list')
    else:
        form = AnneeScolaireForm()
    return render(request, 'admin/annee_scolaire/form.html', {'form': form, 'action': 'Créer'})

#def annee_list(request):
#    return render(request,'admin/annee_scolaire/list.html')

ALL_AGENT_TYPES = [
    'formateur',
    'dir',
    'gestionnaire',
    'agent_comptable',
    'caissier',
    'deps',
    'admin',
    'daf',
    'membre',
]

TYPE_LABELS = [
    ('formateur',       'Formateur'),
    ('dir',             'Directeur inter-régional'),
    ('gestionnaire',    'Directeur de centre'),
    ('agent_comptable', 'Agent comptable'),
    ('caissier',        'Caissière / Caissier'),
    ('deps',            'Direction des Études, de la Planification et des Statistiques'),
    ('admin',           'Administrateur'),
    ('daf',             'Directeur Administratif et Financier'),
    ('membre',          "Membre de l'administration"),
]


@require_permission('accounts.gerer_agents')
def agent_list(request):
    q         = request.GET.get('q', '').strip()
    user_type = request.GET.get('type', '').strip()

    qs = Utilisateur.objects.filter(
        user_type__in=ALL_AGENT_TYPES
    ).order_by('nom', 'prenom')

    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global":
        centre_ids = list(centres_qs.values_list("id", flat=True))
        qs = qs.filter(
            Q(membreadministration__structure_id__in=centre_ids) |
            Q(formateur__centre_id__in=centre_ids)
        )

    if q:
        qs = qs.filter(
            Q(nom__icontains=q)      |
            Q(prenom__icontains=q)   |
            Q(email__icontains=q)    |
            Q(username__icontains=q)
        )
    if user_type:
        qs = qs.filter(user_type=user_type)

    paginator = Paginator(qs, 10)
    agents    = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin/rh/agent_list.html', {
        'agents':       agents,
        'q':            q,
        'user_type':    user_type,
        'type_choices': TYPE_LABELS,
    })


@require_permission('accounts.gerer_agents')

def agent_create(request):
    if request.method == 'POST':
        form = AgentForm(request.POST)
        if form.is_valid():

            username = form.cleaned_data.get("username")
            if Utilisateur.objects.filter(username=username).exists():
                messages.error(request, "Ce nom d'utilisateur est déjà utilisé.")
                return render(request, 'admin/rh/agent_form.html', {
                    'form': form,
                    'action': 'Créer',
                    'filiere_modules_json': _filiere_modules_map(),
                })

            agent = form.save()
            messages.success(request, f'Agent « {agent.nom} {agent.prenom} » créé avec succès !')
            return redirect('bsb_admin:agent_list')
    else:
        form = AgentForm()

    return render(request, 'admin/rh/agent_form.html', {
        'form': form,
        'action': 'Créer',
        'filiere_modules_json': _filiere_modules_map(),
    })


@require_permission('accounts.gerer_agents')
def agent_import_template(request):
    from .bulk_import_registry import SPEC_AGENT
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_AGENT)


@require_permission('accounts.gerer_agents')
def agent_import(request):
    from .bulk_import_registry import SPEC_AGENT
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_AGENT)


@require_permission('accounts.gerer_agents')
def agent_update(request, id):
    agent = get_object_or_404(Utilisateur, id=id, user_type__in=ALL_AGENT_TYPES)
    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global" and not _agent_in_scope(agent, centres_qs):
        raise PermissionDenied("Vous n'avez pas accès à cet agent.")
    if request.method == 'POST':
        form = AgentForm(request.POST, instance=agent)
        if form.is_valid():
            agent = form.save()
            messages.success(request, f'Agent « {agent.nom} {agent.prenom} » modifié avec succès !')
            return redirect('bsb_admin:agent_list')
    else:
        form = AgentForm(instance=agent)
    return render(request, 'admin/rh/agent_form.html', {
        'form': form, 'action': 'Modifier',
        'filiere_modules_json': _filiere_modules_map(),
    })


@require_permission('accounts.gerer_agents')
def agent_delete(request, id):
    agent = get_object_or_404(Utilisateur, id=id, user_type__in=ALL_AGENT_TYPES)
    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global" and not _agent_in_scope(agent, centres_qs):
        raise PermissionDenied("Vous n'avez pas accès à cet agent.")
    if request.method == 'POST':
        nom = f"{agent.nom} {agent.prenom}"
        agent.delete()
        messages.success(request, f'Agent « {nom} » supprimé avec succès !')
        return redirect('bsb_admin:agent_list')
    return render(request, 'admin/rh/agent_confirm_delete.html', {'object': agent})


# == GESTION DES APPRENANTS — modification complète (y compris mot de passe), distincte des écrans d'inscription : ici on corrige/complète le dossier d'un élève déjà existant (identité, contact, mot de passe oublié, etc.) ===

@require_permission('accounts.gerer_eleves')
def eleve_list(request):
    recherche = request.GET.get('recherche', '').strip()

    qs = Eleve.objects.all().order_by('nom', 'prenom')
    if recherche:
        qs = qs.filter(
            Q(nom__icontains=recherche) |
            Q(prenom__icontains=recherche) |
            Q(numero_identifiant__icontains=recherche)
        )

    paginator = Paginator(qs, 10)
    eleves = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin/eleve/list.html', {
        'eleves': eleves,
        'recherche': recherche,
    })


@require_permission('accounts.gerer_eleves')
def eleve_update(request, id):
    eleve = get_object_or_404(Eleve, id=id)
    if request.method == 'POST':
        form = EleveForm(request.POST, instance=eleve)
        if form.is_valid():
            eleve = form.save()
            messages.success(request, f'Apprenant « {eleve.nom} {eleve.prenom} » modifié avec succès !')
            return redirect('bsb_admin:eleve_list')
    else:
        form = EleveForm(instance=eleve)
    return render(request, 'admin/eleve/form.html', {'form': form, 'object': eleve})


# == GESTION DES PERMISSIONS — matrice rôle × action ===========================

# Liste curatee : les permissions Django auto-generees (add/change/delete/view)
# n'ont aucun sens pour un utilisateur non technique.
# Quatrieme element : le theme de regroupement. Les themes sont affiches par
# ordre alphabetique, et les permissions par ordre alphabetique de libelle a
# l'interieur de chacun — un classement previsible vaut mieux qu'un ordre
# metier que seul son auteur connait.
MATRIX_PERMISSIONS = [
    ('gerer_regions', "Créer/modifier/supprimer une région ou province", 'courses', 'Découpage territorial'),
    ('gerer_directions', "Créer/modifier/supprimer une direction inter-régionale", 'courses', 'Découpage territorial'),
    ('gerer_centres', "Créer/modifier/supprimer un centre de formation", 'courses', "Configuration de l'offre"),
    ('gerer_metiers', "Créer/modifier/supprimer un métier", 'courses', "Configuration de l'offre"),
    ('gerer_programmations', "Gérer les associations centre-métier", 'courses', "Configuration de l'offre"),
    ('gerer_modules', "Gérer les modules et cours", 'courses', "Configuration de l'offre"),
    ('gerer_frais', "Gérer les frais et types de frais", 'courses', "Configuration de l'offre"),
    ('gerer_annees', "Gérer les années de formation", 'courses', "Configuration de l'offre"),
    ('gerer_equipe', "Gérer le Directeur Général et l'équipe (page « À propos »)", 'courses', 'Site public'),
    ('gerer_agents', "Gérer les comptes utilisateurs", 'accounts', 'Comptes et accès'),
    ('gerer_eleves', "Gérer les comptes apprenants", 'accounts', 'Comptes et accès'),
    ('gerer_permissions', "Gérer les permissions", 'accounts', 'Comptes et accès'),
    ('voir_historique_connexion', "Voir l'historique des connexions", 'accounts', 'Comptes et accès'),
    ('voir_inscriptions', "Voir les candidatures", 'courses', 'Inscriptions'),
    ('valider_inscription', "Valider une candidature", 'courses', 'Inscriptions'),
    ('rejeter_inscription', "Rejeter une candidature", 'courses', 'Inscriptions'),
    ('encaisser_paiement', "Encaisser un paiement", 'courses', 'Paiements de scolarité'),
    ('gerer_paiements', "Modifier/supprimer un paiement", 'courses', 'Paiements de scolarité'),
    ('rechercher_tous_centres', "Rechercher un apprenant dans tous les centres (paiements)", 'courses', 'Paiements de scolarité'),
    ('telecharger_pieces', "Télécharger les pièces jointes des candidats", 'courses', 'Inscriptions'),
    ('voir_statistiques', "Voir les statistiques", 'courses', 'Statistiques et exports'),
    ('gerer_statistiques_reelles', "Gérer le bilan des effectifs formés (saisie manuelle)", 'courses', 'Statistiques et exports'),
    ('exporter_donnees', "Exporter des données (CSV/Excel/PDF)", 'courses', 'Statistiques et exports'),
    ('gerer_facturation', "Créer/gérer les factures de prestation", 'accounts', 'Facturation de prestations'),
    ('valider_facture_prestation', "Valider une facture proforma en définitive", 'accounts', 'Facturation de prestations'),
    ('encaisser_prestation', "Encaisser un paiement de prestation", 'accounts', 'Facturation de prestations'),
]
MATRIX_CODENAMES = [codename for codename, _, _, _ in MATRIX_PERMISSIONS]

# Ordre d'affichage des colonnes (rôle = groupe Django), aligné sur
# accounts.Utilisateur.ROLE_GROUPS.
MATRIX_ROLES = [
    'Admin', 'Directeur Général', 'Directeur Inter-régional', 'DESP',
    'Directeur de Centre', 'Caissier', 'Agent Comptable', 'Formateur',
    'Membre Administration', 'DAF',
]


@require_permission('accounts.gerer_permissions')
def permissions_matrix_view(request):
    from django.contrib.auth.models import Permission, Group

    groups = [Group.objects.get_or_create(name=name)[0] for name in MATRIX_ROLES]

    if request.method == 'POST':
        checked = {}  # group_id -> set(perm_id)
        for pair in request.POST.getlist('cell'):
            perm_id_str, group_id_str = pair.split(':')
            checked.setdefault(int(group_id_str), set()).add(int(perm_id_str))

        matrix_perms = Permission.objects.filter(codename__in=MATRIX_CODENAMES)
        for group in groups:
            perm_ids = checked.get(group.id, set())
            kept = list(group.permissions.exclude(codename__in=MATRIX_CODENAMES))
            newly_checked = list(matrix_perms.filter(id__in=perm_ids))
            group.permissions.set(kept + newly_checked)

        messages.success(request, "Permissions mises à jour avec succès !")
        return redirect('bsb_admin:permissions_matrix')

    permissions = list(
        Permission.objects.filter(codename__in=MATRIX_CODENAMES)
    )
    # garder l'ordre métier défini ci-dessus plutôt que l'ordre alphabétique de la BD
    perm_by_codename = {p.codename: p for p in permissions}
    ordered_permissions = [perm_by_codename[c] for c, _, _, _ in MATRIX_PERMISSIONS if c in perm_by_codename]

    group_perm_ids = {
        group.id: set(group.permissions.filter(codename__in=MATRIX_CODENAMES).values_list('id', flat=True))
        for group in groups
    }

    # Regroupement par theme. Les libelles metier de MATRIX_PERMISSIONS priment
    # sur Permission.name : c'est ce que l'agent lit a l'ecran, et il doit
    # correspondre au vocabulaire du projet.
    libelles = {c: (libelle, theme) for c, libelle, _, theme in MATRIX_PERMISSIONS}
    par_theme = {}
    for perm in ordered_permissions:
        libelle, theme = libelles[perm.codename]
        par_theme.setdefault(theme, []).append({
            'perm': perm,
            'libelle': libelle,
            'cells': [(group, perm.id in group_perm_ids[group.id]) for group in groups],
        })

    # Themes par ordre alphabetique, permissions par ordre alphabetique de
    # libelle : un classement previsible se retrouve sans le connaitre.
    groupes = [
        {
            # Seul le premier theme est deplie : vingt-six permissions sur dix
            # roles remplissaient l'ecran avant meme de commencer a lire.
            'ouvert': i == 0,
            'cle': 'theme-%d' % i,
            'titre': theme,
            'lignes': sorted(lignes, key=lambda l: l['libelle'].lower()),
            'accordees': sum(1 for l in lignes for _, coche in l['cells'] if coche),
        }
        for i, (theme, lignes) in enumerate(sorted(par_theme.items(), key=lambda kv: kv[0].lower()))
    ]

    return render(request, gabarit('admin/permissions/matrix.html'), {
        'roles': groups,
        'groupes': groupes,
        'total_permissions': len(ordered_permissions),
    })


@require_permission('accounts.gerer_agents')
def agent_toggle_active(request, id):
    agent = get_object_or_404(Utilisateur, id=id, user_type__in=ALL_AGENT_TYPES)
    centres_qs, _, scope = _get_scope(request.user)
    if scope != "global" and not _agent_in_scope(agent, centres_qs):
        raise PermissionDenied("Vous n'avez pas accès à cet agent.")
    if request.method == 'POST':
        agent.is_active = not agent.is_active
        agent.save(update_fields=['is_active'])
        etat = "réactivé" if agent.is_active else "suspendu"
        messages.success(request, f'Compte de « {agent.nom} {agent.prenom} » {etat} avec succès !')
    return redirect('bsb_admin:agent_list')



# == TYPE DE FRAIS CRUD ========================================================

@require_permission('courses.gerer_frais')
def type_frais_list(request):
    types = TypeFrais.objects.annotate(nb_tranches=Count('tranches')).order_by('libelle')

    q = request.GET.get('q', '').strip()
    if q:
        types = types.filter(libelle__icontains=q)

    paginator = Paginator(types, 10)
    page = request.GET.get('page')
    types = paginator.get_page(page)
    return render(request, 'admin/fees/type_frais_list.html', {'types': types, 'q': q})


@require_permission('courses.gerer_frais')
def type_frais_create(request):
    if request.method == 'POST':
        form = TypeFraisForm(request.POST)
        if form.is_valid():
            t = form.save()
            messages.success(request, f'Type de frais "{t.libelle}" créé avec succès!')
            return redirect('bsb_admin:type_frais_list')
    else:
        form = TypeFraisForm()
    return render(request, 'admin/fees/type_frais_form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.gerer_frais')
def type_frais_import_template(request):
    from .bulk_import_registry import SPEC_TYPE_FRAIS
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_TYPE_FRAIS)


@require_permission('courses.gerer_frais')
def type_frais_import(request):
    from .bulk_import_registry import SPEC_TYPE_FRAIS
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_TYPE_FRAIS)


@require_permission('courses.gerer_frais')
def type_frais_update(request, id):
    t = get_object_or_404(TypeFrais, id=id)
    if request.method == 'POST':
        form = TypeFraisForm(request.POST, instance=t)
        if form.is_valid():
            t = form.save()
            messages.success(request, f'Type de frais "{t.libelle}" modifié avec succès!')
            return redirect('bsb_admin:type_frais_list')
    else:
        form = TypeFraisForm(instance=t)
    return render(request, 'admin/fees/type_frais_form.html', {'form': form, 'action': 'Modifier', 'object': t})


@require_permission('courses.gerer_frais')
def type_frais_delete(request, id):
    t = get_object_or_404(TypeFrais, id=id)
    if request.method == 'POST':
        libelle = t.libelle
        t.delete()
        messages.success(request, f'Type de frais "{libelle}" supprimé avec succès!')
        return redirect('bsb_admin:type_frais_list')
    return render(request, 'admin/fees/type_frais_confirm_delete.html', {'object': t})


@require_permission('courses.gerer_frais')
def type_frais_tranches(request, id):
    """
    Gère toutes les tranches d'un type de frais sur une seule page (ajout,
    modification, suppression) : la somme des pourcentages doit faire 100%
    et une seule tranche peut être primordiale, donc elles se valident
    ensemble plutôt qu'une par une.
    """
    type_frais = get_object_or_404(TypeFrais, id=id)
    if request.method == 'POST':
        formset = TrancheFraisFormSet(request.POST, instance=type_frais)
        if formset.is_valid():
            formset.save()
            messages.success(request, f'Tranches de "{type_frais.libelle}" mises à jour avec succès!')
            return redirect('bsb_admin:type_frais_list')
    else:
        formset = TrancheFraisFormSet(instance=type_frais)
    return render(request, 'admin/fees/type_frais_tranches.html', {
        'type_frais': type_frais,
        'formset': formset,
    })


# == ANNÉE SCOLAIRE CRUD  (remplace les 2 vues existantes) =====================

@require_permission('courses.gerer_annees')
def annee_list(request):
    annees = AnneeScolaire.objects.all().order_by('-libelle_anne')

    q = request.GET.get('q', '').strip()
    if q:
        annees = annees.filter(libelle_anne__icontains=q)

    paginator = Paginator(annees, 10)
    page = request.GET.get('page')
    annees = paginator.get_page(page)
    return render(request, 'admin/annee/annee_list.html', {'annees': annees, 'q': q})


@require_permission('courses.gerer_annees')
def annee_create(request):
    if request.method == 'POST':
        form = AnneeScolaireForm(request.POST)
        if form.is_valid():
            annee = form.save()
            messages.success(request, f'Année "{annee.libelle_anne}" enregistrée avec succès!')
            return redirect('bsb_admin:annee_list')
    else:
        form = AnneeScolaireForm()
    return render(request, 'admin/annee/annee_scolaire_form.html', {'form': form, 'action': 'Créer'})


@require_permission('courses.gerer_annees')
def annee_import_template(request):
    from .bulk_import_registry import SPEC_ANNEE
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_ANNEE)


@require_permission('courses.gerer_annees')
def annee_import(request):
    from .bulk_import_registry import SPEC_ANNEE
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_ANNEE)


@require_permission('courses.gerer_annees')
def annee_update(request, id):
    annee = get_object_or_404(AnneeScolaire, id=id)
    if request.method == 'POST':
        form = AnneeScolaireForm(request.POST, instance=annee)
        if form.is_valid():
            annee = form.save()
            messages.success(request, f'Année "{annee.libelle_anne}" modifiée avec succès!')
            return redirect('bsb_admin:annee_list')
    else:
        form = AnneeScolaireForm(instance=annee)
    return render(request, 'admin/annee/annee_scolaire_form.html', {'form': form, 'action': 'Modifier', 'object': annee})


@require_permission('courses.gerer_annees')
def annee_delete(request, id):
    annee = get_object_or_404(AnneeScolaire, id=id)
    if request.method == 'POST':
        libelle = annee.libelle_anne
        annee.delete()
        messages.success(request, f'Année "{libelle}" supprimée avec succès!')
        return redirect('bsb_admin:annee_list')
    return render(request, 'admin/annee/annee_confirm_delete.html', {'object': annee})

# == IMPORT DES MEMBRES DE L'EQUIPE ============================================
@require_permission('courses.gerer_equipe')
def membre_import_template(request):
    from .bulk_import_registry import SPEC_MEMBRE_EQUIPE
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_MEMBRE_EQUIPE)


@require_permission('courses.gerer_equipe')
def membre_import(request):
    from .bulk_import_registry import SPEC_MEMBRE_EQUIPE
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_MEMBRE_EQUIPE)
