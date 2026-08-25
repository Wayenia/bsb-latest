from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.html import format_html
from django.utils.decorators import method_decorator
from django.contrib.auth import logout
from django.core.paginator import Paginator
from django.core.files.storage import FileSystemStorage
from django.db.models import Sum
from django.db import transaction
import qrcode
import uuid
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from courses.forms import PersonalInfoForm,PaiementForm
from .models import (CentreEtFiliere, Filiere, Inscription, PieceJointeInscription
    ,DocumentEleve,Paiement,Dette,CentreFormation,AnneeScolaire,Module
    )
from .forms import FiliereForm
from .filters import CentreFormationFilter, FiliereFilter
from .permissions import require_permission, require_role
from django.core.exceptions import PermissionDenied
from accounts.models import Eleve, Utilisateur
from django.urls import reverse
from django.utils import timezone
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Prefetch
import weasyprint
from django.db.models import Sum
import io
import base64
import os
from reportlab.lib.pagesizes import A5
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from django.conf import settings



############### STUDENT LEVEL #############

# SUBSCRIBE SELECTION : annee scolaire -> centre -> metier
def _param_id(request, nom):
    """Retourne l'identifiant s'il est un entier, sinon '' : un paramètre non
    numérique (ex. '../../etc/passwd') ne doit pas atteindre le filtre SQL."""
    valeur = request.GET.get(nom) or ''
    return valeur if valeur.isdigit() else ''


def subscribe_selection_view(request):
    annees = AnneeScolaire.objects.all()
    centres = CentreFormation.objects.all()
    if 'annee' in request.GET:
        # Choix explicite de l'utilisateur (y compris "revenir à ---Sélectionnez---") : respecté tel quel.
        selected_annee_id = _param_id(request, 'annee')
    else:
        # Premier chargement de la page : présélectionner l'année scolaire la plus récente.
        derniere_annee = AnneeScolaire.objects.order_by('-date_creation').first()
        selected_annee_id = str(derniere_annee.id) if derniere_annee else ''
    selected_centre_id = _param_id(request, 'centre')

    careers = []
    if selected_annee_id and selected_centre_id:
        careers = (
            CentreEtFiliere.objects
            .filter(
                is_active=True,
                annee_prog_id=selected_annee_id,
                centre_id=selected_centre_id,
            )
            .filter(Q(date_limite_inscription__isnull=True) | Q(date_limite_inscription__gte=timezone.now()))
            .select_related('filiere')
            .prefetch_related('frais_set')
            .annotate(total_frais=Sum('frais__montant'))
        )

    careers_data = [
        {
            'id': career.id,
            'nom_filiere': career.filiere.nom_filiere,
            'duree': career.duree_display,
            'niveau_diplome': career.filiere.niveau_diplome or '',
            'titre_professionnel': career.filiere.get_titre_professionnel_display() if career.filiere.titre_professionnel else '',
            'total_frais': career.total_frais or 0,
            'communique_url': career.communique.url if career.communique else '',
            'date_limite': career.date_limite_inscription.strftime('%d/%m/%Y à %H:%M') if career.date_limite_inscription else '',
        }
        for career in careers
    ]

    context = {
        'annees': annees,
        'centres': centres,
        'selected_annee_id': selected_annee_id,
        'selected_centre_id': selected_centre_id,
        'careers': careers,
        'careers_data': careers_data,
    }
    return render(request, 'student/subscription/subscribe_selection.html', context)

# AVAILABLE CAREERS
# @login_required
def available_career_view(request):
    # already_subscribed = Inscription.objects.filter(eleve=request.user).values_list('formation', flat=True)
    # available_career = CentreEtFiliere.objects.filter(is_active=True).exclude(id__in=already_subscribed).select_related('centre', 'filiere').order_by('filiere__nom_filiere')
    available_career = (
        CentreEtFiliere.objects.filter(is_active=True)
        .filter(Q(date_limite_inscription__isnull=True) | Q(date_limite_inscription__gte=timezone.now()))
        .prefetch_related('frais_set').annotate(total_frais=Sum('frais__montant'))
        .select_related('centre', 'filiere').order_by('-date_creation')
    )
    #Ici on doit récupéré le id de la formtion lié a fil et centre pour l'affecter le frais   

    f=CentreFormationFilter(request.GET,queryset=available_career)
    paginator=Paginator(f.qs,10)
    page=request.GET.get('page')
    available_career=paginator.get_page(page)

    # Recherche libre d'un métier + téléchargement de son curricula,
    # accessible sans connexion et indépendamment des offres de formation actives.
    curricula_q = request.GET.get('curricula_q', '').strip()
    if curricula_q:
        curricula_results = Filiere.objects.filter(
            is_active=True, nom_filiere__icontains=curricula_q
        ).order_by('nom_filiere')
    else:
        curricula_results = Filiere.objects.none()

    context = {
        'available_career': available_career,
        'filter': f,
        'curricula_q': curricula_q,
        'curricula_results': curricula_results,
    }
    return render(request, 'student/subscription/available_career.html', context)

# GET CAREER BY ID

@login_required
def get_career_by_id(request, id):
    selected_career = get_object_or_404(CentreEtFiliere, id=id)
    context = {'selected_career': selected_career}
    return render(request, 'student/subscription/personal_info.html', context)

# DOCUMENTS
@login_required
def documents_view(request):
    career_id = request.session.get('career_id')
    if not career_id:
        messages.warning(request, "Choisissez d'abord une formation.")
        return redirect('courses:available_career')
    career = get_object_or_404(CentreEtFiliere, id=career_id, is_active=True)
    required_doc = PieceJointeInscription.objects.filter(formation=career)

    if request.method == 'POST':
        fs = FileSystemStorage()
        uploaded_files = {}
        errors = []

        for doc in required_doc:
            if doc.est_requis and doc.libelle_piece not in request.FILES:
                errors.append(f'Le document « {doc.libelle_piece} » est obligatoire.')

        # Seuls JPEG/JPG/PNG/PDF sont acceptés pour les documents envoyés par
        # l'élève : extension ET signature binaire verifiees, pour empecher un
        # fichier .html/.svg renomme (vecteur XSS stocke contre le personnel
        # qui ouvre ces documents via "Visualiser"). Taille max 5 Mo par
        # fichier, alignee sur client_max_body_size (5m) de nginx.
        for doc in required_doc:
            if doc.libelle_piece in request.FILES:
                err = _valider_fichier_upload(request.FILES[doc.libelle_piece])
                if err:
                    errors.append(err)

        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            # Save each to disk, store path in session
            for doc in required_doc:
                if doc.libelle_piece in request.FILES:
                    requested_file = request.FILES[doc.libelle_piece]
                    file_name_saved = fs.save(f'student/pieces/{requested_file.name}', requested_file)
                    uploaded_files[doc.libelle_piece] = {
                        'url':fs.url(file_name_saved),
                        'path':file_name_saved
                    }

            # Save to session
            request.session['career_id'] = career_id
            request.session['uploaded_files'] = uploaded_files
            return redirect('courses:recap')
    context = {'career': career, 'required_doc': required_doc,}
    print(f'=== CAREER : {career} ======')
    return render(request, 'student/subscription/documents.html', context)

# PERSONAL INFO
@login_required
@login_required
def personal_info_view(request, career_id):
    request.session['career_id'] = career_id
    try:
        eleve = request.user.eleve
    except Eleve.DoesNotExist:
        messages.error(request, "Votre profil élève est introuvable.")
        return redirect("courses:home")  # ou une autre page appropriée

    #eleve = request.user.eleve

    # Réinscription après rejet : on mémorise l'inscription rejetée d'origine en session
    # (elle ne survit pas forcément dans l'URL après le POST) et on préremplit le formulaire
    # avec les informations qu'elle contenait.
    from_rejected_param = request.GET.get('from_rejected')
    if from_rejected_param:
        request.session['from_rejected_id'] = from_rejected_param
    from_rejected_id = request.session.get('from_rejected_id')

    rejected_inscription = None
    if from_rejected_id:
        rejected_inscription = Inscription.objects.filter(
            id=from_rejected_id, eleve=eleve, statut='rejete'
        ).first()

    initial = {
        'nom': eleve.nom,
        'prenom': eleve.prenom,
        'sexe': (eleve.sexe or '').upper(),
        'email': eleve.email,
        'tel': eleve.tel or '',
        'date_naissance': eleve.date_naissance or '',
        'lieu_naissance': eleve.lieu_naissance or '',
        'niveau_scolaire': eleve.niveau_scolaire or '',
    }
    if rejected_inscription:
        initial.update({
            'type_personne_contact': rejected_inscription.type_personne_contact or '',
            'nom_personne': rejected_inscription.personne_contact_nom or '',
            'prenom_personne': rejected_inscription.personne_contact_prenom or '',
            'fonction': rejected_inscription.personne_contact_fonction or '',
            'contact': rejected_inscription.personne_contact_tel or '',
            'email_personne': rejected_inscription.personne_contact_email or '',
            'organisation_nom': rejected_inscription.organisation_nom or '',
            'organisation_adresse': rejected_inscription.organisation_adresse or '',
            'organisation_tel': rejected_inscription.organisation_tel or '',
            'organisation_email': rejected_inscription.organisation_email or '',
        })

    form = PersonalInfoForm(request.POST or None, initial=initial)

    if request.method == 'POST' and form.is_valid():
        # Mettre à jour l'élève avec les nouvelles infos
        eleve.nom = form.cleaned_data['nom']
        eleve.prenom = form.cleaned_data['prenom']
        eleve.sexe = form.cleaned_data['sexe'].lower()  # 'M' → 'm'
        email = (form.cleaned_data.get('email') or '').strip()
        if email:
            if Utilisateur.objects.filter(email=email).exclude(pk=eleve.pk).exists():
                messages.error(request, "Cette adresse email est déjà utilisée par un autre compte.")
                return render(request, 'student/subscription/personal_info.html', {'form': form, 'career_id': career_id})
            eleve.email = email
        else:
            eleve.email = None
        eleve.tel = form.cleaned_data.get('tel', '')
        eleve.date_naissance = form.cleaned_data.get('date_naissance')
        eleve.lieu_naissance = form.cleaned_data.get('lieu_naissance', '')
        eleve.niveau_scolaire = form.cleaned_data.get('niveau_scolaire', '')
        eleve.save()

        request.session['student_data'] = {
            'nom': eleve.nom,
            'prenom': eleve.prenom,
            'sexe': eleve.sexe,
            'email': eleve.email,
            'tel': eleve.tel,
            'date_naissance': str(eleve.date_naissance or ''),
            'lieu_naissance': eleve.lieu_naissance,
            'niveau_scolaire': eleve.get_niveau_scolaire_display(),
            'type_personne_contact': form.cleaned_data.get('type_personne_contact', ''),
            'nom_personne': form.cleaned_data.get('nom_personne', ''),
            'prenom_personne': form.cleaned_data.get('prenom_personne', ''),
            'fonction': form.cleaned_data.get('fonction', ''),
            'contact': form.cleaned_data.get('contact', ''),
            'email_personne': form.cleaned_data.get('email_personne', ''),
            'organisation_nom': form.cleaned_data.get('organisation_nom', ''),
            'organisation_adresse': form.cleaned_data.get('organisation_adresse', ''),
            'organisation_tel': form.cleaned_data.get('organisation_tel', ''),
            'organisation_email': form.cleaned_data.get('organisation_email', ''),
        }
        return redirect('courses:documents')

    context = {'form': form, 'career_id': career_id}
    return render(request, 'student/subscription/personal_info.html', context)
# RECAP
@login_required
def recap_view(request):
    career_id = request.session.get('career_id')
    student_data = request.session.get('student_data')
    uploaded_files = request.session.get('uploaded_files', {})
   

    if not career_id or not student_data:
        messages.warning(request, 'Votre session a expiré. Recommencez svp.')
        return redirect('courses:available_career')

    career = get_object_or_404(CentreEtFiliere, id=career_id)

    if request.method == 'POST': # double safety
        if Inscription.objects.filter(
            eleve=request.user.eleve,
            formation=career
            ).exclude(statut='rejete').exists():
            messages.error(request, 'Vous avez déjà déposé une demande d\'inscription pour cette formation.')
            return redirect('courses:my_subscriptions')

        if career.type_formation == 'initiale':
            # Autant de demandes en Formation Initiale que voulu, dans
            # n'importe quel centre, tant qu'aucune n'est encore validée
            # cette année scolaire — seule une inscription déjà VALIDÉE
            # bloque de nouvelles demandes (les demandes en attente
            # n'empêchent pas d'en déposer d'autres en parallèle).
            conflit = Inscription.objects.filter(
                eleve=request.user.eleve,
                annee_scolaire=career.annee_prog,
                formation__type_formation='initiale',
                statut__in=['valide', 'valide_paye'],
            ).exclude(formation=career).select_related(
                'formation__filiere', 'formation__centre'
            ).first()
            if conflit:
                messages.error(
                    request,
                    "Vous avez déjà une inscription validée en Formation Initiale "
                    f"({conflit.formation.filiere} - {conflit.formation.centre}) pour cette année de formation. "
                    "Une nouvelle inscription en Formation Initiale n'est pas autorisée la même année."
                )
                return redirect('courses:my_subscriptions')

        from_rejected_id = request.session.get('from_rejected_id')
        rejected_inscription = None
        if from_rejected_id:
            rejected_inscription = Inscription.objects.filter(
                id=from_rejected_id, eleve=request.user.eleve, statut='rejete'
            ).first()

        # Create inscription
        inscription=Inscription.objects.create(
            eleve=request.user.eleve,
            formation=career,
            statut='en_cours',
            annee_scolaire=career.annee_prog,  # ← récupérée depuis la formation
            type_personne_contact=student_data.get('type_personne_contact', ''),
            personne_contact_nom=student_data.get('nom_personne', ''),
            personne_contact_prenom=student_data.get('prenom_personne', ''),
            personne_contact_fonction=student_data.get('fonction', ''),
            personne_contact_tel=student_data.get('contact', ''),
            personne_contact_email=student_data.get('email_personne', ''),
            organisation_nom=student_data.get('organisation_nom', ''),
            organisation_adresse=student_data.get('organisation_adresse', ''),
            organisation_tel=student_data.get('organisation_tel', ''),
            organisation_email=student_data.get('organisation_email', ''),
            id_inscription_rejeter=rejected_inscription,
        )
        for libelle,fic in uploaded_files.items():
            try:

                docs=PieceJointeInscription.objects.get(
                    formation=career,
                    libelle_piece=libelle
                )
                doc=DocumentEleve(
                    inscription=inscription,
                    piece_requise=docs,
                )
                doc.piece.name=fic['path']
                doc.save()
            except PieceJointeInscription.DoesNotExist:
             continue

        # finally clear session data
        for key in ['career_id', 'student_data', 'uploaded_files', 'from_rejected_id']:
            request.session.pop(key, None)

        messages.success(request, 'Votre dossier a été soumis avec succès !')
        return redirect('courses:my_subscriptions')
    context = {
        'career': career,
        'student': student_data,
        'uploaded_files': uploaded_files,
    }
    return render(request, 'student/subscription/recap.html', context)

#Ici on affiche d'abord les frais à payer avec action
@login_required
def liste_dettes(request,id):
    inscription=get_object_or_404(Inscription,id=id)
    eleve = getattr(request.user, 'eleve', None)
    if inscription.eleve_id != getattr(eleve, 'pk', None) and not (request.user.is_superuser or request.user.has_perm('courses.voir_inscriptions')):
        raise PermissionDenied("Vous ne pouvez consulter que vos propres dettes.")
    dettes=Dette.objects.filter(inscription=inscription).select_related('inscription','frais_formation__type_frais').prefetch_related('frais_formation__type_frais__tranches','paiements')
    total = dettes.count()
    soldees = dettes.filter(etat_dette='soldé').count()
    non_soldees = dettes.filter(etat_dette='non_soldé').count()

    return render(request,'student/subscription/dette.html',context={'inscription':inscription,'dettes':dettes,'total': total,
        'soldees': soldees,
        'non_soldees': non_soldees,
        })

#Page informative affichée à l'apprenant à la place du formulaire de paiement
@login_required
def paiement_info_centre(request):
    return render(request, 'student/paiement/info_centre.html')


#Ici cest si cliques sur payé en fait
@login_required
def effectuer_paiment(request, id):
    dette = get_object_or_404(Dette, id=id)
    eleve = getattr(request.user, 'eleve', None)
    is_self = dette.inscription.eleve_id == getattr(eleve, 'pk', None)
    if not is_self and not (request.user.is_superuser or request.user.has_perm('courses.encaisser_paiement')):
        raise PermissionDenied("Vous ne pouvez régler que vos propres dettes.")

    # Paiement en ligne désactivé pour les apprenants : ils doivent se rendre
    # au centre BSB le plus proche. Le circuit de paiement lui-même n'est pas
    # touché — cette redirection ne s'applique qu'à l'élève réglant sa propre
    # dette (is_self) ; le personnel autorisé (encaisser_paiement) continue
    # d'utiliser cette vue normalement, rien ci-dessous n'est modifié pour eux.
    if is_self:
        return redirect('courses:paiement_info_centre')

    if dette.reste_a_payer() <= 0:
        messages.warning(request, "Cette dette est déjà soldée.")
        return redirect('courses:liste_dettes', id=dette.inscription.id)

    # Ordre de paiement : la tranche primordiale d'une autre dette de la même
    # inscription doit être intégralement réglée avant celle-ci.
    dette_bloquante, tranche_bloquante = dette.inscription.dette_et_tranche_bloquantes()
    if dette_bloquante and dette_bloquante.id != dette.id:
        messages.error(
            request,
            f"Vous devez d'abord régler entièrement {_libelle_blocage(dette_bloquante, tranche_bloquante)} "
            "avant de pouvoir payer ceci."
        )
        return redirect('courses:liste_dettes', id=dette.inscription.id)

    tranche_cible = dette.tranche_a_payer()
    montant_cible = dette.montant_a_payer()
    tranche_suivante = dette.paiements.count() + 1

    if request.method == 'POST':
        form = PaiementForm(request.POST, request.FILES)
        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.dette = dette
            paiement.tranche = tranche_suivante
            paiement.tranche_frais = tranche_cible

            # Le numéro de quittance est généré automatiquement à l'enregistrement
            # (voir Paiement.save()), avec nouvelle tentative en cas de collision.

            # L'élève doit régler exactement le montant dû (tranche ou dette
            # entière) ; les autres rôles (caisse, admin, gestionnaire...)
            # peuvent régler un montant partiel.
            if is_self and paiement.montant_paiement != montant_cible:
                messages.error(request, f"Vous devez régler exactement le montant dû ({montant_cible:,.0f} FCFA).")
                return render(request, 'student/paiement/form.html', {
                    'form': form, 'dette': dette, 'tranche_cible': tranche_cible, 'montant_cible': montant_cible,
                })
            if paiement.montant_paiement <= 0 or paiement.montant_paiement > montant_cible:
                messages.error(request, f"Le montant saisi dépasse le montant dû ({montant_cible:,.0f} FCFA).")
                return render(request, 'student/paiement/form.html', {
                    'form': form, 'dette': dette, 'tranche_cible': tranche_cible, 'montant_cible': montant_cible,
                })

            # Un frais de dossier se règle en une seule fois, intégralement : ni
            # paiement partiel, ni notion de tranche/dérogation applicable.
            est_frais_dossier = dette.frais_formation.type_frais.est_frais_de_dossier
            if est_frais_dossier and paiement.montant_paiement < montant_cible:
                messages.error(
                    request,
                    f"Le frais de dossier « {dette.frais_formation.type_frais} » doit être réglé intégralement "
                    f"en un seul versement (montant dû : {montant_cible:,.0f} FCFA)."
                )
                return render(request, 'student/paiement/form.html', {
                    'form': form, 'dette': dette, 'tranche_cible': tranche_cible, 'montant_cible': montant_cible,
                })

            # Sous-paiement d'une tranche primordiale par un caissier/staff : motif +
            # pièce jointe obligatoires avant de pouvoir valider le paiement.
            if not is_self and not est_frais_dossier and tranche_cible and tranche_cible.est_primordiale and paiement.montant_paiement < montant_cible:
                motif = request.POST.get('motif_derogation', '').strip()
                piece_jointe = request.FILES.get('piece_jointe_derogation')
                if not motif or not piece_jointe:
                    messages.error(
                        request,
                        "Un motif et une pièce jointe justificative sont obligatoires pour valider un "
                        "règlement inférieur au montant dû de la tranche primordiale."
                    )
                    return render(request, 'student/paiement/form.html', {
                        'form': form, 'dette': dette, 'tranche_cible': tranche_cible, 'montant_cible': montant_cible,
                    })
                err = _valider_fichier_upload(piece_jointe)
                if err:
                    messages.error(request, err)
                    return render(request, 'student/paiement/form.html', {
                        'form': form, 'dette': dette, 'tranche_cible': tranche_cible, 'montant_cible': montant_cible,
                    })
                paiement.motif_derogation = motif
                paiement.piece_jointe_derogation = piece_jointe

            paiement.cree_par = request.user
            paiement.groupe_id = uuid.uuid4()
            paiement.save()

            if dette.reste_a_payer() <= 0:
                dette.etat_dette = "soldé"
                dette.save()

            messages.success(request, f'Paiement de {paiement.montant_paiement:,.0f} FCFA enregistré.')
            return redirect('courses:mes_paiements')
    else:
        form = PaiementForm(initial={'tranche': tranche_suivante, 'montant_paiement': montant_cible})

    return render(request, 'student/paiement/form.html', {
        'form': form,
        'dette': dette,
        'tranche_cible': tranche_cible,
        'montant_cible': montant_cible,
    })

#Afficher tous les paiemnents de l'élève conneecté en fait c'est mieux on a pas beoin des dettes on affiche tout
# require_role('eleve') et pas seulement login_required : la vue lit
# request.user.eleve, qui leve AttributeError sur un visiteur anonyme et
# Eleve.DoesNotExist sur un agent — dans les deux cas une erreur 500 au lieu
# d'un refus propre. Les donnees restaient filtrees par eleve (aucune fuite),
# mais l'absence de garde produisait des 500 exploitables en deni de service.
@require_role('eleve')
def liste_paiement(request):
    paiements=Paiement.objects.filter(dette__inscription__eleve=request.user.eleve).select_related('dette','dette__inscription','dette__inscription__eleve').order_by('-date_paiement')
    paginator=Paginator(paiements,10)
    page=request.GET.get('page')
    paiements=paginator.get_page(page)
    return render(request, 'student/paiement/mes_paiements.html', {
        'paiements': paiements,
    })
              
# ─────────────────────────────────────────────
# EN-TÊTE OFFICIEL PARTAGÉ POUR LES PDF GÉNÉRÉS
# ─────────────────────────────────────────────
def _pdf_header_lines(centre=None, direction=None):
    """Retourne (lignes_gauche, lignes_droite) de l'en-tête officiel.

    `direction` (Direction_reg) et `centre` (CentreFormation) sont optionnels :
    si absents, l'en-tête s'arrête à "Direction Générale" (cas d'un rapport
    non circonscrit à une direction/un centre précis).
    """
    left = [
        "MINISTÈRE DE L'ENSEIGNEMENT SECONDAIRE",
        "ET DE LA FORMATION PROFESSIONNELLE ET TECHNIQUE",
        "**********",
        "BURKINA SUUDU BAWDE",
        "**********",
        "DIRECTION GENERALE",
    ]
    resolved_direction = direction or (centre.direction if centre else None)
    if resolved_direction:
        left.append("**********")
        left.append(resolved_direction.nom_direction.upper())
    if centre:
        left.append("**********")
        left.append(centre.nom_centre.upper())

    right = ["BURKINA FASO", "**********", "la patrie ou la mort,", "nous vaincrons"]
    return left, right


_FAVICON_DATA_URI_CACHE = None


def _pdf_logo_data_uri():
    """Logo encodé en base64, à insérer directement dans le HTML des PDF
    weasyprint (`<img src="{{ favicon_data_uri }}">`). Nécessaire car
    `weasyprint.HTML(string=...)` ne peut pas résoudre une URL `{% static %}`
    en environnement conteneurisé (pas de serveur HTTP réellement joignable
    depuis le worker qui génère le PDF) — on évite tout aller-retour réseau
    ou résolution de fichier au moment du rendu."""
    global _FAVICON_DATA_URI_CACHE
    if _FAVICON_DATA_URI_CACHE is None:
        favicon_path = os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
        with open(favicon_path, 'rb') as f:
            encoded = base64.b64encode(f.read()).decode('ascii')
        _FAVICON_DATA_URI_CACHE = f"data:image/png;base64,{encoded}"
    return _FAVICON_DATA_URI_CACHE


def _draw_pdf_watermark(p, width, height, favicon_path=None):
    """Filigrane discret (logo centré, faible opacité) dessiné en premier —
    donc derrière tout le reste — sur un canvas reportlab (A5/A4)."""
    favicon_path = favicon_path or os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
    size = min(width, height) * 0.6
    try:
        p.saveState()
        p.setFillAlpha(0.06)
        p.drawImage(
            ImageReader(favicon_path),
            x=(width - size) / 2, y=(height - size) / 2,
            width=size, height=size,
            preserveAspectRatio=True, mask='auto',
        )
        p.restoreState()
    except Exception:
        pass

# ─────────────────────────────────────────────
# ELEVE — Télécharger la quittance PDF
# ─────────────────────────────────────────────
@require_role('eleve')
def telecharger_quittance(request, id):
    
    paiement = get_object_or_404(
        Paiement.objects.filter(dette__inscription__eleve=request.user.eleve).select_related(
            'dette__inscription__formation',
            'dette__inscription__eleve',
            'dette__inscription__annee_scolaire',
        ),
        id=id
    )

    if paiement.dette.inscription.eleve != request.user.eleve:
         messages.error(request, "Action non autorisée.")
         return redirect('courses:mes_paiements')

    if paiement.annule:
        messages.error(request, "Ce versement a été annulé, sa quittance n'est plus disponible au téléchargement.")
        return redirect('courses:mes_paiements')

    centre = paiement.dette.inscription.formation.centre
    header_left, header_right = _pdf_header_lines(centre)
    html_string = render_to_string('student/paiement/quittance_pdf.html', {
         'paiement': paiement,
         'eleve':paiement.dette.inscription.eleve,
         'annee_scolaire':paiement.dette.inscription.annee_scolaire,
         'dette':paiement.dette,
         'centre':centre,
         'filiere':paiement.dette.inscription.formation.filiere,
         'frais':paiement.dette.frais_formation,
         'type_de_frais':paiement.dette.frais_formation.type_frais.libelle,
         'montant_frais':paiement.dette.frais_formation.montant,
         'inscription':paiement.dette.inscription,
         'quittance_numero':paiement.numero_quittance,
         'header_left': header_left,
         'header_right': header_right,
         'favicon_data_uri': _pdf_logo_data_uri(),
     })
    pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="quittance_{paiement.numero_quittance}.pdf"'
    return response


# ─────────────────────────────────────────────
# RÉCÉPISSÉ DE DEMANDE D'INSCRIPTION (dépôt ou validation, sans paiement)
# ─────────────────────────────────────────────
@login_required
def telecharger_recepisse(request, id):
    inscription = get_object_or_404(
        Inscription.objects.select_related(
            'eleve', 'formation__centre', 'formation__filiere', 'annee_scolaire'
        ),
        id=id
    )

    if inscription.eleve != request.user.eleve:
        messages.error(request, "Action non autorisée.")
        return redirect('courses:my_subscriptions')

    centre = inscription.formation.centre
    header_left, header_right = _pdf_header_lines(centre)

    if inscription.statut == 'valide':
        titre_document = "Récépissé de validation de candidature"
    elif inscription.statut == 'rejete':
        titre_document = "Récépissé de demande d'inscription"
    else:
        titre_document = "Récépissé de dépôt de demande"

    html_string = render_to_string('student/subscription/recepisse_pdf.html', {
        'inscription': inscription,
        'eleve': inscription.eleve,
        'annee_scolaire': inscription.annee_scolaire,
        'centre': centre,
        'filiere': inscription.formation.filiere,
        'numero_dossier': f"DOSS-{inscription.id:06d}",
        'titre_document': titre_document,
        'header_left': header_left,
        'header_right': header_right,
        'favicon_data_uri': _pdf_logo_data_uri(),
    })
    pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="recepisse_{inscription.id}.pdf"'
    return response


# ─────────────────────────────────────────────
# ATTESTATION D'INSCRIPTION (nécessite au moins un paiement)
# ─────────────────────────────────────────────
@login_required
def telecharger_attestation(request, id):
    inscription = get_object_or_404(
        Inscription.objects.select_related(
            'eleve', 'formation__centre', 'formation__filiere', 'annee_scolaire'
        ),
        id=id
    )

    if inscription.eleve != request.user.eleve:
        messages.error(request, "Action non autorisée.")
        return redirect('courses:my_subscriptions')

    a_un_paiement = Paiement.objects.filter(dette__inscription=inscription).exists()
    if not a_un_paiement:
        messages.error(
            request,
            "L'attestation d'inscription est disponible après un premier versement pour cette inscription."
        )
        return redirect('courses:my_subscriptions')

    centre = inscription.formation.centre
    header_left, header_right = _pdf_header_lines(centre)

    # "Directeur du centre" = le membre de l'administration rattaché à ce
    # centre avec le rôle "gestionnaire" (= Directeur de Centre — voir
    # ROLE_GROUPS). À défaut (poste vacant), on retombe sur une formule
    # générique plutôt que de laisser un nom vide sur le document officiel.
    directeur_centre = MembreAdministration.objects.filter(
        structure=centre, user_type='gestionnaire'
    ).first()
    directeur_nom = f"{directeur_centre.prenom} {directeur_centre.nom}" if directeur_centre else "Le Directeur du centre"
    if directeur_centre and directeur_centre.sexe == 'f':
        directeur_civilite = "Mme"
        directeur_titre = "Directrice"
        directeur_titre_article = "La Directrice"
    else:
        # Par défaut (sexe masculin, ou poste vacant/genre inconnu) : formes
        # masculines, cohérent avec le "Le Directeur du centre" générique
        # utilisé plus haut quand le poste est vacant.
        directeur_civilite = "M." if directeur_centre else ""
        directeur_titre = "Directeur"
        directeur_titre_article = "Le Directeur"
    ville = centre.province.chef_lieu if centre.province_id else centre.nom_centre

    html_string = render_to_string('student/subscription/attestation_pdf.html', {
        'inscription': inscription,
        'eleve': inscription.eleve,
        'annee_scolaire': inscription.annee_scolaire,
        'centre': centre,
        'filiere': inscription.formation.filiere,
        'directeur_nom': directeur_nom,
        'directeur_civilite': directeur_civilite,
        'directeur_titre': directeur_titre,
        'directeur_titre_article': directeur_titre_article,
        'ville': ville,
        'header_left': header_left,
        'header_right': header_right,
        'favicon_data_uri': _pdf_logo_data_uri(),
    })
    pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attestation_{inscription.id}.pdf"'
    return response


@require_role('eleve')
def download_quittance(request,id):
    paiement = get_object_or_404(
        Paiement.objects.select_related(
            'dette__inscription__eleve',
            'dette__inscription__formation',
            'dette__inscription__annee_scolaire',
        ),
        id=id,
        dette__inscription__eleve=request.user.eleve
    )

    if paiement.annule:
        messages.error(request, "Ce versement a été annulé, sa quittance n'est plus disponible au téléchargement.")
        return redirect('courses:mes_paiements')

    dette=paiement.dette
    inscription=paiement.dette.inscription
    eleve=inscription.eleve
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A5)
    width, height = A5
    favicon_path = os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
    _draw_pdf_watermark(p, width, height, favicon_path)
    header_left, header_right = _pdf_header_lines(inscription.formation.centre)
    line_h = 0.28*cm
    y_left = height - 0.6*cm
    p.setFont("Helvetica-Bold", 5.5)
    for line in header_left:
        p.drawString(0.6*cm, y_left, line)
        y_left -= line_h
    y_right = height - 0.6*cm
    for line in header_right:
        p.drawRightString(width-0.6*cm, y_right, line)
        y_right -= line_h
    p.drawImage(ImageReader(favicon_path), x=width/2-0.9*cm, y=height-2.2*cm, width=1.8*cm, height=1.8*cm, preserveAspectRatio=True, mask='auto')
    #  TITRE
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height-3.8*cm, "QUITTANCE DE PAIEMENT")
    p.setFont("Helvetica", 9)
    p.drawCentredString(width/2, height-4.4*cm, "Burkina Suudu Bawde")
    # Tampon "ANNULÉE" : le paiement reste conservé pour l'audit (numéro de
    # quittance jamais réutilisé) mais ce document ne doit jamais pouvoir
    # passer pour un reçu valide s'il est réimprimé après annulation.
    if paiement.annule:
        p.saveState()
        p.setFillColor(colors.red)
        p.setFillAlpha(0.35)
        p.setFont("Helvetica-Bold", 34)
        p.translate(width/2, height/2)
        p.rotate(30)
        p.drawCentredString(0, 0, "ANNULÉE")
        p.restoreState()
    # LIGNE SEPARATRICE
    y = height - 5.2*cm
    p.setLineWidth(0.8)
    p.line(1.5*cm, y, width-1.5*cm, y)
    # FONCTION HELPER — bascule automatiquement sur plusieurs lignes si la
    # valeur (ex. nom de centre ou de métier à rallonge) dépasse la largeur
    # disponible entre la colonne valeur et la marge droite.
    def ligne(label, valeur, y_pos):
        p.setFont("Helvetica-Bold", 10)
        p.drawString(1.5*cm, y_pos, label)
        p.setFont("Helvetica", 10)
        valeur = str(valeur)
        max_width = (width - 1.5*cm) - 7*cm
        if p.stringWidth(valeur, "Helvetica", 10) <= max_width:
            p.drawString(7*cm, y_pos, valeur)
            return y_pos - 0.5*cm  # <-- était 0.7*cm (chevauchait le QR code plus bas)
        mots = valeur.split()
        lignes, courante = [], ""
        for mot in mots:
            essai = f"{courante} {mot}".strip()
            if p.stringWidth(essai, "Helvetica", 10) <= max_width:
                courante = essai
            else:
                if courante:
                    lignes.append(courante)
                courante = mot
        if courante:
            lignes.append(courante)
        line_h = 0.42*cm
        for i, texte in enumerate(lignes):
            p.drawString(7*cm, y_pos - i*line_h, texte)
        return y_pos - len(lignes)*line_h - 0.1*cm
    #  INFOS QUITTANCE
    y -= 0.5*cm
    y = ligne("Numéro de quittance :", paiement.numero_quittance, y)
    y = ligne("Date de paiement :", paiement.date_paiement.strftime("%d/%m/%Y à %H:%M"), y)
    y -= 0.3*cm
    p.setLineWidth(0.3)
    p.setDash(3, 3)
    p.line(1.5*cm, y, width-1.5*cm, y)
    p.setDash()
    y -= 0.5*cm
    # INFOS APPRENANT
    y = ligne("Apprenant :", f"{eleve.nom} {eleve.prenom}", y)
    y = ligne("Matricule :", eleve.matricule or "—", y)
    y = ligne("Centre de Formation :", str(inscription.formation.centre), y)
    y=ligne("Métier :" ,str(inscription.formation.filiere),y)
    y = ligne("Année de formation :", str(inscription.annee_scolaire), y)
    y -= 0.3*cm
    p.setDash(3, 3)
    p.line(1.5*cm, y, width-1.5*cm, y)
    p.setDash()
    y -= 0.5*cm
    # ── DETAILS PAIEMENT ───────────────────────────────────
    tranche_label = paiement.tranche_frais.libelle if paiement.tranche_frais else f"Tranche {paiement.tranche}"
    y = ligne("Type de frais :", str(dette.frais_formation.type_frais.libelle), y)
    y = ligne("Tranche :", tranche_label, y)
    y = ligne("Mode de paiement :", paiement.get_mode_paiement_display(), y)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(1.5*cm, y, "Montant payé :")
    p.drawString(7*cm, y, f"{paiement.montant_paiement:,.0f} FCFA")
    y -= 0.7*cm
    y -= 0.3*cm
    p.setDash(3, 3)
    p.line(1.5*cm, y, width-1.5*cm, y)
    p.setDash()
    y -= 0.5*cm
    # ── RECAP DETTE ────────────────────────────────────────
    y = ligne("Total dû :",      f"{dette.montant_total:,.0f} FCFA", y)
    y = ligne("Total payé :",    f"{dette.montant_paye():,.0f} FCFA", y)
    y = ligne("Reste à payer :", f"{dette.reste_a_payer():,.0f} FCFA", y)
    y = ligne("État de la dette :", dette.get_etat_dette_display(), y)

    # ── QR CODE ────────────────────────────────────────────
    qr_data = (
        f"Quittance : {paiement.numero_quittance}\n"
        f"Date : {paiement.date_paiement.strftime('%d/%m/%Y à %H:%M')}\n"
        f"Apprenant : {eleve.nom} {eleve.prenom}\n"
        f"Centre : {inscription.formation.centre}\n"
        f"Métier : {inscription.formation.filiere}\n"
        f"Année de formation : {inscription.annee_scolaire}\n"
        f"Type de frais : {dette.frais_formation.type_frais.libelle}\n"
        f"Tranche : {tranche_label}\n"
        f"Mode de paiement : {paiement.get_mode_paiement_display()}\n"
        f"Montant payé : {paiement.montant_paiement:,.0f} FCFA\n"
        f"Total dû : {dette.montant_total:,.0f} FCFA\n"
        f"Total payé : {dette.montant_paye():,.0f} FCFA\n"
        f"Reste à payer : {dette.reste_a_payer():,.0f} FCFA\n"
        f"État : {dette.get_etat_dette_display()}"
    )
    import qrcode
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)

    # Position calculée à partir de la fin du texte (et non plus une valeur
    # fixe) pour que le QR code ne chevauche jamais les lignes ci-dessus,
    # même si certaines ont débordé sur plusieurs lignes.
    qr_size = 3*cm
    qr_x = (width - qr_size) / 2  # centré horizontalement
    qr_y = max(y - 0.3*cm - qr_size, 0.9*cm)
    p.drawImage(ImageReader(qr_buffer), x=qr_x, y=qr_y, width=qr_size, height=qr_size)
    p.setFont("Helvetica-Oblique", 7)
    p.setFillColor(colors.grey)
    p.drawCentredString(width/2, qr_y - 0.25*cm, "Scannez pour vérifier")

    #  PIED DE PAGE — à droite, en petit, pour ne pas chevaucher le QR code centré
    p.setFont("Helvetica-Oblique", 6)
    p.setFillColor(colors.grey)
    p.drawRightString(width - 1.5*cm, max(qr_y - 0.65*cm, 0.3*cm), f"BSB — généré sur YU-PAAN le : {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    p.showPage()
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="quittance_{paiement.numero_quittance}.pdf"'
    return response
# STUDENT DASHBOARD
@login_required
def student_dashboard(request):
    already_subscribed = Inscription.objects.filter(eleve=request.user.eleve).values_list('formation', flat=True)
    available_career_count = CentreEtFiliere.objects.filter(is_active=True).exclude(id__in=already_subscribed).count()
    active_careers = (
        CentreEtFiliere.objects
        .filter(is_active=True)
        .select_related('centre', 'filiere', 'annee_prog')
        .prefetch_related('frais_set')
        .annotate(total_frais=Sum('frais__montant'))
        .order_by('-date_lancement')
    )
    context = {
        'my_subscriptions': Inscription.objects.filter(eleve=request.user.eleve).count(),
        'available_career_count': available_career_count,
        'active_careers': active_careers,
    }
    return render(request, 'student/dashboard/dashboard.html', context)

# MY SUBSCRIPTIONS
@login_required
def my_subscriptions(request):
    subscriptions = (
        Inscription.objects.filter(eleve=request.user.eleve)
        .select_related('formation__centre', 'formation__filiere')
        .prefetch_related('dettes__paiements')
        .order_by('-date_inscription')
    )
    # Inscriptions rejetées pour lesquelles une réinscription (non re-rejetée) existe déjà :
    # leur bouton "Se réinscrire" doit être désactivé.
    deja_reinscrites_ids = set(
        Inscription.objects.filter(
            eleve=request.user.eleve,
            id_inscription_rejeter__isnull=False,
        ).exclude(statut='rejete').values_list('id_inscription_rejeter_id', flat=True)
    )

    paginator = Paginator(subscriptions, 10)
    page = request.GET.get('page')
    subscriptions = paginator.get_page(page)

    # Attestation téléchargeable seulement si au moins un paiement existe
    # (voir dettes__paiements préchargé ci-dessus — pas de requête par ligne).
    for insc in subscriptions:
        insc.has_payment = any(dette.paiements.all() for dette in insc.dettes.all())

    context = {'subscriptions': subscriptions, 'deja_reinscrites_ids': deja_reinscrites_ids}
    return render(request, 'student/dashboard/my_subscriptions.html', context)

@login_required
def api(request):
    return render(request,'student/paiement/quittance_pdf.html')

###############CENTRE USER PANEL #############


@login_required
@login_required
def member_dashboard(request):
    user = request.user
    utype = user.user_type

    active_careers = (
        CentreEtFiliere.objects
        .filter(is_active=True)
        .select_related('centre', 'filiere', 'annee_prog')
        .prefetch_related('frais_set')
        .annotate(total_frais=Sum('frais__montant'))
        .order_by('-date_lancement')
    )

    # ── deps/membre (personnel du siège, sans centre) → accès global en
    # lecture ; ce que chacun peut FAIRE reste gouverné par ses permissions ─
    if utype in ['deps', 'admin', 'dg', 'membre'] or user.is_superuser:
        stats = {
            'total_inscriptions': Inscription.objects.count(),
            'filieres': Filiere.objects.distinct().count(),
            'etudiants': Eleve.objects.distinct().count(),
        }
        return render(request, 'member/member_dashboard/dashboard.html', {
            'active_careers': active_careers,
            'stats': stats,
            'membre': None,
            'centre': None,
            'centres_visibles': CentreFormation.objects.all(),
            'centres_visibles_ids': list(CentreFormation.objects.values_list('id', flat=True)),
        })

    # ── dir → toute sa direction (tous ses centres), pas de membreadministration
    if utype == 'dir':
        try:
            dir_obj = DirecteurInterRegional.objects.get(pk=user.pk)
        except DirecteurInterRegional.DoesNotExist:
            messages.error(request, "Accès refusé.")
            return redirect('accounts:login')

        direction = dir_obj.direction
        if not direction:
            messages.error(request, "Accès refusé : aucune direction associée à votre profil.")
            return redirect('accounts:login')

        centres_visibles = CentreFormation.objects.filter(direction=direction)
        stats = {
            'total_inscriptions': Inscription.objects.filter(formation__centre__direction=direction).count(),
            'filieres': Filiere.objects.filter(centreetfiliere__centre__direction=direction).distinct().count(),
            'etudiants': Eleve.objects.filter(inscription__formation__centre__direction=direction).distinct().count(),
        }
        return render(request, 'member/member_dashboard/dashboard.html', {
            'active_careers': active_careers.filter(centre__direction=direction),
            'stats': stats,
            'membre': None,
            'centre': None,
            'direction': direction,
            'centres_visibles': centres_visibles,
            'centres_visibles_ids': list(centres_visibles.values_list('id', flat=True)),
        })

    # ── tous les autres → ont forcément un membreadministration ───────────
    try:
        membre = request.user.membreadministration
    except Exception:
        messages.error(request, "Accès refusé.")
        return redirect('accounts:login')

    if not membre.structure:
        messages.error(request, "Accès refusé : vous n'êtes pas membre d'un centre.")
        return redirect('accounts:login')

    structure = membre.structure
    stats = {
        'total_inscriptions': Inscription.objects.filter(formation__centre=structure).count(),
        'filieres': Filiere.objects.filter(centreetfiliere__centre=structure).distinct().count(),
        'etudiants': Eleve.objects.filter(inscription__formation__centre=structure).distinct().count(),
    }
    centres_visibles = membre.get_centres_visibles()
    centres_visibles_ids = [c.id for c in centres_visibles]

    return render(request, 'member/member_dashboard/dashboard.html', {
        'centre': structure,
        'membre': membre,
        'stats': stats,
        'centres_visibles': centres_visibles,
        'centres_visibles_ids': centres_visibles_ids,
        'active_careers': active_careers,
    })


#La liste de tous les inscriptions du cenrte
@require_permission('courses.voir_inscriptions')
def member_inscriptions_list(request):
    #centre=get_object_or_404(CentreFormation,id=id)
    from django.db.models import Q
    member=request.user
    membre = getattr(member, 'membreadministration', None)
    centre = membre.structure if membre else None
    inscriptions=Inscription.objects.select_related('eleve','formation__filiere').filter(formation__centre=centre)

    recherche = request.GET.get('recherche', '').strip()
    if recherche:
        inscriptions = inscriptions.filter(
            Q(eleve__nom__icontains=recherche) |
            Q(eleve__prenom__icontains=recherche) |
            Q(eleve__matricule__icontains=recherche)
        )

    paginator=Paginator(inscriptions,10)
    page=request.GET.get('page')
    inscriptions=paginator.get_page(page)

    context={
        'subscriptions':inscriptions,
        'centre':centre,
        'recherche': recherche,
    }
    return render(request,'member/inscriptions/list.html',context)


##La list des isncriptions en cours du centre
@require_permission('courses.voir_inscriptions')
def member_inscription_en_cours(request):
    #centre=get_object_or_404(CentreFormation,id=id)
    from django.db.models import Q
    member=request.user
    membre = getattr(member, 'membreadministration', None)
    centre = membre.structure if membre else None
    inscriptions=Inscription.objects.select_related('eleve','formation__filiere').filter(formation__centre=centre,statut='en_cours').order_by('-date_inscription')

    recherche = request.GET.get('recherche', '').strip()
    if recherche:
        inscriptions = inscriptions.filter(
            Q(eleve__nom__icontains=recherche) |
            Q(eleve__prenom__icontains=recherche) |
            Q(eleve__matricule__icontains=recherche)
        )

    paginator=Paginator(inscriptions,10)
    page=request.GET.get('page')
    inscriptions=paginator.get_page(page)

    non_valide=Inscription.objects.filter(formation__centre=centre,statut='en_cours').count()

    context={
        'subscriptions':inscriptions,
        'number':non_valide,
        'recherche': recherche,
    }
    return render(request,'member/inscriptions/valide_inscription.html',context)


##Valider une inscription du dentre en cours en fait Ici c'est bo en fait on a pas besoin de
@require_permission('courses.voir_inscriptions')
def member_inscription_detail(request,id):
    membre = getattr(request.user, 'membreadministration', None)
    centre = membre.structure if membre else None
    inscription=get_object_or_404(Inscription,id=id,formation__centre=centre)
    eleve_docs=DocumentEleve.objects.select_related('piece_requise').filter(inscription=inscription)
    return render(request,'member/inscriptions/inscription_detail.html',{'detail':inscription,'eleve_docs':eleve_docs})

@require_permission('courses.valider_inscription')
def gerer_inscription(request,id):
     membre = getattr(request.user, 'membreadministration', None)
     centre = membre.structure if membre else None
     subscription=get_object_or_404(Inscription,id=id,formation__centre=centre)
     if request.method == 'POST':
         action=request.POST.get('action')
         if action == 'valide':
             if subscription.formation and subscription.formation.type_formation == 'initiale':
                 conflit = Inscription.objects.filter(
                     eleve=subscription.eleve,
                     annee_scolaire=subscription.annee_scolaire,
                     formation__type_formation='initiale',
                     statut__in=['valide', 'valide_paye'],
                 ).exclude(pk=subscription.pk).select_related(
                     'formation__filiere', 'formation__centre'
                 ).first()
                 if conflit:
                     messages.error(
                         request,
                         "Cet apprenant a déjà une inscription validée en Formation Initiale "
                         f"({conflit.formation.filiere} - {conflit.formation.centre}) pour cette année de formation."
                     )
                     return redirect("courses:valide_inscription")
             subscription.statut='valide'
             subscription.date_validation=timezone.now()
             subscription.motif_rejet=None
             messages.success(request, "Inscription validée et dettes générées.")
             subscription.save()
     return redirect("courses:valide_inscription")

@require_permission('courses.rejeter_inscription')
def rejeter_inscription(request,id):
    membre = getattr(request.user, 'membreadministration', None)
    centre = membre.structure if membre else None
    subscription=get_object_or_404(Inscription,id=id,formation__centre=centre)

    if request.method == 'POST':
        motif=request.POST.get('motif')
        if not motif:
            messages.error(request,"Veuillez renseigner le motif du rejet du dossier ")
            return redirect("courses:valide_inscription")

        subscription.statut = "rejete"
        subscription.motif_rejet = motif
        subscription.date_validation = timezone.now()
        subscription.save()
        messages.warning(request, "Inscription rejetée")
        return redirect("courses:valide_inscription")

    return render(request, "member/inscriptions/rejeter_inscription.html", {"subscription": subscription})

@require_permission('courses.encaisser_paiement', 'courses.gerer_paiements')
def paiement_list(request):
    """
    Par défaut : inscriptions dont le reste à payer est strictement positif,
    dans le périmètre de l'utilisateur connecté (son centre pour un
    gestionnaire/caissier, les centres de sa direction pour un directeur
    inter-régional, tous les centres pour les rôles à portée nationale —
    voir _get_scope). Avec une recherche (nom, prénom, identifiant,
    téléphone), la recherche porte sur tous les centres si l'utilisateur a
    une portée globale ou la permission 'rechercher_tous_centres', pour
    pouvoir encaisser le paiement de n'importe quel apprenant.

    Note : la vue reposait auparavant sur `request.user.membreadministration`
    seul, ce qui ne couvre que les gestionnaires/caissiers. Un directeur
    inter-régional (modèle DirecteurInterRegional, distinct de
    MembreAdministration) ou un rôle à portée nationale (admin, dg, deps,
    agent_comptable — pas forcément rattachés à un centre) tombait alors sur
    `centre = None`, ce qui filtrait sur `formation__centre=None` et
    renvoyait toujours une liste vide, quelles que soient les permissions
    accordées. On réutilise donc `_get_scope`, déjà utilisé pour les
    statistiques, qui gère correctement ces cas.
    """
    q = request.GET.get('q', '').strip()
    centre_id = request.GET.get('centre', '').strip()
    centres_qs, _directions_qs, scope = _get_scope(request.user)
    multi_centre = scope in ('global', 'direction')
    peut_rechercher_tous_centres = (
        request.user.is_superuser or
        scope == 'global' or
        request.user.has_perm('courses.rechercher_tous_centres')
    )
    if q and not peut_rechercher_tous_centres:
        q = ''
    if not multi_centre:
        centre_id = ''

    base_qs = Inscription.objects.select_related(
        'eleve',
        'formation__filiere',
        'formation__centre',
        'annee_scolaire',
    ).prefetch_related(
        'dettes',
        'dettes__paiements',
        'dettes__frais_formation',
        'dettes__frais_formation__type_frais',
    )

    if q:
        inscriptions_qs = base_qs.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__numero_identifiant__icontains=q) |
            Q(eleve__tel__icontains=q) |
            Q(eleve__matricule__icontains=q) |
            Q(dettes__paiements__numero_quittance__icontains=q)
        ).distinct().order_by('-date_inscription')
    elif scope == 'none':
        inscriptions_qs = base_qs.none()
    else:
        inscriptions_qs = base_qs.filter(
            formation__centre_id__in=centres_qs.values_list('id', flat=True)
        ).order_by('-date_inscription')

    # Le reste à payer est calculé à partir des dettes/paiements déjà préchargés :
    # filtrer en Python plutôt qu'avec une agrégation SQL fragile sur deux
    # niveaux de relations (dettes -> paiements). Les paiements annulés ne
    # comptent plus comme payés (cf. Dette.montant_paye()).
    inscriptions_avec_reste = []
    for insc in inscriptions_qs:
        insc.total_du = sum(d.montant_total for d in insc.dettes.all())
        insc.total_paye = sum(
            p.montant_paiement
            for d in insc.dettes.all()
            for p in d.paiements.all()
            if not p.annule
        )
        insc.reste = insc.total_du - insc.total_paye
        # En parcours libre (sans recherche), on ne montre que ce qui reste dû —
        # utile pour un caissier. Mais une recherche doit retrouver un apprenant
        # même à reste nul (ex. annuler un versement après coup sur un dossier
        # déjà entièrement réglé) : ce n'est prévu nulle part ailleurs.
        if q or insc.reste > 0:
            inscriptions_avec_reste.append(insc)

    scope_labels = {
        'global': "tous les centres",
        'direction': "les centres de votre direction",
        'centre': "votre centre",
        'none': "aucun centre (aucune structure ne vous est rattachée)",
    }

    # Recherche cross-centre ou portée mono-centre : liste plate, comme
    # avant (une recherche cible un apprenant précis, peu importe son centre ;
    # un gestionnaire/caissier n'a qu'un seul centre, pas besoin d'accordéon).
    if q or not multi_centre:
        paginator = Paginator(inscriptions_avec_reste, 10)
        inscriptions = paginator.get_page(request.GET.get('page'))
        return render(request, 'member/paiement/list.html', {
            'mode': 'plat',
            'inscriptions': inscriptions,
            'q': q,
            'scope': scope,
            'scope_label': scope_labels.get(scope, "votre centre"),
            'multi_centre': multi_centre,
            'peut_rechercher_tous_centres': peut_rechercher_tous_centres,
        })

    # Portée multi-centres, sans recherche : accordéon centres -> inscriptions
    # (même principe que régions -> provinces) : la liste des centres est
    # paginée (10/page) et, pour le centre déplié (paramètre ?centre=), ses
    # inscriptions à reste à payer sont elles-mêmes paginées (10/page,
    # paramètre ?ipage=) — sans recharger toute la liste des centres.
    compte_par_centre = {}
    inscriptions_par_centre = {}
    for insc in inscriptions_avec_reste:
        cid = insc.formation.centre_id
        compte_par_centre[cid] = compte_par_centre.get(cid, 0) + 1
        inscriptions_par_centre.setdefault(cid, []).append(insc)

    centres_annotes = list(centres_qs.order_by('nom_centre'))
    for c in centres_annotes:
        c.nb_inscriptions = compte_par_centre.get(c.id, 0)

    paginator = Paginator(centres_annotes, 10)
    centres_page = paginator.get_page(request.GET.get('page'))

    centre_ouvert = None
    inscriptions = None
    if centre_id:
        centre_ouvert = centres_qs.filter(pk=centre_id).first()
        if centre_ouvert:
            ipaginator = Paginator(inscriptions_par_centre.get(centre_ouvert.id, []), 10)
            inscriptions = ipaginator.get_page(request.GET.get('ipage'))

    return render(request, 'member/paiement/list.html', {
        'mode': 'accordeon',
        'centres_page': centres_page,
        'centre_ouvert': centre_ouvert,
        'inscriptions': inscriptions,
        'q': q,
        'scope': scope,
        'scope_label': scope_labels.get(scope, "votre centre"),
        'multi_centre': multi_centre,
        'peut_rechercher_tous_centres': peut_rechercher_tous_centres,
    })


@require_permission('courses.encaisser_paiement', 'courses.gerer_paiements')
def paiement_historique(request):
    """
    Historique des paiements, dans le périmètre de l'utilisateur connecté
    (même logique de portée que paiement_list — voir _get_scope).
    """
    centres_qs, _directions_qs, scope = _get_scope(request.user)
    multi_centre = scope in ('global', 'direction')

    paiements = Paiement.objects.select_related(
        'dette__inscription__eleve',
        'dette__inscription__formation__filiere',
        'dette__inscription__formation__centre',
        'dette__frais_formation__type_frais',
        'cree_par',
    ).order_by('-date_paiement')

    if scope == 'none':
        paiements = paiements.none()
    else:
        paiements = paiements.filter(
            dette__inscription__formation__centre_id__in=centres_qs.values_list('id', flat=True)
        )

    q = request.GET.get('q', '').strip()
    if q:
        paiements = paiements.filter(
            Q(dette__inscription__eleve__nom__icontains=q) |
            Q(dette__inscription__eleve__prenom__icontains=q) |
            Q(dette__inscription__eleve__matricule__icontains=q) |
            Q(numero_quittance__icontains=q)
        )

    paginator = Paginator(paiements, 10)
    page = request.GET.get('page')
    paiements = paginator.get_page(page)

    scope_labels = {
        'global': "tous les centres",
        'direction': "les centres de votre direction",
        'centre': "votre centre",
        'none': "aucun centre (aucune structure ne vous est rattachée)",
    }

    return render(request, 'member/paiement/historique.html', {
        'paiements': paiements,
        'scope': scope,
        'scope_label': scope_labels.get(scope, "votre centre"),
        'multi_centre': multi_centre,
        'q': q,
    })

############### TEACHER LEVEL #############

# TEACHER DASHBOARD
@login_required
@login_required
def teacher_dashboard(request):
    return redirect('courses:formateur_dashboard')

@login_required
def member_dashboard_direction(request, id):
    active_careers = (
        CentreEtFiliere.objects
        .filter(is_active=True)
        .select_related('centre', 'filiere', 'annee_prog')
        .prefetch_related('frais_set')
        .annotate(total_frais=Sum('frais__montant'))
        .order_by('-date_lancement')
    )
    context = {
        'active_careers': active_careers,
    }
    return render(request, "teacher/dashboard/dashboard.html", context)




# ─────────────────────────────────────────────────────────────────────────────
#  CRÉER UN MÉTIER  (anciennement create_fees)
# ─────────────────────────────────────────────────────────────────────────────
def _process_metier_modules(request, filiere, form):
    """
    Associe à ce métier les modules existants cochés dans le formulaire, plus
    les nouveaux modules créés à la volée (champs new_module_nom[]/new_module_volume[]).
    """
    module_ids = list(form.cleaned_data.get('modules_existants').values_list('id', flat=True))
    noms = request.POST.getlist('new_module_nom')
    volumes = request.POST.getlist('new_module_volume')
    for nom, volume in zip(noms, volumes):
        nom = nom.strip()
        if not nom:
            continue
        try:
            vol = int(volume)
        except (TypeError, ValueError):
            vol = 0
        module = Module.objects.create(nom_module=nom, volume_h_cours=vol)
        module_ids.append(module.id)
    filiere.modules.set(module_ids)


@require_permission('courses.gerer_metiers')
def create_metier(request):
    """Crée un nouveau métier (Filiere)."""
    if request.method == 'POST':
        form = FiliereForm(request.POST, request.FILES)
        if form.is_valid():
            metier = form.save()
            _process_metier_modules(request, metier, form)
            messages.success(request, f'Métier « {metier.nom_filiere} » créé avec succès !')
            return redirect('courses:member_field_list')
    else:
        form = FiliereForm()

    return render(request, 'member/filiere/metier_form.html', {
        'form': form,
        'action': 'Créer',
    })


@require_permission('courses.gerer_metiers')
def field_import_template(request):
    from .bulk_import_registry import SPEC_FILIERE
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_FILIERE)


@require_permission('courses.gerer_metiers')
def field_import(request):
    from .bulk_import_registry import SPEC_FILIERE
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_FILIERE)


# ─────────────────────────────────────────────────────────────────────────────
#  LISTE DES MÉTIERS D'UN CENTRE  (anciennement member_filiere_list)
# ─────────────────────────────────────────────────────────────────────────────
@require_permission('courses.gerer_metiers')
def member_metier_list(request):
    from django.db.models import Count
    metiers = Filiere.objects.annotate(nb_modules=Count('modules', distinct=True)).prefetch_related('modules').order_by('nom_filiere')
    f = FiliereFilter(request.GET, queryset=metiers)

    paginator = Paginator(f.qs, 10)
    metiers_page = paginator.get_page(request.GET.get('page'))

    context = {
        'filieres': metiers_page,
        'filter': f,
    }
    return render(request, 'member/filiere/metier_list.html', context)
        

@require_permission('courses.gerer_metiers')
def metier_delete(request, id):
    field = get_object_or_404(Filiere, id=id)
    if request.method == 'POST':
        nom = field.nom_filiere
        field.delete()
        messages.success(request, f'Métier "{nom}" supprimé avec succès !')
        return redirect('bsb_admin:field_list')
    return render(request, 'member/filiere/confirm_delete.html', {'object': field})
############### THIRD PAGES LEVEL #############

#  HOME
def home(request):
    from .models import CentreEtFiliere
    careers_qs = (
        CentreEtFiliere.objects
        .filter(is_active=True)
        .filter(Q(date_limite_inscription__isnull=True) | Q(date_limite_inscription__gte=timezone.now()))
        .select_related('centre', 'filiere', 'annee_prog')
        .order_by('-date_lancement')
    )
    # Un carrousel représente un métier, mais un même métier peut être
    # programmé sur plusieurs centres/années (plusieurs CentreEtFiliere) :
    # on garde une seule carte par métier (la plus récemment lancée), mais on
    # calcule quand même, sur l'ENSEMBLE de ses formations actives, la date
    # limite d'inscription qui expire le plus tôt.
    seen_filieres = set()
    active_careers = []
    earliest_deadlines = {}
    for career in careers_qs:
        if career.date_limite_inscription:
            current = earliest_deadlines.get(career.filiere_id)
            if current is None or career.date_limite_inscription < current:
                earliest_deadlines[career.filiere_id] = career.date_limite_inscription
        if career.filiere_id in seen_filieres:
            continue
        seen_filieres.add(career.filiere_id)
        active_careers.append(career)

    for career in active_careers:
        career.date_limite_proche = earliest_deadlines.get(career.filiere_id)

    return render(request, "third_pages/home.html", {'active_careers': active_careers})

# ABOUT
def about_view(request):
    from .models import DG, Membre
    context = {
        'dg': DG.objects.filter(is_active=True).first(),
        'members': Membre.objects.filter(is_active=True).order_by('order'),
        }
    return render(request, 'third_pages/about.html', context)



############### HELPER METHOD #############

def _bounce_to_login(request, error_message):
    """
    Déconnecte l'utilisateur avant de le renvoyer vers la connexion.
    Indispensable ici : `accounts:login` renvoie tout utilisateur déjà
    authentifié directement vers `redirect_to_dashboard` — sans cette
    déconnexion préalable, un profil incomplet/non reconnu créerait une
    boucle de redirection infinie (login → dashboard → login → ...).
    """
    logout(request)
    messages.error(request, error_message)
    return redirect('accounts:login')


@login_required
def redirect_to_dashboard(request):
    user = request.user

    # Superuser → accès total
    if user.is_superuser:
        return redirect('bsb_admin:admin_dashboard')

    utype = user.user_type

    # ── Admin / Directeur Général → même tableau de bord ───────────────────
    if utype in ('admin', 'dg'):
        return redirect('bsb_admin:admin_dashboard')

    # ── Deps, agent_comptable et membre (personnel du siège, sans centre ni
    # direction de rattachement) → member_dashboard (accès global en lecture,
    # ce que chacun peut y FAIRE reste gouverné par ses permissions) ────────
    if utype in ['deps', 'agent_comptable', 'membre']:
        return redirect('courses:member_dashboard')

    # ── Directeur inter-régional → member_dashboard (vue direction entière) ─
    if utype == 'dir':
        try:
            dir_obj = DirecteurInterRegional.objects.get(pk=user.pk)
            if dir_obj.direction:
                return redirect('courses:member_dashboard')
            return _bounce_to_login(request, "Aucune direction associée à votre profil. Contactez l'administrateur.")
        except DirecteurInterRegional.DoesNotExist:
            return _bounce_to_login(request, "Profil directeur introuvable. Contactez l'administrateur.")

    # ── Formateur ──────────────────────────────────────────────────────────
    if utype == 'formateur':
        try:
            _ = user.formateur
            return redirect('courses:formateur_dashboard')
        except Exception:
            return _bounce_to_login(request, "Profil formateur introuvable. Contactez l'administrateur.")

    # ── Gestionnaire / Caissier (ont toujours une structure) ───────────────
    if utype in ['gestionnaire', 'caissier']:
        try:
            membre = user.membreadministration
            if membre.structure is not None:
                return redirect('courses:member_dashboard')
            elif membre.direction is not None:
                return redirect('courses:member_dashboard_direction', id=membre.direction.id)
            return _bounce_to_login(request, "Aucune structure ni direction associée. Contactez l'administrateur.")
        except Exception:
            return _bounce_to_login(request, "Profil membre introuvable. Contactez l'administrateur.")

    # ── Élève ──────────────────────────────────────────────────────────────
    if utype == 'eleve':
        try:
            _ = user.eleve
            return redirect('courses:student_dashboard')
        except Exception:
            return _bounce_to_login(request, "Profil élève introuvable. Contactez l'administrateur.")

    # ── DAF (Directeur Administratif et Financier) — module Facturation ────
    if utype == 'daf':
        return redirect('accounts:daf_dashboard')

    # ── Fallback ───────────────────────────────────────────────────────────
    return _bounce_to_login(request, f"Type d'utilisateur non reconnu : {utype}. Contactez l'administrateur.")

#statistiques
############### STATISTIQUES & PAIEMENT MEMBRE ###############

"""
courses/views/statistiques_view.py
Logique d'accès par rôle :
  - admin | dg | agent_comptable | deps  → tout voir
  - dir                                  → sa direction + ses centres
  - gestionnaire | caissier              → son centre uniquement
"""

from django.db.models import Count, Sum, Q
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from .models import (
    CentreFormation, Direction_reg, Filiere, AnneeScolaire,
    Inscription, Dette, Paiement, CentreEtFiliere, Frais,
)
from accounts.models import Eleve, DirecteurInterRegional, MembreAdministration

import csv, io, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ─── Helpers ──────────────────────────────────────────────────────────────────

ROUGE  = "#C0392B"
OR     = "#D4A017"
GRIS   = "#6B7280"

def _get_scope(user):
    """
    Retourne (centres_qs, directions_qs, scope_label)
    selon le rôle de l'utilisateur connecté.
    """
    # Un superuser créé via `createsuperuser` ne demande pas user_type
    # et laisse donc la valeur par défaut "eleve" il faut donc le capter ici
    if user.is_superuser:
        return (
            CentreFormation.objects.all(),
            Direction_reg.objects.all(),
            "global",
        )

    utype = user.user_type

    # Niveau 1 : accès total
    if utype in ("admin", "dg", "agent_comptable", "deps"):
        return (
            CentreFormation.objects.all(),
            Direction_reg.objects.all(),
            "global",
        )

    # Directeur Inter-régional : toujours limité à sa propre direction
    # (un directeur ayant besoin d'une portée globale doit être promu vers un
    # rôle qui l'a par défaut — admin/dg/deps — plutôt que via un booléen caché)
    if utype == "dir":
        try:
            dir_obj = DirecteurInterRegional.objects.get(pk=user.pk)
            centres = CentreFormation.objects.filter(direction=dir_obj.direction)
            directions = Direction_reg.objects.filter(pk=dir_obj.direction_id)
            return centres, directions, "direction"
        except DirecteurInterRegional.DoesNotExist:
            pass

    # Gestionnaire / Caissier → son centre uniquement
    if utype in ("gestionnaire", "caissier"):
        try:
            membre = MembreAdministration.objects.get(pk=user.pk)
            if membre.structure:
                centres = CentreFormation.objects.filter(pk=membre.structure_id)
                directions = Direction_reg.objects.filter(
                    pk=membre.structure.direction_id
                ) if membre.structure.direction_id else Direction_reg.objects.none()
                return centres, directions, "centre"
        except MembreAdministration.DoesNotExist:
            pass

    # Fallback : rien
    return CentreFormation.objects.none(), Direction_reg.objects.none(), "none"


def _base_qs(user):
    """
    Retourne les querysets de base filtrés selon le scope de l'utilisateur.
    """
    centres_qs, directions_qs, scope = _get_scope(user)
    centre_ids = list(centres_qs.values_list("id", flat=True))

    inscriptions = Inscription.objects.filter(
        formation__centre_id__in=centre_ids
    )
    dettes = Dette.objects.filter(
        inscription__formation__centre_id__in=centre_ids
    )
    # annule=False : un versement annulé n'est plus considéré comme encaissé
    # (cf. Dette.montant_paye()) — exclu ici une bonne fois pour toutes les
    # statistiques/exports qui réutilisent ce queryset (totaux, recouvrement,
    # répartition par mode de paiement, liste des paiements encaissés).
    paiements = Paiement.objects.filter(
        dette__inscription__formation__centre_id__in=centre_ids, annule=False
    )
    return inscriptions, dettes, paiements, centres_qs, directions_qs, scope


def _apply_stats_filters(request, inscriptions_qs, dettes_qs, paiements_qs, scope):
    """
    Applique aux 3 querysets (inscriptions/dettes/paiements) les filtres lus
    dans request.GET. Utilisée à la fois par statistiques_view (affichage
    écran) et par les exports (CSV/Excel/PDF), pour garantir que les exports
    reflètent TOUJOURS exactement ce qui est filtré à l'écran.
    """
    centre_id    = request.GET.get("centre")
    direction_id = request.GET.get("direction")
    filiere_id   = request.GET.get("filiere")
    annee_id     = request.GET.get("annee")
    statut_f     = request.GET.get("statut")
    region_id    = request.GET.get("region")
    genre        = request.GET.get("genre")
    date_debut   = request.GET.get("date_debut")
    date_fin     = request.GET.get("date_fin")
    statut_paiement_f = request.GET.get("statut_paiement")

    if direction_id and scope == "global":
        inscriptions_qs = inscriptions_qs.filter(formation__centre__direction_id=direction_id)
        dettes_qs = dettes_qs.filter(inscription__formation__centre__direction_id=direction_id)
        paiements_qs = paiements_qs.filter(dette__inscription__formation__centre__direction_id=direction_id)

    if centre_id and scope in ("global", "direction"):
        inscriptions_qs = inscriptions_qs.filter(formation__centre_id=centre_id)
        dettes_qs = dettes_qs.filter(inscription__formation__centre_id=centre_id)
        paiements_qs = paiements_qs.filter(dette__inscription__formation__centre_id=centre_id)

    if filiere_id:
        inscriptions_qs = inscriptions_qs.filter(formation__filiere_id=filiere_id)
        dettes_qs = dettes_qs.filter(inscription__formation__filiere_id=filiere_id)
        paiements_qs = paiements_qs.filter(dette__inscription__formation__filiere_id=filiere_id)

    if annee_id:
        inscriptions_qs = inscriptions_qs.filter(annee_scolaire_id=annee_id)
        dettes_qs = dettes_qs.filter(inscription__annee_scolaire_id=annee_id)
        paiements_qs = paiements_qs.filter(dette__inscription__annee_scolaire_id=annee_id)

    if statut_f:
        inscriptions_qs = inscriptions_qs.filter(statut=statut_f)
        dettes_qs = dettes_qs.filter(inscription__statut=statut_f)
        paiements_qs = paiements_qs.filter(dette__inscription__statut=statut_f)

    if region_id:
        inscriptions_qs = inscriptions_qs.filter(formation__centre__province__region_id=region_id)
        dettes_qs = dettes_qs.filter(inscription__formation__centre__province__region_id=region_id)
        paiements_qs = paiements_qs.filter(dette__inscription__formation__centre__province__region_id=region_id)

    if genre:
        inscriptions_qs = inscriptions_qs.filter(eleve__sexe=genre)
        dettes_qs = dettes_qs.filter(inscription__eleve__sexe=genre)
        paiements_qs = paiements_qs.filter(dette__inscription__eleve__sexe=genre)

    if date_debut:
        inscriptions_qs = inscriptions_qs.filter(date_inscription__date__gte=date_debut)
        dettes_qs = dettes_qs.filter(inscription__date_inscription__date__gte=date_debut)
        paiements_qs = paiements_qs.filter(dette__inscription__date_inscription__date__gte=date_debut)

    if date_fin:
        inscriptions_qs = inscriptions_qs.filter(date_inscription__date__lte=date_fin)
        dettes_qs = dettes_qs.filter(inscription__date_inscription__date__lte=date_fin)
        paiements_qs = paiements_qs.filter(dette__inscription__date_inscription__date__lte=date_fin)

    # Statut de paiement (totalement/partiellement/pas réglé) : ne se traduit
    # pas par un simple champ à filtrer (il faut sommer dettes/paiements par
    # inscription), et sommer sur deux relations inverses en cascade dans un
    # seul .annotate() multiplierait les montants (piège classique de l'ORM
    # Django) — calculé en Python, comme déjà fait pour la liste des
    # paiements (paiement_list), puis appliqué en id__in.
    if statut_paiement_f in ("totalement", "partiellement", "aucun"):
        ids_ok = []
        for insc in inscriptions_qs.prefetch_related('dettes__paiements'):
            total_du = sum(d.montant_total for d in insc.dettes.all())
            total_paye = sum(
                p.montant_paiement for d in insc.dettes.all() for p in d.paiements.all() if not p.annule
            )
            if total_du <= 0:
                continue
            if statut_paiement_f == "totalement" and total_paye >= total_du:
                ids_ok.append(insc.id)
            elif statut_paiement_f == "partiellement" and 0 < total_paye < total_du:
                ids_ok.append(insc.id)
            elif statut_paiement_f == "aucun" and total_paye <= 0:
                ids_ok.append(insc.id)
        inscriptions_qs = inscriptions_qs.filter(id__in=ids_ok)
        dettes_qs = dettes_qs.filter(inscription_id__in=ids_ok)
        paiements_qs = paiements_qs.filter(dette__inscription_id__in=ids_ok)

    filters = {
        "centre_id": centre_id, "direction_id": direction_id, "filiere_id": filiere_id,
        "annee_id": annee_id, "statut_f": statut_f, "region_id": region_id,
        "genre": genre, "date_debut": date_debut, "date_fin": date_fin,
        "statut_paiement_f": statut_paiement_f,
    }
    return inscriptions_qs, dettes_qs, paiements_qs, filters


def _resume_filtres_stats(filters):
    """Phrase récapitulant les filtres actifs du tableau de bord statistiques,
    reprise dans les fichiers exportés (CSV/Excel/PDF) pour que leur contenu
    reste traçable une fois détaché de l'écran qui les a produits."""
    from .models import Region
    from accounts.models import Utilisateur

    parties = []

    if filters.get("centre_id"):
        centre = CentreFormation.objects.filter(pk=filters["centre_id"]).first()
        parties.append(f"Centre : {centre.nom_centre if centre else '—'}")

    if filters.get("direction_id"):
        direction = Direction_reg.objects.filter(pk=filters["direction_id"]).first()
        parties.append(f"Direction : {direction.nom_direction if direction else '—'}")

    if filters.get("filiere_id"):
        filiere = Filiere.objects.filter(pk=filters["filiere_id"]).first()
        parties.append(f"Métier : {filiere.nom_filiere if filiere else '—'}")

    if filters.get("annee_id"):
        annee = AnneeScolaire.objects.filter(pk=filters["annee_id"]).first()
        parties.append(f"Année de formation : {annee.libelle_anne if annee else '—'}")

    if filters.get("statut_f"):
        statut_labels = dict(Inscription.STATUT_CHOICE)
        parties.append(f"Statut inscription : {statut_labels.get(filters['statut_f'], filters['statut_f'])}")

    if filters.get("region_id"):
        region = Region.objects.filter(pk=filters["region_id"]).first()
        parties.append(f"Région : {region.nom_region if region else '—'}")

    if filters.get("genre"):
        genre_labels = dict(Utilisateur.SEXE_CHOICE)
        parties.append(f"Genre : {genre_labels.get(filters['genre'], filters['genre'])}")

    if filters.get("date_debut") or filters.get("date_fin"):
        parties.append(f"Période : du {filters.get('date_debut') or '…'} au {filters.get('date_fin') or '…'}")

    if filters.get("statut_paiement_f"):
        statut_paiement_labels = {
            "totalement": "Totalement réglé",
            "partiellement": "Partiellement réglé",
            "aucun": "Rien n'est réglé",
        }
        parties.append(
            f"Statut de paiement : {statut_paiement_labels.get(filters['statut_paiement_f'], filters['statut_paiement_f'])}"
        )

    return " | ".join(parties) if parties else "Aucun filtre appliqué"


def _can_access_eleve_finances(user, eleve):
    """
    Contrôle d'accès pour les vues financières d'un élève (dettes/quittances) :
    l'élève lui-même, un superuser, ou un membre du personnel dont le périmètre
    (centre/direction/global) couvre au moins une des inscriptions de l'élève —
    ou qui a `rechercher_tous_centres`, qui étend volontairement l'accès à
    tous les centres (voir migration 0031 : cette permission existe pour que
    la recherche cross-centre d'un caissier/directeur reste cohérente avec ce
    qu'`encaisser_paiement` autorise déjà — action jamais limitée par centre).
    """
    if user.is_superuser:
        return True
    if getattr(user, 'pk', None) == getattr(eleve, 'pk', None):
        return True
    if not user.has_perm('courses.voir_inscriptions'):
        return False
    if user.has_perm('courses.rechercher_tous_centres'):
        return True
    centres_qs, _, scope = _get_scope(user)
    if scope == "none":
        return False
    if scope == "global":
        return True
    centre_ids = list(centres_qs.values_list("id", flat=True))
    return Inscription.objects.filter(eleve=eleve, formation__centre_id__in=centre_ids).exists()


TAILLE_MAX_UPLOAD = 5 * 1024 * 1024  # 5 Mo, aligne sur client_max_body_size de nginx

def _valider_fichier_upload(fichier):
    """Valide un fichier envoye par un utilisateur (piece jointe justificative,
    document requis...) : extension ET signature binaire verifiees (empeche
    un fichier renomme, ex. .html en .pdf), taille max 5 Mo. Formats
    acceptes : PDF, JPEG, JPG, PNG. Retourne un message d'erreur (str) si
    invalide, None si le fichier est valide."""
    if fichier.size > TAILLE_MAX_UPLOAD:
        return f"« {fichier.name} » dépasse la taille maximale de 5 Mo."

    nom = fichier.name.lower()
    entete = fichier.read(8)
    fichier.seek(0)

    if nom.endswith('.pdf'):
        if entete[:5] != b'%PDF-':
            return f"« {fichier.name} » n'est pas un fichier PDF valide."
    elif nom.endswith('.jpg') or nom.endswith('.jpeg'):
        if entete[:3] != b'\xff\xd8\xff':
            return f"« {fichier.name} » n'est pas une image JPEG valide."
    elif nom.endswith('.png'):
        if entete[:8] != b'\x89PNG\r\n\x1a\n':
            return f"« {fichier.name} » n'est pas une image PNG valide."
    else:
        return f"« {fichier.name} » : formats acceptés — JPEG, JPG, PNG, PDF uniquement."

    return None


def _libelle_blocage(dette_bloquante, tranche_bloquante):
    """Fragment de phrase décrivant ce qui bloque le paiement — une tranche
    précise, ou le frais entier quand la dette bloquante (ex. frais de
    dossier) n'a pas de tranches."""
    if tranche_bloquante:
        return f"la tranche « {tranche_bloquante.libelle} » de « {dette_bloquante.frais_formation.type_frais} »"
    return f"le frais « {dette_bloquante.frais_formation.type_frais} »"


def _reste_blocage(dette_bloquante, tranche_bloquante):
    """Montant restant à régler sur la dette/tranche bloquante."""
    if tranche_bloquante:
        return dette_bloquante.reste_pour_tranche(tranche_bloquante)
    return dette_bloquante.reste_a_payer()


def _can_access_dette_finances(user, dette):
    """Même contrôle que _can_access_eleve_finances, mais scopé à UNE dette précise."""
    if user.is_superuser:
        return True
    eleve = dette.inscription.eleve
    if getattr(user, 'pk', None) == getattr(eleve, 'pk', None):
        return True
    if not user.has_perm('courses.voir_inscriptions'):
        return False
    if user.has_perm('courses.rechercher_tous_centres'):
        return True
    centres_qs, _, scope = _get_scope(user)
    if scope == "none":
        return False
    if scope == "global":
        return True
    centre_id = dette.inscription.formation.centre_id if dette.inscription.formation_id else None
    return centre_id is not None and centres_qs.filter(pk=centre_id).exists()


# ─── Vue principale ───────────────────────────────────────────────────────────

@login_required
def statistiques_view(request):
    user = request.user
    inscriptions_qs, dettes_qs, paiements_qs, centres_scope, directions_scope, scope = _base_qs(user)

    inscriptions_qs, dettes_qs, paiements_qs, filters = _apply_stats_filters(
        request, inscriptions_qs, dettes_qs, paiements_qs, scope
    )
    centre_id    = filters["centre_id"]
    direction_id = filters["direction_id"]
    filiere_id   = filters["filiere_id"]
    annee_id     = filters["annee_id"]
    statut_f     = filters["statut_f"]
    region_id    = filters["region_id"]
    genre        = filters["genre"]
    date_debut   = filters["date_debut"]
    date_fin     = filters["date_fin"]
    statut_paiement_f = filters["statut_paiement_f"]

    # Narrowing du dropdown "centre" affiché à l'écran quand une direction est sélectionnée.
    if direction_id and scope == "global":
        centres_scope = centres_scope.filter(direction_id=direction_id)

    # ── KPIs ─────────────────────────────────────────────────────────────────
    total_encaisse = paiements_qs.aggregate(s=Sum("montant_paiement"))["s"] or 0
    total_du       = dettes_qs.aggregate(s=Sum("montant_total"))["s"] or 0
    total_restant  = max(total_du - total_encaisse, 0)
    taux_global    = round(total_encaisse / total_du * 100, 1) if total_du > 0 else 0

    centre_ids_scope = list(centres_scope.values_list("id", flat=True))

    # Champ "Métier" : si un centre précis est sélectionné, ses métiers priment ;
    # sinon on retombe sur les métiers déjà lancés dans les centres du périmètre
    # (centres de la direction sélectionnée pour l'accès national, centres de la
    # direction/du centre de portée pour les accès directionnel/centre) ; en
    # accès national sans aucun filtre, tous les métiers actifs sont proposés.
    if centre_id and scope in ("global", "direction"):
        filieres_scope = Filiere.objects.filter(
            is_active=True, centreetfiliere__centre_id=centre_id
        ).distinct()
    elif scope == "global" and not direction_id:
        filieres_scope = Filiere.objects.filter(is_active=True)
    else:
        filieres_scope = Filiere.objects.filter(
            is_active=True, centreetfiliere__centre_id__in=centre_ids_scope
        ).distinct()

    stats = {
        "total_eleves":           Eleve.objects.filter(inscription__in=inscriptions_qs).distinct().count(),
        "inscriptions_validees":  inscriptions_qs.filter(statut__in=["valide", "valide_paye", "Valide"]).count(),
        "inscriptions_en_cours":  inscriptions_qs.filter(statut="en_cours").count(),
        "inscriptions_rejetees":  inscriptions_qs.filter(statut="rejete").count(),
        "total_encaisse":         total_encaisse,
        "total_restant":          total_restant,
        "total_du":               total_du,
        "taux_global":            taux_global,
        "total_centres":          centres_scope.count(),
        "total_filieres":         filieres_scope.count(),
        "total_directions":       directions_scope.count(),
    }

    # ── Top métiers ───────────────────────────────────────────────────────────
    top_filieres_qs = (
        inscriptions_qs
        .values("formation__filiere__nom_filiere")
        .annotate(count=Count("id"))
        .order_by("-count")[:8]
    )
    top_filieres = [
        {"nom": f["formation__filiere__nom_filiere"] or "—", "count": f["count"]}
        for f in top_filieres_qs
    ]

    # ── Taux de recouvrement par centre ───────────────────────────────────────
    recouvrement_centres = []
    for centre in centres_scope.order_by("nom_centre"):
        c_dettes    = dettes_qs.filter(inscription__formation__centre=centre)
        c_paiements = paiements_qs.filter(dette__inscription__formation__centre=centre)
        c_du  = c_dettes.aggregate(s=Sum("montant_total"))["s"] or 0
        c_enc = c_paiements.aggregate(s=Sum("montant_paiement"))["s"] or 0
        c_rest = max(c_du - c_enc, 0)
        taux = round(c_enc / c_du * 100, 1) if c_du > 0 else 0
        recouvrement_centres.append({
            "nom_centre":   centre.nom_centre,
            "direction":    centre.direction.nom_direction if centre.direction else "—",
            "total_du":     c_du,
            "encaisse":     c_enc,
            "restant":      c_rest,
            "taux":         taux,
            "inscrits":     inscriptions_qs.filter(formation__centre=centre).count(),
        })
    recouvrement_centres.sort(key=lambda x: x["taux"], reverse=True)

    recouvrement_paginator = Paginator(recouvrement_centres, 5)
    recouvrement_page = recouvrement_paginator.get_page(request.GET.get("rpage"))

    # ── Évolution mensuelle inscriptions (12 derniers mois) ──────────────────
    from django.db.models.functions import TruncMonth
    evol_qs = (
        inscriptions_qs
        .annotate(mois=TruncMonth("date_inscription"))
        .values("mois")
        .annotate(count=Count("id"))
        .order_by("mois")
    )
    evol_labels = []
    evol_data   = []
    MOIS_FR = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]
    for e in evol_qs:
        if e["mois"]:
            evol_labels.append(f"{MOIS_FR[e['mois'].month-1]} {e['mois'].year}")
            evol_data.append(e["count"])

    # ── Paiements par mode ────────────────────────────────────────────────────
    modes = (
        paiements_qs
        .values("mode_paiement")
        .annotate(total=Sum("montant_paiement"))
        .order_by("-total")
    )
    mode_labels = [m["mode_paiement"].capitalize() for m in modes]
    mode_data   = [m["total"] or 0 for m in modes]

    # ── Historique des inscriptions (paginé, 5/page) ──────────────────────────
    inscriptions_liste_qs = (
        inscriptions_qs
        .select_related("eleve", "formation__filiere", "formation__centre", "annee_scolaire")
        .order_by("-date_inscription")
    )
    inscriptions_paginator = Paginator(inscriptions_liste_qs, 5)
    dernieres_inscriptions = inscriptions_paginator.get_page(request.GET.get("ipage"))

    from .models import Region
    from accounts.models import Utilisateur

    def _direction_regions(direction_obj):
        """Régions couvertes par une direction (Direction_reg.region, texte
        séparé par virgules — voir DirectionRegForm)."""
        if not direction_obj or not direction_obj.region:
            return Region.objects.none()
        noms = [r.strip() for r in direction_obj.region.split(',') if r.strip()]
        return Region.objects.filter(nom_region__in=noms)

    if scope == "global":
        if direction_id:
            regions_scope = _direction_regions(Direction_reg.objects.filter(pk=direction_id).first())
        else:
            regions_scope = Region.objects.all()
    elif scope == "direction":
        regions_scope = _direction_regions(directions_scope.first())
    else:
        regions_scope = Region.objects.filter(
            provinces__centre_formations__id__in=centre_ids_scope
        ).distinct()

    # Querystrings pour les liens de pagination des deux sections paginées
    # séparément (rpage/ipage) : on garde tous les filtres actifs, on retire
    # juste le paramètre de page de la section concernée.
    qd_recouvrement = request.GET.copy()
    qd_recouvrement.pop("rpage", None)
    querystring_recouvrement = qd_recouvrement.urlencode()

    qd_inscriptions = request.GET.copy()
    qd_inscriptions.pop("ipage", None)
    querystring_inscriptions = qd_inscriptions.urlencode()

    context = {
        "stats":                  stats,
        "top_filieres":           top_filieres,
        "recouvrement_centres":   recouvrement_page,
        "dernieres_inscriptions": dernieres_inscriptions,
        "querystring_recouvrement": querystring_recouvrement,
        "querystring_inscriptions": querystring_inscriptions,
        "scope":                  scope,
        # Filtres disponibles
        "centres":    centres_scope.order_by("nom_centre"),
        "directions": directions_scope.order_by("nom_direction"),
        "filieres":   filieres_scope.order_by("nom_filiere"),
        "annees":     AnneeScolaire.objects.all().order_by("-libelle_anne"),
        "regions":    regions_scope.order_by("nom_region"),
        "genres":     Utilisateur.SEXE_CHOICE,
        # Valeurs actives des filtres
        "f_centre":     centre_id,
        "f_direction":  direction_id,
        "f_filiere":    filiere_id,
        "f_annee":      annee_id,
        "f_statut":     statut_f,
        "f_region":     region_id,
        "f_genre":      genre,
        "f_date_debut": date_debut,
        "f_date_fin":   date_fin,
        "f_statut_paiement": statut_paiement_f,
        # Donnees des graphiques. Transmises brutes : le template les serialise
        # avec le filtre `json_script`, qui echappe <, > et & — ce que
        # json.dumps ne fait pas. Un nom de metier contenant « </script> »
        # sortait autrement du bloc <script> (XSS stocke).
        "evol_labels":       evol_labels,
        "evol_data":         evol_data,
        "mode_labels":       mode_labels,
        "mode_data":         mode_data,
        "top_filieres_noms":  [f["nom"] for f in top_filieres],
        "top_filieres_count": [f["count"] for f in top_filieres],
    }
    return render(request, "member/statistiques/statistiques.html", context)


# ─── Export CSV ───────────────────────────────────────────────────────────────

@login_required
def export_csv(request):
    user = request.user
    inscriptions_qs, dettes_qs, paiements_qs, centres_scope, _, scope = _base_qs(user)
    inscriptions_qs, dettes_qs, paiements_qs, filters = _apply_stats_filters(
        request, inscriptions_qs, dettes_qs, paiements_qs, scope
    )

    export_type = request.GET.get("type", "inscriptions")

    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = f'attachment; filename="export_{export_type}.csv"'

    writer = csv.writer(response, delimiter=";")
    writer.writerow(["Filtres appliqués :", _resume_filtres_stats(filters)])
    writer.writerow([])

    if export_type == "inscriptions":
        writer.writerow(["N°", "Apprenant", "Matricule", "Sexe", "Téléphone", "Email", "Métier", "Centre", "Direction", "Année", "Statut", "Date inscription"])
        for i, insc in enumerate(
            inscriptions_qs.select_related(
                "eleve", "formation__filiere", "formation__centre__direction", "annee_scolaire"
            ).order_by("-date_inscription"), 1
        ):
            writer.writerow([
                i,
                f"{insc.eleve.nom} {insc.eleve.prenom}" if insc.eleve else "—",
                (insc.eleve.matricule or "—") if insc.eleve else "—",
                insc.eleve.get_sexe_display() if insc.eleve else "—",
                insc.eleve.tel if insc.eleve else "—",
                insc.eleve.email if insc.eleve else "—",
                insc.formation.filiere.nom_filiere if insc.formation and insc.formation.filiere else "—",
                insc.formation.centre.nom_centre if insc.formation and insc.formation.centre else "—",
                insc.formation.centre.direction.nom_direction if insc.formation and insc.formation.centre and insc.formation.centre.direction else "—",
                insc.annee_scolaire.libelle_anne if insc.annee_scolaire else "—",
                insc.get_statut_display(),
                insc.date_inscription.strftime("%d/%m/%Y") if insc.date_inscription else "—",
            ])

    elif export_type == "paiements":
        writer.writerow(["N°", "Apprenant", "Quittance", "Montant (FCFA)", "Mode", "Date", "Centre"])
        for i, p in enumerate(
            paiements_qs.select_related(
                "dette__inscription__eleve",
                "dette__inscription__formation__centre",
            ).order_by("-date_paiement"), 1
        ):
            insc = p.dette.inscription if p.dette else None
            eleve = insc.eleve if insc else None
            centre = insc.formation.centre if insc and insc.formation else None
            writer.writerow([
                i,
                f"{eleve.nom} {eleve.prenom}" if eleve else "—",
                p.numero_quittance or "—",
                p.montant_paiement,
                p.mode_paiement,
                p.date_paiement.strftime("%d/%m/%Y") if p.date_paiement else "—",
                centre.nom_centre if centre else "—",
            ])

    elif export_type == "recouvrement":
        writer.writerow(["Centre", "Direction", "Total dû (FCFA)", "Encaissé (FCFA)", "Restant (FCFA)", "Taux (%)"])
        for centre in centres_scope.order_by("nom_centre"):
            c_dettes    = dettes_qs.filter(inscription__formation__centre=centre)
            c_paiements = paiements_qs.filter(dette__inscription__formation__centre=centre)
            c_du   = c_dettes.aggregate(s=Sum("montant_total"))["s"] or 0
            c_enc  = c_paiements.aggregate(s=Sum("montant_paiement"))["s"] or 0
            c_rest = max(c_du - c_enc, 0)
            taux   = round(c_enc / c_du * 100, 1) if c_du > 0 else 0
            writer.writerow([
                centre.nom_centre,
                centre.direction.nom_direction if centre.direction else "—",
                c_du, c_enc, c_rest, taux,
            ])

    return response


# ─── Export Excel ─────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    user = request.user
    inscriptions_qs, dettes_qs, paiements_qs, centres_scope, _, scope = _base_qs(user)
    inscriptions_qs, dettes_qs, paiements_qs, filters = _apply_stats_filters(
        request, inscriptions_qs, dettes_qs, paiements_qs, scope
    )
    export_type = request.GET.get("type", "inscriptions")

    wb = Workbook()
    ws = wb.active

    rouge_fill = PatternFill("solid", fgColor="C0392B")
    or_fill    = PatternFill("solid", fgColor="D4A017")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    filtres_font = Font(italic=True, color="6B7280", size=9)

    def style_header(row_cells):
        for cell in row_cells:
            cell.fill = rouge_fill
            cell.font = header_font
            cell.alignment = center_align

    def ecrire_resume_filtres():
        ws.append([f"Filtres appliqués : {_resume_filtres_stats(filters)}"])
        ws["A" + str(ws.max_row)].font = filtres_font
        ws.append([])

    if export_type == "inscriptions":
        ws.title = "Inscriptions"
        ecrire_resume_filtres()
        headers = ["N°","Apprenant","Matricule","Sexe","Téléphone","Email","Métier","Centre","Direction","Année","Statut","Date"]
        ws.append(headers)
        style_header(ws[ws.max_row])
        for i, insc in enumerate(
            inscriptions_qs.select_related(
                "eleve","formation__filiere","formation__centre__direction","annee_scolaire"
            ).order_by("-date_inscription"), 1
        ):
            ws.append([
                i,
                f"{insc.eleve.nom} {insc.eleve.prenom}" if insc.eleve else "—",
                (insc.eleve.matricule or "—") if insc.eleve else "—",
                insc.eleve.get_sexe_display() if insc.eleve else "—",
                insc.eleve.tel if insc.eleve else "—",
                insc.eleve.email if insc.eleve else "—",
                insc.formation.filiere.nom_filiere if insc.formation and insc.formation.filiere else "—",
                insc.formation.centre.nom_centre if insc.formation and insc.formation.centre else "—",
                insc.formation.centre.direction.nom_direction if insc.formation and insc.formation.centre and insc.formation.centre.direction else "—",
                insc.annee_scolaire.libelle_anne if insc.annee_scolaire else "—",
                insc.get_statut_display(),
                insc.date_inscription.strftime("%d/%m/%Y") if insc.date_inscription else "—",
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    elif export_type == "recouvrement":
        ws.title = "Recouvrement"
        ecrire_resume_filtres()
        headers = ["Centre","Direction","Total dû (FCFA)","Encaissé (FCFA)","Restant (FCFA)","Taux (%)"]
        ws.append(headers)
        style_header(ws[ws.max_row])
        for centre in centres_scope.order_by("nom_centre"):
            c_dettes    = dettes_qs.filter(inscription__formation__centre=centre)
            c_paiements = paiements_qs.filter(dette__inscription__formation__centre=centre)
            c_du   = c_dettes.aggregate(s=Sum("montant_total"))["s"] or 0
            c_enc  = c_paiements.aggregate(s=Sum("montant_paiement"))["s"] or 0
            c_rest = max(c_du - c_enc, 0)
            taux   = round(c_enc / c_du * 100, 1) if c_du > 0 else 0
            ws.append([
                centre.nom_centre,
                centre.direction.nom_direction if centre.direction else "—",
                c_du, c_enc, c_rest, taux,
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="export_{export_type}.xlsx"'
    wb.save(response)
    return response


# ─── Export PDF ───────────────────────────────────────────────────────────────

@login_required
def export_pdf(request):
    user = request.user
    inscriptions_qs, dettes_qs, paiements_qs, centres_scope, directions_scope, scope = _base_qs(user)
    inscriptions_qs, dettes_qs, paiements_qs, filters = _apply_stats_filters(
        request, inscriptions_qs, dettes_qs, paiements_qs, scope
    )
    export_type = request.GET.get("type", "inscriptions")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_bsb",
        parent=styles["Title"],
        fontSize=16,
        textColor=rl_colors.HexColor("#C0392B"),
        spaceAfter=12,
    )
    sub_style = ParagraphStyle(
        "sub_bsb",
        parent=styles["Normal"],
        fontSize=9,
        textColor=rl_colors.HexColor("#6B7280"),
        spaceAfter=16,
    )
    cell_style = ParagraphStyle(
        "cell_bsb", parent=styles["Normal"], fontSize=8, leading=10,
    )
    filtres_style = ParagraphStyle(
        "filtres_bsb",
        parent=styles["Normal"],
        fontSize=8,
        fontName="Helvetica-Oblique",
        textColor=rl_colors.HexColor("#6B7280"),
        spaceAfter=10,
    )

    def cell(texte):
        """Cellule de tableau avec retour à la ligne au lieu d'une troncature."""
        return Paragraph(str(texte), cell_style)

    rouge = rl_colors.HexColor("#C0392B")
    or_cl = rl_colors.HexColor("#D4A017")
    gris  = rl_colors.HexColor("#F3F4F6")

    def base_table_style(header_rows=1):
        return TableStyle([
            ("BACKGROUND",  (0, 0), (-1, header_rows-1), rouge),
            ("TEXTCOLOR",   (0, 0), (-1, header_rows-1), rl_colors.white),
            ("FONTNAME",    (0, 0), (-1, header_rows-1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, header_rows-1), 9),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [rl_colors.white, gris]),
            ("FONTSIZE",    (0, header_rows), (-1, -1), 8),
            ("GRID",        (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#E5E7EB")),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])

    story = []
    now = timezone.now().strftime("%d/%m/%Y %H:%M")

    header_left, header_right = _pdf_header_lines(
        centre=centres_scope.first() if scope == "centre" else None,
        direction=directions_scope.first() if scope == "direction" else None,
    )
    header_line_style = ParagraphStyle(
        "pdf_header_line", parent=styles["Normal"],
        fontSize=6, leading=8, alignment=1, fontName="Helvetica-Bold",
    )
    favicon_path = os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
    header_table = Table(
        [[
            Paragraph("<br/>".join(header_left), header_line_style),
            Image(favicon_path, width=1.6*cm, height=1.6*cm),
            Paragraph("<br/>".join(header_right), header_line_style),
        ]],
        colWidths=[10*cm, 3*cm, 10*cm],
    )
    header_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 12))

    if export_type == "inscriptions":
        story.append(Paragraph("Rapport des Inscriptions — BSB", title_style))
        story.append(Paragraph(f"Généré le {now}  |  {inscriptions_qs.count()} inscription(s)", sub_style))
        story.append(Paragraph(f"Filtres appliqués : {_resume_filtres_stats(filters)}", filtres_style))

        data = [["N°", "Apprenant", "Matricule", "Métier", "Centre", "Année", "Statut", "Date"]]
        for i, insc in enumerate(
            inscriptions_qs.select_related(
                "eleve","formation__filiere","formation__centre","annee_scolaire"
            ).order_by("-date_inscription")[:500], 1
        ):
            data.append([
                str(i),
                cell(f"{insc.eleve.nom} {insc.eleve.prenom}") if insc.eleve else "—",
                cell(insc.eleve.matricule) if insc.eleve and insc.eleve.matricule else "—",
                cell(insc.formation.filiere.nom_filiere) if insc.formation and insc.formation.filiere else "—",
                cell(insc.formation.centre.nom_centre) if insc.formation and insc.formation.centre else "—",
                insc.annee_scolaire.libelle_anne if insc.annee_scolaire else "—",
                cell(insc.get_statut_display()),
                insc.date_inscription.strftime("%d/%m/%Y") if insc.date_inscription else "—",
            ])

        col_widths = [1.2*cm, 4.5*cm, 3.5*cm, 4*cm, 4*cm, 2.5*cm, 4*cm, 2.5*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(base_table_style())
        story.append(t)

    elif export_type == "recouvrement":
        story.append(Paragraph("Rapport de Recouvrement par Centre — BSB", title_style))
        story.append(Paragraph(f"Généré le {now}", sub_style))
        story.append(Paragraph(f"Filtres appliqués : {_resume_filtres_stats(filters)}", filtres_style))

        data = [["Centre", "Direction", "Total dû (FCFA)", "Encaissé (FCFA)", "Restant (FCFA)", "Taux (%)"]]
        for centre in centres_scope.order_by("nom_centre"):
            c_dettes    = dettes_qs.filter(inscription__formation__centre=centre)
            c_paiements = paiements_qs.filter(dette__inscription__formation__centre=centre)
            c_du   = c_dettes.aggregate(s=Sum("montant_total"))["s"] or 0
            c_enc  = c_paiements.aggregate(s=Sum("montant_paiement"))["s"] or 0
            c_rest = max(c_du - c_enc, 0)
            taux   = round(c_enc / c_du * 100, 1) if c_du > 0 else 0
            data.append([
                cell(centre.nom_centre),
                cell(centre.direction.nom_direction) if centre.direction else "—",
                f"{c_du:,.0f}".replace(",", " "),
                f"{c_enc:,.0f}".replace(",", " "),
                f"{c_rest:,.0f}".replace(",", " "),
                f"{taux}%",
            ])

        col_widths = [5*cm, 5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 3*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(base_table_style())
        story.append(t)

    signataire = {
        "centre": "Le Gestionnaire du centre",
        "direction": "Le Directeur Inter-Régional",
        "global": "Le Directeur Général",
    }.get(scope, "Le Directeur Général")
    signature_style = ParagraphStyle(
        "signature_bsb", parent=styles["Normal"], fontSize=10,
        alignment=2, spaceBefore=28,
    )
    story.append(Paragraph(signataire, signature_style))

    footer_style_left = ParagraphStyle(
        "footer_bsb_left", parent=styles["Normal"], fontSize=7,
        textColor=rl_colors.grey, alignment=0,
    )
    footer_style_right = ParagraphStyle(
        "footer_bsb_right", parent=styles["Normal"], fontSize=7,
        textColor=rl_colors.grey, alignment=2,
    )
    footer_table = Table(
        [[Paragraph("BSB", footer_style_left), Paragraph(f"généré sur YU-PAAN le : {now}", footer_style_right)]],
        colWidths=[doc.width / 2, doc.width / 2],
    )
    footer_table.setStyle(TableStyle([
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(Spacer(1, 24))
    story.append(footer_table)

    def _watermark_page(canvas_obj, doc_obj):
        _draw_pdf_watermark(canvas_obj, doc_obj.pagesize[0], doc_obj.pagesize[1])

    doc.build(story, onFirstPage=_watermark_page, onLaterPages=_watermark_page)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="rapport_{export_type}.pdf"'
    return response


# ─────────────────────────────────────────────
# DETTES D'UN ÉLÈVE (groupées par inscription)
# ─────────────────────────────────────────────
@login_required
def stats_dettes_eleve_view(request, eleve_id):
    eleve = get_object_or_404(Eleve, id=eleve_id)

    if not _can_access_eleve_finances(request.user, eleve):
        raise PermissionDenied("Vous n'avez pas accès aux informations financières de cet apprenant.")

    inscriptions = (
        Inscription.objects
        .filter(eleve=eleve)
        .exclude(statut='rejete')
        .select_related('formation__filiere', 'formation__centre', 'annee_scolaire')
        .prefetch_related('dettes__frais_formation__type_frais__tranches', 'dettes__paiements')
        .order_by('-date_inscription')
    )

    if (
        not request.user.is_superuser
        and getattr(request.user, 'pk', None) != eleve.pk
        and not request.user.has_perm('courses.rechercher_tous_centres')
    ):
        centres_qs, _, scope = _get_scope(request.user)
        if scope != "global":
            centre_ids = list(centres_qs.values_list("id", flat=True))
            inscriptions = inscriptions.filter(formation__centre_id__in=centre_ids)

    inscription_id = request.GET.get('inscription')
    if inscription_id:
        inscriptions = inscriptions.filter(id=inscription_id)

    inscriptions_dettes = []
    total_restant_global = 0

    for insc in inscriptions:
        dettes_data = []
        insc_du = 0
        insc_paye = 0
        insc_reste = 0
        dossier_impaye = False

        for dette in insc.dettes.all():
            paye = dette.montant_paye()
            reste = dette.reste_a_payer()
            taux = (paye / dette.montant_total * 100) if dette.montant_total > 0 else 0
            insc_du += dette.montant_total
            insc_paye += paye
            insc_reste += max(reste, 0)
            total_restant_global += max(reste, 0)

            tranche_cible = dette.tranche_a_payer()
            if dette.frais_formation.type_frais.est_frais_de_dossier and reste > 0:
                dossier_impaye = True

            dettes_data.append({
                'id': dette.id,
                'frais_formation': dette.frais_formation,
                'montant_total': dette.montant_total,
                'montant_paye': paye,
                'reste': max(reste, 0),
                'taux': min(round(taux, 1), 100),
                'etat_dette': dette.etat_dette,
                'tranches': dette.tranches_detail(),
                'bloquee': dette.bloquee_par_autre_dette(),
                'montant_a_payer': dette.montant_a_payer(),
                'tranche_cible_primordiale': bool(tranche_cible and tranche_cible.est_primordiale),
                'est_frais_dossier': dette.frais_formation.type_frais.est_frais_de_dossier,
            })

        dette_bloquante, tranche_bloquante = insc.dette_et_tranche_bloquantes()
        primordiale_bloquante_reste = (
            _reste_blocage(dette_bloquante, tranche_bloquante) if dette_bloquante else 0
        )

        inscriptions_dettes.append({
            'inscription': insc,
            'dettes': dettes_data,
            'total_du': insc_du,
            'total_paye': insc_paye,
            'total_reste': insc_reste,
            'dossier_impaye': dossier_impaye,
            'primordiale_bloquante_reste': primordiale_bloquante_reste,
        })

    return render(request, 'member/statistiques/stats_dettes_eleve.html', {
        'eleve': eleve,
        'inscriptions_dettes': inscriptions_dettes,
        'total_restant': total_restant_global,
    })


class _CascadeInterrompue(Exception):
    """Levée en cours de cascade multi-dettes quand un versement ne peut pas
    être appliqué tel quel (tranche primordiale sous-payée sans motif/pièce
    jointe fournis, ou frais de dossier réglé partiellement) — permet
    d'annuler (rollback) les paiements déjà enregistrés dans la même
    transaction."""
    def __init__(self, message):
        self.message = message
        super().__init__(message)


def _encaisser_montant_dette(dette, montant, mode_paiement, user, motif_derogation=None, piece_jointe_derogation=None, groupe_id=None):
    """
    Encaisse `montant` sur cette dette : tranche primordiale d'abord (ou
    versement unique si le type de frais n'a pas de tranches), puis le reste
    sur les tranches suivantes dans l'ordre. Si `montant` est insuffisant
    pour couvrir entièrement la tranche primordiale non soldée, tout le
    montant lui est affecté en sous-paiement dérogatoire (motif et pièce
    jointe supposés déjà validés par l'appelant). `groupe_id` identifie le
    lot d'encaissement (une action utilisateur peut créer plusieurs
    paiements — c'est ce lot entier qui est annulable, pas une ligne isolée).
    Retourne (nombre de paiements créés, montant réellement encaissé,
    montant non utilisé).
    """
    tranche_num = dette.paiements.count()
    restant = montant
    nb_paiements = 0
    encaisse = 0

    tranches = list(dette.frais_formation.type_frais.tranches.all())
    if not tranches:
        reste = dette.reste_a_payer()
        prise = min(restant, reste)
        if prise > 0:
            tranche_num += 1
            Paiement.objects.create(
                dette=dette, montant_paiement=prise, mode_paiement=mode_paiement,
                tranche=tranche_num, tranche_frais=None,
                date_paiement=timezone.now(), cree_par=user, groupe_id=groupe_id,
            )
            restant -= prise
            encaisse += prise
            nb_paiements += 1
    else:
        primordiale = next((t for t in tranches if t.est_primordiale), None)
        ordre = ([primordiale] if primordiale else []) + [
            t for t in sorted(tranches, key=lambda t: t.ordre) if not primordiale or t.id != primordiale.id
        ]
        for t in ordre:
            if restant <= 0:
                break
            reste_t = dette.reste_pour_tranche(t)
            if reste_t <= 0:
                continue
            if t.est_primordiale and restant < reste_t:
                tranche_num += 1
                Paiement.objects.create(
                    dette=dette, montant_paiement=restant, mode_paiement=mode_paiement,
                    tranche=tranche_num, tranche_frais=t,
                    date_paiement=timezone.now(), cree_par=user, groupe_id=groupe_id,
                    motif_derogation=motif_derogation,
                    piece_jointe_derogation=piece_jointe_derogation,
                )
                encaisse += restant
                nb_paiements += 1
                restant = 0
                break
            prise = min(restant, reste_t)
            tranche_num += 1
            Paiement.objects.create(
                dette=dette, montant_paiement=prise, mode_paiement=mode_paiement,
                tranche=tranche_num, tranche_frais=t,
                date_paiement=timezone.now(), cree_par=user, groupe_id=groupe_id,
            )
            encaisse += prise
            nb_paiements += 1
            restant -= prise

    if dette.reste_a_payer() <= 0:
        dette.etat_dette = 'soldé'
        dette.save()

    return nb_paiements, encaisse, restant


# ─────────────────────────────────────────────
# ENCAISSER UN MONTANT SUR UN TYPE DE FRAIS (une dette)
# ─────────────────────────────────────────────
@login_required
def stats_encaisser_solde_dette_view(request, dette_id):
    dette = get_object_or_404(
        Dette.objects.select_related('inscription__eleve', 'frais_formation__type_frais'),
        id=dette_id
    )

    if not _can_access_dette_finances(request.user, dette):
        raise PermissionDenied("Vous n'avez pas accès aux informations financières de cette dette.")

    if request.method != 'POST':
        return redirect('courses:stats_dettes_eleve', eleve_id=dette.inscription.eleve_id)

    if not (request.user.is_superuser or request.user.has_perm('courses.encaisser_paiement')):
        raise PermissionDenied("Vous n'avez pas la permission d'encaisser un paiement.")

    redirect_url = f"{reverse('courses:stats_dettes_eleve', args=[dette.inscription.eleve_id])}?inscription={dette.inscription_id}"

    dette_bloquante, tranche_bloquante = dette.inscription.dette_et_tranche_bloquantes()
    if dette_bloquante and dette_bloquante.id != dette.id:
        messages.error(
            request,
            f"Il faut d'abord régler entièrement {_libelle_blocage(dette_bloquante, tranche_bloquante)}."
        )
        return redirect(redirect_url)

    mode = request.POST.get('mode_paiement', 'espece')
    montant_str = request.POST.get('montant_paiement', '').strip()
    try:
        montant = float(montant_str)
    except (ValueError, TypeError):
        messages.error(request, "Montant invalide.")
        return redirect(redirect_url)

    if montant <= 0:
        messages.error(request, "Le montant doit être supérieur à 0.")
        return redirect(redirect_url)

    reste_dette = dette.reste_a_payer()
    if montant > reste_dette:
        messages.error(request, f"Le montant saisi ({montant:,.0f} FCFA) dépasse le reste dû ({reste_dette:,.0f} FCFA).")
        return redirect(redirect_url)

    # Un frais de dossier se règle en une seule fois, intégralement : ni
    # paiement partiel, ni notion de tranche/dérogation applicable.
    est_frais_dossier = dette.frais_formation.type_frais.est_frais_de_dossier
    if est_frais_dossier and montant < reste_dette:
        messages.error(
            request,
            f"Le frais de dossier « {dette.frais_formation.type_frais} » doit être réglé intégralement "
            f"en un seul versement (montant dû : {reste_dette:,.0f} FCFA)."
        )
        return redirect(redirect_url)

    tranche_cible = dette.tranche_a_payer()
    motif_derogation = None
    piece_jointe_derogation = None
    if not est_frais_dossier and tranche_cible and tranche_cible.est_primordiale and montant < dette.reste_pour_tranche(tranche_cible):
        motif_derogation = request.POST.get('motif_derogation', '').strip()
        piece_jointe_derogation = request.FILES.get('piece_jointe_derogation')
        if not motif_derogation or not piece_jointe_derogation:
            messages.error(
                request,
                "Un motif et une pièce jointe justificative sont obligatoires pour valider un "
                "règlement inférieur au montant dû de la tranche primordiale."
            )
            return redirect(redirect_url)
        err = _valider_fichier_upload(piece_jointe_derogation)
        if err:
            messages.error(request, err)
            return redirect(redirect_url)

    groupe_id = uuid.uuid4()
    nb, total, _ = _encaisser_montant_dette(dette, montant, mode, request.user, motif_derogation, piece_jointe_derogation, groupe_id=groupe_id)

    messages.success(
        request,
        f"Paiement de « {dette.frais_formation.type_frais} » encaissé : "
        f"{total:,.0f} FCFA en {nb} versement{'s' if nb > 1 else ''}."
    )
    return redirect(redirect_url)


# ─────────────────────────────────────────────
# ENCAISSER UN MONTANT SUR L'ENSEMBLE DES TYPES DE FRAIS D'UNE INSCRIPTION
# ─────────────────────────────────────────────
@login_required
def stats_encaisser_solde_inscription_view(request, inscription_id):
    inscription = get_object_or_404(
        Inscription.objects.select_related('eleve').prefetch_related('dettes__frais_formation__type_frais'),
        id=inscription_id
    )

    if not _can_access_eleve_finances(request.user, inscription.eleve):
        raise PermissionDenied("Vous n'avez pas accès aux informations financières de cet apprenant.")

    if request.method != 'POST':
        return redirect('courses:stats_dettes_eleve', eleve_id=inscription.eleve_id)

    if not (request.user.is_superuser or request.user.has_perm('courses.encaisser_paiement')):
        raise PermissionDenied("Vous n'avez pas la permission d'encaisser un paiement.")

    redirect_url = f"{reverse('courses:stats_dettes_eleve', args=[inscription.eleve_id])}?inscription={inscription.id}"

    # Le frais de dossier se règle intégralement via son propre bouton
    # « Solder ce frais » — « Solder l'inscription » ne devient utilisable
    # qu'une fois ce frais soldé (cf. bouton grisé côté template).
    dette_dossier_impayee = next(
        (d for d in inscription.dettes.all()
         if d.frais_formation.type_frais.est_frais_de_dossier and d.reste_a_payer() > 0),
        None
    )
    if dette_dossier_impayee:
        messages.error(
            request,
            f"Réglez d'abord entièrement le frais de dossier « {dette_dossier_impayee.frais_formation.type_frais} » "
            "(bouton « Solder ce frais ») avant de pouvoir solder l'inscription."
        )
        return redirect(redirect_url)

    mode = request.POST.get('mode_paiement', 'espece')
    montant_str = request.POST.get('montant_paiement', '').strip()
    try:
        montant = float(montant_str)
    except (ValueError, TypeError):
        messages.error(request, "Montant invalide.")
        return redirect(redirect_url)

    if montant <= 0:
        messages.error(request, "Le montant doit être supérieur à 0.")
        return redirect(redirect_url)

    dettes = list(inscription.dettes.order_by('id'))
    reste_total = sum(max(d.reste_a_payer(), 0) for d in dettes)
    if montant > reste_total:
        messages.error(request, f"Le montant saisi ({montant:,.0f} FCFA) dépasse le reste dû total ({reste_total:,.0f} FCFA).")
        return redirect(redirect_url)

    # La dette bloquante (tranche primordiale non réglée) est traitée en
    # premier, pour respecter le même ordre que l'encaissement tranche par
    # tranche ; le reste du montant cascade ensuite sur les tranches et
    # types de frais suivants.
    dette_bloquante, _ = inscription.dette_et_tranche_bloquantes()
    if dette_bloquante:
        dettes.sort(key=lambda d: 0 if d.id == dette_bloquante.id else 1)

    motif_derogation = request.POST.get('motif_derogation', '').strip() or None
    piece_jointe_derogation = request.FILES.get('piece_jointe_derogation')
    if piece_jointe_derogation:
        err = _valider_fichier_upload(piece_jointe_derogation)
        if err:
            messages.error(request, err)
            return redirect(redirect_url)

    groupe_id = uuid.uuid4()
    try:
        with transaction.atomic():
            restant = montant
            nb_total = 0
            montant_total = 0
            for dette in dettes:
                if restant <= 0:
                    break
                reste_dette = dette.reste_a_payer()
                if reste_dette <= 0:
                    continue

                # Un frais de dossier se règle en une seule fois, intégralement :
                # ni paiement partiel, ni dérogation applicable. Tant qu'il n'est
                # pas soldé, aucune autre dette n'est de toute façon atteinte ici
                # (elle est bloquante — cf. dette_et_tranche_bloquantes).
                if dette.frais_formation.type_frais.est_frais_de_dossier:
                    if restant < reste_dette:
                        raise _CascadeInterrompue(
                            f"Le frais de dossier « {dette.frais_formation.type_frais} » doit être réglé "
                            f"intégralement en un seul versement (montant dû : {reste_dette:,.0f} FCFA)."
                        )
                    nb, total, restant = _encaisser_montant_dette(dette, restant, mode, request.user, groupe_id=groupe_id)
                    nb_total += nb
                    montant_total += total
                    continue

                tranche_cible = dette.tranche_a_payer()
                motif, piece = None, None
                if tranche_cible and tranche_cible.est_primordiale and restant < dette.reste_pour_tranche(tranche_cible):
                    if not motif_derogation or not piece_jointe_derogation:
                        raise _CascadeInterrompue(
                            "Un motif et une pièce jointe justificative sont obligatoires pour valider un "
                            f"règlement inférieur au montant dû de la tranche primordiale de « {dette.frais_formation.type_frais} »."
                        )
                    motif, piece = motif_derogation, piece_jointe_derogation

                nb, total, restant = _encaisser_montant_dette(dette, restant, mode, request.user, motif, piece, groupe_id=groupe_id)
                nb_total += nb
                montant_total += total
    except _CascadeInterrompue as exc:
        messages.error(request, exc.message)
        return redirect(redirect_url)

    if nb_total == 0:
        messages.info(request, "Aucun paiement n'a pu être enregistré.")
    else:
        messages.success(
            request,
            f"Paiement de l'inscription encaissé : {montant_total:,.0f} FCFA en "
            f"{nb_total} versement{'s' if nb_total > 1 else ''}."
        )
    return redirect(redirect_url)


# ─────────────────────────────────────────────
# DÉTAIL D'UNE DETTE (tranches + modal paiement)
# ─────────────────────────────────────────────
@login_required
def stats_detail_dette_view(request, dette_id):
    dette = get_object_or_404(
        Dette.objects.select_related(
            'inscription__eleve',
            'inscription__formation__filiere',
            'inscription__formation__centre',
            'inscription__annee_scolaire',
            'frais_formation__type_frais',
        ).prefetch_related('paiements'),
        id=dette_id
    )
    eleve = dette.inscription.eleve

    if not _can_access_dette_finances(request.user, dette):
        raise PermissionDenied("Vous n'avez pas accès aux informations financières de cette dette.")

    if request.method == 'POST':
        if not (request.user.is_superuser or request.user.has_perm('courses.encaisser_paiement')):
            raise PermissionDenied("Vous n'avez pas la permission d'encaisser un paiement.")

        # Ordre de paiement : la tranche primordiale d'une autre dette de la
        # même inscription doit être intégralement réglée avant celle-ci.
        dette_bloquante, tranche_bloquante = dette.inscription.dette_et_tranche_bloquantes()
        if dette_bloquante and dette_bloquante.id != dette.id:
            messages.error(
                request,
                f"Il faut d'abord régler entièrement {_libelle_blocage(dette_bloquante, tranche_bloquante)}."
            )
            return redirect('courses:stats_detail_dette', dette_id=dette_id)

        montant_str = request.POST.get('montant_paiement', '').strip()
        mode = request.POST.get('mode_paiement', 'espece')

        try:
            montant = float(montant_str)
        except (ValueError, TypeError):
            messages.error(request, "Montant invalide.")
            return redirect('courses:stats_detail_dette', dette_id=dette_id)

        tranche_cible = dette.tranche_a_payer()
        montant_cible = dette.montant_a_payer()

        if montant <= 0:
            messages.error(request, "Le montant doit être supérieur à 0.")
            return redirect('courses:stats_detail_dette', dette_id=dette_id)

        if montant > montant_cible:
            messages.error(request, f"Le montant saisi ({montant:,.0f} FCFA) dépasse le montant dû ({montant_cible:,.0f} FCFA).")
            return redirect('courses:stats_detail_dette', dette_id=dette_id)

        # Un frais de dossier se règle en une seule fois, intégralement : ni
        # paiement partiel, ni notion de tranche/dérogation applicable.
        est_frais_dossier = dette.frais_formation.type_frais.est_frais_de_dossier
        if est_frais_dossier and montant < montant_cible:
            messages.error(
                request,
                f"Le frais de dossier « {dette.frais_formation.type_frais} » doit être réglé intégralement "
                f"en un seul versement (montant dû : {montant_cible:,.0f} FCFA)."
            )
            return redirect('courses:stats_detail_dette', dette_id=dette_id)

        # Sous-paiement d'une tranche primordiale : motif + pièce jointe obligatoires.
        motif_derogation = None
        piece_jointe_derogation = None
        if not est_frais_dossier and tranche_cible and tranche_cible.est_primordiale and montant < montant_cible:
            motif_derogation = request.POST.get('motif_derogation', '').strip()
            piece_jointe_derogation = request.FILES.get('piece_jointe_derogation')
            if not motif_derogation or not piece_jointe_derogation:
                messages.error(
                    request,
                    "Un motif et une pièce jointe justificative sont obligatoires pour valider un "
                    "règlement inférieur au montant dû de la tranche primordiale."
                )
                return redirect('courses:stats_detail_dette', dette_id=dette_id)
            err = _valider_fichier_upload(piece_jointe_derogation)
            if err:
                messages.error(request, err)
                return redirect('courses:stats_detail_dette', dette_id=dette_id)

        tranche_num = dette.paiements.count() + 1

        paiement = Paiement(
            dette=dette,
            montant_paiement=montant,
            mode_paiement=mode,
            tranche=tranche_num,
            tranche_frais=tranche_cible,
            date_paiement=timezone.now(),
            cree_par=request.user,  # ← ici
            motif_derogation=motif_derogation,
            piece_jointe_derogation=piece_jointe_derogation,
            groupe_id=uuid.uuid4(),
        )
        paiement.save()

        # Mettre à jour l'état de la dette si soldée. `dette.paiements` a été
        # préchargée (prefetch_related) avant la création du paiement ci-dessus ;
        # `dette.paiements.all()` réutiliserait ce cache obsolète (sans le
        # paiement qu'on vient de créer), d'où une requête indépendante ici.
        total_paye = Paiement.objects.filter(dette_id=dette.id, annule=False).aggregate(s=Sum('montant_paiement'))['s'] or 0
        if dette.montant_total - total_paye <= 0:
            dette.etat_dette = 'soldé'
            dette.save()

        messages.success(request, f"Paiement de {montant:,.0f} FCFA enregistré — Tranche {tranche_num}.")
        return redirect('courses:stats_detail_dette', dette_id=dette_id)

    paiements = list(
        dette.paiements.select_related('dette__inscription').order_by('-date_paiement', '-tranche')
    )
    for p in paiements:
        p.annulable = (not p.annule) and _est_dernier_versement_inscription(p)
    montant_paye = dette.montant_paye()
    reste = dette.reste_a_payer()
    taux = (montant_paye / dette.montant_total * 100) if dette.montant_total > 0 else 0

    tranche_cible = dette.tranche_a_payer()
    montant_cible = dette.montant_a_payer()
    bloquee = dette.bloquee_par_autre_dette()

    return render(request, 'member/statistiques/stats_detail_dette.html', {
        'dette': dette,
        'eleve': eleve,
        'paiements': paiements,
        'montant_paye': montant_paye,
        'reste': max(reste, 0),
        'taux': min(round(taux, 1), 100),
        'tranches_data': dette.tranches_detail(),
        'tranche_cible': tranche_cible,
        'montant_cible': montant_cible,
        'bloquee': bloquee,
    })


# ─────────────────────────────────────────────
# ANNULATION D'UN PAIEMENT (ou du lot auquel il appartient)
# ─────────────────────────────────────────────
def _paiements_du_lot(paiement):
    """Tous les paiements non annulés créés par la même action que
    `paiement` (même groupe_id — ex. "Solder ce frais"/"Solder
    l'inscription" génère un versement par tranche/frais en un seul clic).
    Un paiement sans groupe_id (créé avant l'introduction des lots) forme
    son propre lot, réduit à lui-même."""
    if paiement.groupe_id:
        return Paiement.objects.filter(
            dette__inscription=paiement.dette.inscription,
            groupe_id=paiement.groupe_id, annule=False,
        )
    return Paiement.objects.filter(pk=paiement.pk, annule=False)


def _est_dernier_versement_inscription(paiement):
    """True si le lot de `paiement` est le versement non annulé le plus
    récent de son inscription — condition nécessaire pour pouvoir l'annuler
    sans casser l'ordre primordiale/frais de dossier (LIFO : on ne défait
    que le dernier mouvement, jamais un versement du milieu). Portée
    limitée à l'inscription : deux apprenants différents ne s'influencent
    jamais l'un l'autre."""
    lot = _paiements_du_lot(paiement)
    lot_ids = set(lot.values_list('pk', flat=True))
    dernier_du_lot = lot.order_by('-date_paiement').first()
    if not dernier_du_lot:
        return False

    plus_recent_ailleurs = (
        Paiement.objects.filter(dette__inscription=paiement.dette.inscription, annule=False)
        .exclude(pk__in=lot_ids)
        .order_by('-date_paiement')
        .first()
    )
    if not plus_recent_ailleurs:
        return True
    return dernier_du_lot.date_paiement >= plus_recent_ailleurs.date_paiement


def _annuler_paiement(paiement, user, motif):
    """Annule le lot entier de `paiement` (transaction.atomic — tout ou
    rien) : chaque ligne reste en base (numéro de quittance jamais
    réutilisé) mais n'est plus comptée dans Dette.montant_paye(), et
    l'état de chaque dette touchée est remis à jour. Retourne le nombre de
    paiements annulés."""
    lot = list(_paiements_du_lot(paiement))
    maintenant = timezone.now()
    dettes_touchees = set()
    with transaction.atomic():
        for p in lot:
            p.annule = True
            p.motif_annulation = motif
            p.annule_par = user
            p.date_annulation = maintenant
            p.save()
            dettes_touchees.add(p.dette_id)
        for dette_id in dettes_touchees:
            dette = Dette.objects.get(pk=dette_id)
            dette.etat_dette = 'soldé' if dette.reste_a_payer() <= 0 else 'non_soldé'
            dette.save()
    return len(lot)


@login_required
def stats_annuler_paiement_view(request, paiement_id):
    paiement = get_object_or_404(
        Paiement.objects.select_related('dette__inscription__eleve', 'dette__frais_formation__type_frais'),
        id=paiement_id
    )
    dette = paiement.dette

    if not _can_access_dette_finances(request.user, dette):
        raise PermissionDenied("Vous n'avez pas accès aux informations financières de cette dette.")

    redirect_url = f"{reverse('courses:stats_dettes_eleve', args=[dette.inscription.eleve_id])}?inscription={dette.inscription_id}"

    # Les deux controles d'autorisation sont evalues avant le filtre sur la
    # methode : un utilisateur sans la permission doit recevoir un 403, y compris
    # en GET. Place apres le test POST, ce controle renvoyait une redirection
    # silencieuse — la mutation restait protegee, mais le refus n'etait ni
    # visible pour l'utilisateur ni journalise.
    if not (request.user.is_superuser or request.user.has_perm('courses.gerer_paiements')):
        raise PermissionDenied("Vous n'avez pas la permission d'annuler un paiement.")

    if request.method != 'POST':
        return redirect(redirect_url)

    if paiement.annule:
        messages.info(request, "Ce paiement est déjà annulé.")
        return redirect(redirect_url)

    motif = request.POST.get('motif_annulation', '').strip()
    if not motif:
        messages.error(request, "Un motif est obligatoire pour annuler un paiement.")
        return redirect(redirect_url)

    if not _est_dernier_versement_inscription(paiement):
        messages.error(
            request,
            "Impossible d'annuler ce versement : des versements plus récents existent sur cette "
            "inscription. Annulez-les d'abord, du plus récent au plus ancien."
        )
        return redirect(redirect_url)

    nb = _annuler_paiement(paiement, request.user, motif)
    messages.success(request, f"Versement annulé ({nb} paiement{'s' if nb > 1 else ''}).")
    return redirect(redirect_url)


# ─────────────────────────────────────────────
# QUITTANCE D'UNE TRANCHE (liste des paiements)
# ─────────────────────────────────────────────
@login_required
def stats_quittance_tranche_view(request, dette_id, tranche):
    dette = get_object_or_404(
        Dette.objects.select_related(
            'inscription__eleve',
            'inscription__formation__filiere',
            'inscription__formation__centre',
            'inscription__annee_scolaire',
            'frais_formation__type_frais',
        ),
        id=dette_id
    )

    if not _can_access_dette_finances(request.user, dette):
        raise PermissionDenied("Vous n'avez pas accès aux informations financières de cette dette.")

    paiements = dette.paiements.filter(tranche=tranche).order_by('date_paiement')
    est_apprenant = getattr(request.user, 'pk', None) == getattr(dette.inscription.eleve, 'pk', None)

    return render(request, 'member/statistiques/stats_quittance_tranche.html', {
        'dette': dette,
        'eleve': dette.inscription.eleve,
        'tranche': tranche,
        'paiements': paiements,
        'est_apprenant': est_apprenant,
    })


# ─────────────────────────────────────────────
# TÉLÉCHARGER QUITTANCE PDF (réutilisable)
# ─────────────────────────────────────────────
@login_required
def stats_download_quittance_view(request, paiement_id):
    paiement = get_object_or_404(
        Paiement.objects.select_related(
            'dette__inscription__eleve',
            'dette__inscription__formation__filiere',
            'dette__inscription__formation__centre',
            'dette__inscription__annee_scolaire',
            'dette__frais_formation__type_frais',
        ),
        id=paiement_id
    )
    dette = paiement.dette
    inscription = dette.inscription
    eleve = inscription.eleve

    if not _can_access_dette_finances(request.user, dette):
        raise PermissionDenied("Vous n'avez pas accès à cette quittance.")

    # L'apprenant ne doit jamais pouvoir télécharger la quittance d'un
    # versement annulé (seul le personnel peut la réimprimer, tamponnée
    # "ANNULÉE", à des fins d'audit — cf. plus bas).
    est_apprenant = getattr(request.user, 'pk', None) == getattr(eleve, 'pk', None)
    if paiement.annule and est_apprenant:
        messages.error(request, "Ce versement a été annulé, sa quittance n'est plus disponible au téléchargement.")
        return redirect('courses:mes_paiements')

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A5)
    width, height = A5

    favicon_path = os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
    _draw_pdf_watermark(p, width, height, favicon_path)
    header_left, header_right = _pdf_header_lines(inscription.formation.centre)
    line_h = 0.28*cm
    y_left = height - 0.6*cm
    p.setFont("Helvetica-Bold", 5.5)
    for line in header_left:
        p.drawString(0.6*cm, y_left, line)
        y_left -= line_h
    y_right = height - 0.6*cm
    for line in header_right:
        p.drawRightString(width-0.6*cm, y_right, line)
        y_right -= line_h
    try:
        p.drawImage(ImageReader(favicon_path), x=width/2-0.9*cm, y=height-2.2*cm,
                    width=1.8*cm, height=1.8*cm, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height-3.8*cm, "QUITTANCE DE PAIEMENT")
    p.setFont("Helvetica", 9)
    p.drawCentredString(width/2, height-4.4*cm, "Burkina Suudu Bawdè")
    if paiement.annule:
        p.saveState()
        p.setFillColor(colors.red)
        p.setFillAlpha(0.35)
        p.setFont("Helvetica-Bold", 34)
        p.translate(width/2, height/2)
        p.rotate(30)
        p.drawCentredString(0, 0, "ANNULÉE")
        p.restoreState()

    y = height - 5.2*cm
    p.setLineWidth(0.8)
    p.line(1.5*cm, y, width-1.5*cm, y)

    # Bascule automatiquement sur plusieurs lignes si la valeur (ex. nom de
    # centre ou de métier à rallonge) dépasse la largeur disponible.
    def ligne(label, valeur, y_pos):
        p.setFont("Helvetica-Bold", 10)
        p.drawString(1.5*cm, y_pos, label)
        p.setFont("Helvetica", 10)
        valeur = str(valeur)
        max_width = (width - 1.5*cm) - 7*cm
        if p.stringWidth(valeur, "Helvetica", 10) <= max_width:
            p.drawString(7*cm, y_pos, valeur)
            return y_pos - 0.5*cm   # <-- était 0.7*cm
        mots = valeur.split()
        lignes, courante = [], ""
        for mot in mots:
            essai = f"{courante} {mot}".strip()
            if p.stringWidth(essai, "Helvetica", 10) <= max_width:
                courante = essai
            else:
                if courante:
                    lignes.append(courante)
                courante = mot
        if courante:
            lignes.append(courante)
        line_h = 0.42*cm
        for i, texte in enumerate(lignes):
            p.drawString(7*cm, y_pos - i*line_h, texte)
        return y_pos - len(lignes)*line_h - 0.1*cm

    y -= 0.5*cm
    y = ligne("Numéro de quittance :", paiement.numero_quittance, y)
    y = ligne("Date de paiement :", paiement.date_paiement.strftime("%d/%m/%Y à %H:%M"), y)
    y -= 0.3*cm
    p.setDash(3, 3); p.line(1.5*cm, y, width-1.5*cm, y); p.setDash(); y -= 0.5*cm

    y = ligne("Apprenant :", f"{eleve.nom} {eleve.prenom}", y)
    y = ligne("Matricule :", eleve.matricule or "—", y)
    y = ligne("Centre :", str(inscription.formation.centre), y)
    y = ligne("Métier :", str(inscription.formation.filiere), y)
    y = ligne("Année de formation :", str(inscription.annee_scolaire or "—"), y)
    y -= 0.3*cm
    p.setDash(3, 3); p.line(1.5*cm, y, width-1.5*cm, y); p.setDash(); y -= 0.5*cm

    tranche_label = paiement.tranche_frais.libelle if paiement.tranche_frais else f"Tranche {paiement.tranche}"
    y = ligne("Type de frais :", str(dette.frais_formation.type_frais.libelle), y)
    y = ligne("Tranche :", tranche_label, y)
    y = ligne("Mode de paiement :", paiement.get_mode_paiement_display(), y)

    p.setFont("Helvetica-Bold", 12)
    p.drawString(1.5*cm, y, "Montant payé :")
    p.drawString(7*cm, y, f"{paiement.montant_paiement:,.0f} FCFA")
    y -= 0.7*cm; y -= 0.3*cm
    p.setDash(3, 3); p.line(1.5*cm, y, width-1.5*cm, y); p.setDash(); y -= 0.5*cm

    y = ligne("Total dû :", f"{dette.montant_total:,.0f} FCFA", y)
    y = ligne("Total payé :", f"{dette.montant_paye():,.0f} FCFA", y)
    y = ligne("Reste à payer :", f"{dette.reste_a_payer():,.0f} FCFA", y)
    y = ligne("État de la dette :", dette.get_etat_dette_display(), y)

    # QR Code — identifiant élève volontairement absent de la quittance (et du
    # QR) : ce document circule hors de la plateforme et n'a pas à exposer cet
    # identifiant. Position calculée à partir de la fin du texte (et non plus
    # une valeur fixe) pour ne jamais chevaucher les lignes ci-dessus, même si
    # certaines ont débordé sur plusieurs lignes.
    qr_data = (
        f"Quittance : {paiement.numero_quittance}\n"
        f"Date : {paiement.date_paiement.strftime('%d/%m/%Y à %H:%M')}\n"
        f"Apprenant : {eleve.nom} {eleve.prenom}\n"
        f"Centre : {inscription.formation.centre}\n"
        f"Métier : {inscription.formation.filiere}\n"
        f"Montant payé : {paiement.montant_paiement:,.0f} FCFA\n"
        f"Reste à payer : {dette.reste_a_payer():,.0f} FCFA"
    )
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)

    qr_size = 3*cm
    qr_y = max(y - 0.3*cm - qr_size, 0.9*cm)
    p.drawImage(ImageReader(qr_buffer), x=(width-qr_size)/2, y=qr_y,
                width=qr_size, height=qr_size)
    p.setFont("Helvetica-Oblique", 7)
    p.setFillColor(colors.grey)
    p.drawCentredString(width/2, qr_y - 0.25*cm, "Scannez pour vérifier")
    # Pied de page à droite, en petit, pour ne pas chevaucher le QR code centré
    p.setFont("Helvetica-Oblique", 6)
    p.drawRightString(width - 1.5*cm, max(qr_y - 0.65*cm, 0.3*cm),
                       f"BSB — généré sur YU-PAAN le : {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="quittance_{paiement.numero_quittance}.pdf"'
    )
    return response

# courses/views/center_views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.db.models import Q

from courses.models import CentreFormation, Filiere, CentreEtFiliere, Direction_reg
from courses.forms import CentreFormationForm


# ── LIST ──────────────────────────────────────────────────────────────────────
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views import View
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q

from courses.models import CentreFormation, Direction_reg


@method_decorator(login_required, name='dispatch')
@method_decorator(require_permission('courses.gerer_centres'), name='dispatch')
class CenterListView(View):
    template_name = 'admin/center/list.html'

    def get(self, request):
        q = request.GET.get('q', '').strip()
        direction_id = request.GET.get('direction', '').strip()
        niveau = request.GET.get('niveau', '').strip()

        qs = (
            CentreFormation.objects
            .filter(pk__in=_get_scope(request.user)[0])
            .select_related('direction', 'province')
            .order_by('nom_centre')
        )

        # 🔍 SEARCH GLOBAL
        if q:
            qs = qs.filter(
                Q(nom_centre__icontains=q) |
                Q(adresse__icontains=q) |
                Q(province__nom_province__icontains=q)  # ← adapter au nom exact du champ sur Province
            ).distinct()

        # 🏢 FILTER DIRECTION (SAFE)
        if direction_id and direction_id.isdigit():
            qs = qs.filter(direction_id=int(direction_id))

        # 🎯 FILTER NIVEAU (SAFE - IMPORTANT FIX)
        if niveau and niveau.isdigit():
            qs = qs.filter(niveau_centre=int(niveau))
        else:
            niveau = ''  # ← reset pour éviter qu'une valeur invalide soit renvoyée au template

        # 📄 PAGINATION SAFE
        try:
            page_number = int(request.GET.get('page', 1))
            if page_number < 1:
                page_number = 1
        except (ValueError, TypeError):
            page_number = 1

        paginator = Paginator(qs, 10)
        centers = paginator.get_page(page_number)

        # 📌 LISTE DIRECTIONS
        directions = Direction_reg.objects.all().order_by('nom_direction')

        # 📌 LISTE NIVEAUX PROPRES (sans NULL)
        niveaux = (
            CentreFormation.objects
            .exclude(niveau_centre__isnull=True)
            .values_list('niveau_centre', flat=True)
            .distinct()
            .order_by('niveau_centre')
        )

        return render(request, self.template_name, {
            'centers': centers,
            'q': q,
            'direction_id': direction_id,
            'niveau': niveau,
            'directions': directions,
            'niveaux': niveaux,
        })

# ── CREATE ────────────────────────────────────────────────────────────────────
@method_decorator(require_permission('courses.gerer_centres'), name='dispatch')
class CenterCreateView(View):
    template_name = 'admin/center/form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form':     CentreFormationForm(direction_queryset=_get_scope(request.user)[1]),
            #'filieres': Filiere.objects.all().order_by('nom_filiere'),
            'title':    'Créer un centre',
            'action':   'Créer',
        })

    def post(self, request):
        form = CentreFormationForm(request.POST, direction_queryset=_get_scope(request.user)[1])
        if form.is_valid():
            centre = form.save()
            messages.success(request, f'Le centre « {centre.nom_centre} » a été créé avec succès.')
            next_url = request.POST.get('next')
            if next_url and url_has_allowed_host_and_scheme(
                next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
            ):
                return redirect(next_url)
            return redirect('bsb_admin:center_list')
        return render(request, self.template_name, {
            'form': form, 'title': 'Créer un centre', 'action': 'Créer',
        })


# ── UPDATE ────────────────────────────────────────────────────────────────────
@method_decorator(require_permission('courses.gerer_centres'), name='dispatch')
class CenterUpdateView(View):
    template_name = 'admin/center/form.html'

    def get(self, request, pk):
        centre   = get_object_or_404(CentreFormation, pk=pk, pk__in=_get_scope(request.user)[0])
        filieres = Filiere.objects.all().order_by('nom_filiere')
        current_filieres = list(
            CentreEtFiliere.objects.filter(centre=centre).values_list('filiere_id', flat=True)
        )
        return render(request, self.template_name, {
            'form':   CentreFormationForm(instance=centre, direction_queryset=_get_scope(request.user)[1]),
            'title':  f'Modifier — {centre.nom_centre}',
            'action': 'Modifier',
        })

    def post(self, request, pk):
        centre   = get_object_or_404(CentreFormation, pk=pk, pk__in=_get_scope(request.user)[0])
        form     = CentreFormationForm(request.POST, instance=centre, direction_queryset=_get_scope(request.user)[1])
        filieres = Filiere.objects.all().order_by('nom_filiere')
        if form.is_valid():
            centre = form.save()
            messages.success(request, f'Le centre « {centre.nom_centre} » a été modifié avec succès.')
            next_url = request.POST.get('next')
            if next_url and url_has_allowed_host_and_scheme(
                next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
            ):
                return redirect(next_url)
            return redirect('bsb_admin:center_list')
        return render(request, self.template_name, {
            'form': form, 'title': f'Modifier — {centre.nom_centre}', 'action': 'Modifier',
        })


# ── DELETE ────────────────────────────────────────────────────────────────────
@method_decorator(require_permission('courses.gerer_centres'), name='dispatch')
class CenterDeleteView(View):
    template_name = 'admin/center/confirm_delete.html'

    def get(self, request, pk):
        centre = get_object_or_404(CentreFormation, pk=pk, pk__in=_get_scope(request.user)[0])
        return render(request, self.template_name, {'object': centre})

    def post(self, request, pk):
        centre = get_object_or_404(CentreFormation, pk=pk, pk__in=_get_scope(request.user)[0])
        nom    = centre.nom_centre
        centre.delete()
        messages.success(request, f'Le centre « {nom} » a été supprimé définitivement.')
        return redirect('bsb_admin:center_list')
    
    # ── CENTER FILIERES DETAIL ────────────────────────────────────────────────────
@method_decorator(login_required, name='dispatch')
class CenterFiliereListView(View):
    template_name = 'admin/center/filieres.html'

    def get(self, request, pk):
        centre = get_object_or_404(CentreFormation, pk=pk)
        centre_filieres = (
            CentreEtFiliere.objects
            .filter(centre=centre)
            .select_related('filiere', 'annee_prog')
            .order_by('filiere__nom_filiere')
        )
        paginator = Paginator(centre_filieres, 10)
        page_obj  = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'centre':     centre,
            'page_obj':   page_obj,
        })
        
from django.http import HttpResponse
from django.db.models import Count, Sum, Q
import csv
import io

@require_role('formateur')
def formateur_dashboard(request):
    try:
        formateur = request.user.formateur
    except Exception:
        messages.error(request, "Profil formateur introuvable.")
        return redirect('accounts:login')

    centre = formateur.centre

    # Formations actives dans son centre pour son métier
    if formateur.filiere_id:
        formations = CentreEtFiliere.objects.filter(
            centre=centre,
            filiere_id=formateur.filiere_id,
            is_active=True
        ).select_related('filiere', 'annee_prog').prefetch_related('frais_set')
    else:
        formations = CentreEtFiliere.objects.none()

    # Stats globales
    total_inscrits = Inscription.objects.filter(
        formation__in=formations
    ).count()

    total_valides = Inscription.objects.filter(
        formation__in=formations,
        statut__in=['valide', 'valide_paye', 'Valide']
    ).count()

    # Vraiement inscrits = ont payé au moins quelque chose
    total_vrais = Inscription.objects.filter(
        formation__in=formations,
        dettes__paiements__isnull=False
    ).distinct().count()

    # Stats par filière
    stats_filieres = []
    for formation in formations:
        inscrits = Inscription.objects.filter(formation=formation)
        vrais = inscrits.filter(dettes__paiements__isnull=False).distinct()
        stats_filieres.append({
            'formation': formation,
            'total_inscrits': inscrits.count(),
            'total_valides': inscrits.filter(
                statut__in=['valide', 'valide_paye', 'Valide']
            ).count(),
            'vrais_inscrits': vrais.count(),
            'total_encaisse': Paiement.objects.filter(
                dette__inscription__formation=formation, annule=False
            ).aggregate(s=Sum('montant_paiement'))['s'] or 0,
        })

    context = {
        'formateur': formateur,
        'centre': centre,
        'stats_filieres': stats_filieres,
        'total_inscrits': total_inscrits,
        'total_valides': total_valides,
        'total_vrais': total_vrais,
        'total_formations': formations.count(),
    }
    return render(request, 'teacher/dashboard/dashboard/dashboard.html', context)


@require_role('formateur')
def formateur_filieres(request):
    try:
        formateur = request.user.formateur
    except Exception:
        return redirect('accounts:login')

    if formateur.filiere_id:
        formations = CentreEtFiliere.objects.filter(
            centre=formateur.centre,
            filiere_id=formateur.filiere_id,
        ).select_related('filiere', 'annee_prog').order_by('filiere__nom_filiere')
    else:
        formations = CentreEtFiliere.objects.none()

    paginator = Paginator(formations, 10)
    page = request.GET.get('page')
    formations = paginator.get_page(page)

    return render(request, 'teacher/filieres/list.html', {
        'formations': formations,
        'formateur': formateur,
    })


@require_role('formateur')
def formateur_etudiants(request, formation_id):
    try:
        formateur = request.user.formateur
    except Exception:
        return redirect('accounts:login')

    formation = get_object_or_404(
        CentreEtFiliere,
        id=formation_id,
        centre=formateur.centre,
        filiere_id=formateur.filiere_id
    )

    # Filtres
    statut_filter = request.GET.get('statut', '')  # 'vrais' | 'valide' | '' (tous)
    q = request.GET.get('q', '').strip()

    inscriptions = Inscription.objects.filter(
        formation=formation
    ).select_related(
        'eleve', 'annee_scolaire'
    ).prefetch_related(
        'dettes__paiements'
    ).order_by('eleve__nom', 'eleve__prenom')

    if q:
        inscriptions = inscriptions.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__numero_identifiant__icontains=q)
        )

    if statut_filter == 'vrais':
        inscriptions = inscriptions.filter(
            dettes__paiements__isnull=False
        ).distinct()
    elif statut_filter == 'valide':
        inscriptions = inscriptions.filter(
            statut__in=['valide', 'valide_paye', 'Valide']
        )

    # Enrichissement pour affichage
    inscrits_data = []
    for insc in inscriptions:
        total_du = sum(d.montant_total for d in insc.dettes.all())
        total_paye = sum(
            p.montant_paiement
            for d in insc.dettes.all()
            for p in d.paiements.all()
            if not p.annule
        )
        inscrits_data.append({
            'inscription': insc,
            'eleve': insc.eleve,
            'total_du': total_du,
            'total_paye': total_paye,
            'reste': total_du - total_paye,
            'a_paye': total_paye > 0,
        })

    paginator = Paginator(inscrits_data, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'formation': formation,
        'formateur': formateur,
        'page_obj': page_obj,
        'inscrits_data': inscrits_data,
        'statut_filter': statut_filter,
        'q': q,
        'total_tous': inscriptions.count(),
        'total_vrais': sum(1 for d in inscrits_data if d['a_paye']),
    }
    return render(request, 'teacher/filieres/etudiants.html', context)


@require_role('formateur')
def formateur_export(request, formation_id, format):
    try:
        formateur = request.user.formateur
    except Exception:
        return redirect('accounts:login')

    formation = get_object_or_404(
        CentreEtFiliere,
        id=formation_id,
        centre=formateur.centre,
        filiere_id=formateur.filiere_id
    )

    statut_filter = request.GET.get('statut', '')
    q = request.GET.get('q', '').strip()

    inscriptions = Inscription.objects.filter(
        formation=formation
    ).select_related('eleve', 'annee_scolaire').prefetch_related('dettes__paiements')

    if q:
        inscriptions = inscriptions.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__numero_identifiant__icontains=q)
        )
    if statut_filter == 'vrais':
        inscriptions = inscriptions.filter(
            dettes__paiements__isnull=False
        ).distinct()
    elif statut_filter == 'valide':
        inscriptions = inscriptions.filter(
            statut__in=['valide', 'valide_paye', 'Valide']
        )

    # Préparer les données
    rows = []
    for i, insc in enumerate(inscriptions, 1):
        total_du = sum(d.montant_total for d in insc.dettes.all())
        total_paye = sum(
            p.montant_paiement for d in insc.dettes.all() for p in d.paiements.all() if not p.annule
        )
        rows.append({
            'N°': i,
            'Matricule': insc.eleve.matricule or '—',
            'Nom': insc.eleve.nom,
            'Prénom': insc.eleve.prenom,
            'Sexe': insc.eleve.get_sexe_display() if hasattr(insc.eleve, 'get_sexe_display') else insc.eleve.sexe,
            'Téléphone': insc.eleve.tel or '—',
            'Email': insc.eleve.email or '—',
            'Statut inscription': insc.get_statut_display(),
            'Total dû (FCFA)': total_du,
            'Total payé (FCFA)': total_paye,
            'Reste (FCFA)': total_du - total_paye,
            'A payé': 'Oui' if total_paye > 0 else 'Non',
        })

    label = f"{formation.filiere.nom_filiere}_{formation.centre.nom_centre}"

    # En-têtes fixes (mêmes noms de colonnes que les données), écrites
    # inconditionnellement — même logique que les exports de statistiques.
    headers = [
        'N°', 'Matricule', 'Nom', 'Prénom', 'Sexe', 'Téléphone', 'Email',
        'Statut inscription', 'Total dû (FCFA)', 'Total payé (FCFA)', 'Reste (FCFA)', 'A payé',
    ]

    # ── CSV ──────────────────────────────────────────────────────────────────
    if format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="apprenants_{label}.csv"'
        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        return response

    # ── XLSX ─────────────────────────────────────────────────────────────────
    elif format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messages.error(request, "openpyxl non installé. Lancez : pip install openpyxl")
            return redirect('courses:formateur_etudiants', formation_id=formation_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Apprenants"

        # En-tête titre
        titre = f"Liste des apprenants — {formation.filiere.nom_filiere} — {formation.centre.nom_centre}"
        ws.merge_cells('A1:L1')
        ws['A1'] = titre
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = Alignment(horizontal='center')

        # En-têtes colonnes — écrites inconditionnellement, qu'il y ait ou non des lignes.
        header_fill = PatternFill("solid", fgColor="D4A017")
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append([])  # ligne vide après titre
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Données
        for row in rows:
            ws.append(list(row.values()))
            data_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='left')
                # Colorier en vert si a payé
                if headers[col_idx - 1] == 'A payé' and cell.value == 'Oui':
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif headers[col_idx - 1] == 'A payé' and cell.value == 'Non':
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

        # Largeur colonnes auto
        for col in ws.columns:
            max_len = max(
                (len(str(c.value or '')) for c in col),
                default=10
            )
            first_valid_cell = next(
                (cell for cell in col if not isinstance(cell, MergedCell)),
                None
            )
            if first_valid_cell:
                column_letter = get_column_letter(first_valid_cell.column)
                ws.column_dimensions[column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="apprenants_{label}.xlsx"'
        return response

    # ── PDF ──────────────────────────────────────────────────────────────────
    elif format == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm
        )
        styles = getSampleStyleSheet()
        elements = []

        # En-tête (logo centré, petit)
        favicon_path = os.path.join(settings.BASE_DIR, 'static/images/favicon.png')
        header_line_style = ParagraphStyle(
            'header_line', parent=styles['Normal'], fontSize=6, leading=8,
            alignment=1, fontName='Helvetica-Bold',
        )
        header_left_fe, header_right_fe = _pdf_header_lines(formation.centre)
        header_table = Table(
            [[
                Paragraph('<br/>'.join(header_left_fe), header_line_style),
                Image(favicon_path, width=1.6*cm, height=1.6*cm),
                Paragraph('<br/>'.join(header_right_fe), header_line_style),
            ]],
            colWidths=[10*cm, 3*cm, 10*cm],
        )
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 12))

        # Titre
        title_style = ParagraphStyle(
            'title', parent=styles['Heading1'],
            fontSize=14, spaceAfter=6, alignment=1
        )
        sub_style = ParagraphStyle(
            'sub', parent=styles['Normal'],
            fontSize=9, spaceAfter=12, alignment=1, textColor=rl_colors.grey
        )
        elements.append(Paragraph(
            f"Liste des apprenants — {formation.filiere.nom_filiere}", title_style
        ))
        elements.append(Paragraph(
            f"Centre : {formation.centre.nom_centre} | Filtre : {statut_filter or 'Tous'} | Total : {len(rows)}",
            sub_style
        ))
        elements.append(Spacer(1, 0.3*cm))

        data = [headers] + [list(r.values()) for r in rows]
        if not rows:
            elements.append(Paragraph("Aucun apprenant trouvé.", styles['Normal']))
        col_count = len(headers)
        page_w = landscape(A4)[0] - 3*cm
        col_w = page_w / col_count

        table = Table(data, colWidths=[col_w] * col_count, repeatRows=1)
        table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#D4A017')),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F9FAFB')]),
                ('GRID', (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#E5E7EB')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

        signature_style = ParagraphStyle(
            'signature', parent=styles['Normal'], fontSize=10,
            alignment=2, spaceBefore=28,
        )
        elements.append(Paragraph(f"Le Formateur — {formateur.nom} {formateur.prenom}", signature_style))

        footer_style_left = ParagraphStyle(
            'footer_bsb_left', parent=styles['Normal'], fontSize=7,
            textColor=rl_colors.grey, alignment=0,
        )
        footer_style_right = ParagraphStyle(
            'footer_bsb_right', parent=styles['Normal'], fontSize=7,
            textColor=rl_colors.grey, alignment=2,
        )
        footer_table = Table(
            [[
                Paragraph("BSB", footer_style_left),
                Paragraph(f"généré sur YU-PAAN le : {timezone.now().strftime('%d/%m/%Y à %H:%M')}", footer_style_right),
            ]],
            colWidths=[doc.width / 2, doc.width / 2],
        )
        footer_table.setStyle(TableStyle([
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(Spacer(1, 24))
        elements.append(footer_table)

        def _watermark_page_fe(canvas_obj, doc_obj):
            _draw_pdf_watermark(canvas_obj, doc_obj.pagesize[0], doc_obj.pagesize[1], favicon_path)

        doc.build(elements, onFirstPage=_watermark_page_fe, onLaterPages=_watermark_page_fe)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="apprenants_{label}.pdf"'
        return response

    return redirect('courses:formateur_etudiants', formation_id=formation_id)

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count

# Adapte ces imports selon ton app
from .models import Region, Province


# ──────────────────────────────────────────────
#  RÉGION — Liste + filtres
# ──────────────────────────────────────────────

@login_required
def region_list(request):
    """Liste des régions avec leurs provinces, filtres et pagination."""
    q = request.GET.get("q", "").strip()
    chef_lieu_filter = request.GET.get("chef_lieu", "").strip()

    regions = Region.objects.annotate(nb_provinces=Count("provinces")).order_by("nom_region")

    if q:
        regions = regions.filter(
            Q(nom_region__icontains=q) | Q(chef_lieu__icontains=q)
        )
    if chef_lieu_filter:
        regions = regions.filter(chef_lieu__icontains=chef_lieu_filter)

    paginator = Paginator(regions, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Pour le filtre chef-lieu (valeurs distinctes)
    chefs_lieux = Region.objects.values_list("chef_lieu", flat=True).distinct().order_by("chef_lieu")

    return render(request, "admin/region/region_list.html", {
        "regions": page_obj,
        "q": q,
        "chef_lieu_filter": chef_lieu_filter,
        "chefs_lieux": chefs_lieux,
    })


# ──────────────────────────────────────────────
#  RÉGION — Créer / Modifier
# ──────────────────────────────────────────────

@require_permission('courses.gerer_regions')
def region_create(request):
    if request.method == "POST":
        nom = request.POST.get("nom_region", "").strip()
        chef_lieu = request.POST.get("chef_lieu", "").strip()
        errors = {}

        if not nom:
            errors["nom_region"] = "Le nom de la région est obligatoire."
        elif Region.objects.filter(nom_region__iexact=nom).exists():
            errors["nom_region"] = "Une région avec ce nom existe déjà."
        if not chef_lieu:
            errors["chef_lieu"] = "Le chef-lieu est obligatoire."

        if not errors:
            Region.objects.create(nom_region=nom, chef_lieu=chef_lieu)
            messages.success(request, f"Région « {nom} » créée avec succès.")
            return redirect("bsb_admin:region_list")

        return render(request, "admin/region/region_form.html", {
            "errors": errors,
            "values": request.POST,
            "action": "Créer",
            "title": "Nouvelle région",
        })

    return render(request, "admin/region/region_form.html", {
        "action": "Créer",
        "title": "Nouvelle région",
    })


@require_permission('courses.gerer_regions')
def region_update(request, pk):
    region = get_object_or_404(Region, pk=pk)

    if request.method == "POST":
        nom = request.POST.get("nom_region", "").strip()
        chef_lieu = request.POST.get("chef_lieu", "").strip()
        errors = {}

        if not nom:
            errors["nom_region"] = "Le nom de la région est obligatoire."
        elif Region.objects.filter(nom_region__iexact=nom).exclude(pk=pk).exists():
            errors["nom_region"] = "Une autre région porte déjà ce nom."
        if not chef_lieu:
            errors["chef_lieu"] = "Le chef-lieu est obligatoire."

        if not errors:
            region.nom_region = nom
            region.chef_lieu = chef_lieu
            region.save()
            messages.success(request, f"Région « {nom} » modifiée avec succès.")
            return redirect("bsb_admin:region_list")

        return render(request, "admin/region/region_form.html", {
            "errors": errors,
            "values": request.POST,
            "region": region,
            "action": "Modifier",
            "title": f"Modifier — {region.nom_region}",
        })

    return render(request, "admin/region/region_form.html", {
        "region": region,
        "action": "Modifier",
        "title": f"Modifier — {region.nom_region}",
    })


@require_permission('courses.gerer_regions')
def region_delete(request, pk):
    region = get_object_or_404(Region, pk=pk)
    if request.method == "POST":
        nom = region.nom_region
        region.delete()
        messages.success(request, f"Région « {nom} » supprimée.")
        return redirect("bsb_admin:region_list")
    return render(request, "admin/region/region_confirm_delete.html", {"region": region})


# ──────────────────────────────────────────────
#  PROVINCE — Créer / Modifier / Supprimer
# ──────────────────────────────────────────────

@require_permission('courses.gerer_regions')
def province_create(request):
    regions = Region.objects.order_by("nom_region")

    if request.method == "POST":
        nom = request.POST.get("nom_province", "").strip()
        chef_lieu = request.POST.get("chef_lieu", "").strip()
        region_id = request.POST.get("region", "").strip()
        errors = {}

        if not nom:
            errors["nom_province"] = "Le nom de la province est obligatoire."
        elif Province.objects.filter(nom_province__iexact=nom).exists():
            errors["nom_province"] = "Une province avec ce nom existe déjà."
        if not chef_lieu:
            errors["chef_lieu"] = "Le chef-lieu est obligatoire."
        if not region_id:
            errors["region"] = "Veuillez sélectionner une région."

        if not errors:
            region = get_object_or_404(Region, pk=region_id)
            Province.objects.create(nom_province=nom, chef_lieu=chef_lieu, region=region)
            messages.success(request, f"Province « {nom} » créée avec succès.")
            return redirect("bsb_admin:region_list")

        return render(request, "admin/region/province_form.html", {
            "errors": errors,
            "values": request.POST,
            "regions": regions,
            "action": "Créer",
            "title": "Nouvelle province",
        })

    # Pré-sélection région si passée en GET
    region_id = request.GET.get("region")
    return render(request, "admin/region/province_form.html", {
        "regions": regions,
        "preselected_region": region_id,
        "action": "Créer",
        "title": "Nouvelle province",
    })


# ─── Import Excel/CSV — Region et Province ─────────────────────────────────

@require_permission('courses.gerer_regions')
def region_import_template(request):
    from .bulk_import_registry import SPEC_REGION
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_REGION)


@require_permission('courses.gerer_regions')
def region_import(request):
    from .bulk_import_registry import SPEC_REGION
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_REGION)


@require_permission('courses.gerer_regions')
def province_import_template(request):
    from .bulk_import_registry import SPEC_PROVINCE
    from .bulk_import.views_helpers import render_import_template
    return render_import_template(request, SPEC_PROVINCE)


@require_permission('courses.gerer_regions')
def province_import(request):
    from .bulk_import_registry import SPEC_PROVINCE
    from .bulk_import.views_helpers import handle_import_upload
    return handle_import_upload(request, SPEC_PROVINCE)


@require_permission('courses.gerer_regions')
def province_update(request, pk):
    province = get_object_or_404(Province, pk=pk)
    regions = Region.objects.order_by("nom_region")

    if request.method == "POST":
        nom = request.POST.get("nom_province", "").strip()
        chef_lieu = request.POST.get("chef_lieu", "").strip()
        region_id = request.POST.get("region", "").strip()
        errors = {}

        if not nom:
            errors["nom_province"] = "Le nom de la province est obligatoire."
        elif Province.objects.filter(nom_province__iexact=nom).exclude(pk=pk).exists():
            errors["nom_province"] = "Une autre province porte déjà ce nom."
        if not chef_lieu:
            errors["chef_lieu"] = "Le chef-lieu est obligatoire."
        if not region_id:
            errors["region"] = "Veuillez sélectionner une région."

        if not errors:
            province.nom_province = nom
            province.chef_lieu = chef_lieu
            province.region = get_object_or_404(Region, pk=region_id)
            province.save()
            messages.success(request, f"Province « {nom} » modifiée avec succès.")
            return redirect("bsb_admin:region_list")

        return render(request, "admin/region/province_form.html", {
            "errors": errors,
            "values": request.POST,
            "province": province,
            "regions": regions,
            "action": "Modifier",
            "title": f"Modifier — {province.nom_province}",
        })

    return render(request, "admin/region/province_form.html", {
        "province": province,
        "regions": regions,
        "action": "Modifier",
        "title": f"Modifier — {province.nom_province}",
    })


@require_permission('courses.gerer_regions')
def province_delete(request, pk):
    province = get_object_or_404(Province, pk=pk)
    if request.method == "POST":
        nom = province.nom_province
        province.delete()
        messages.success(request, f"Province « {nom} » supprimée.")
        return redirect("bsb_admin:region_list")
    return render(request, "admin/region/province_confirm_delete.html", {"province": province})


@login_required
def page_notifications(request):
    inscriptions_notif = Inscription.objects.filter(
        eleve=request.user,
        statut__in=["valide", "rejete"]
    ).select_related('formation__filiere', 'formation').order_by('-date_validation')

    vues = request.session.get('notifs_vues', [])
    tous_ids = list(inscriptions_notif.values_list('id', flat=True))
    request.session['notifs_vues'] = list(set(vues + tous_ids))
    request.session.modified = True

    notifications = []
    for inscription in inscriptions_notif:
        if inscription.statut == "valide":
            total_frais = Frais.objects.filter(
                formation=inscription.formation
            ).aggregate(
                total=Sum("montant")
            )["total"] or 0

            # format_html et non f-string : le message est rendu tel quel dans
            # le template (balises <strong> volontaires), donc tout ce qui vient
            # de la base doit etre echappe. `motif_rejet` en particulier est du
            # texte libre saisi par un agent - sans echappement, c'est un XSS
            # stocke de l'agent vers l'apprenant.
            if total_frais:
                message = format_html(
                    "✅ Félicitations ! Votre dossier d'inscription à la formation "
                    "<strong>{}</strong> a été <strong>validé</strong>. "
                    "Pour finaliser votre inscription, vous devez payer "
                    "<strong>75% du montant de la formation</strong>, soit <strong>{} FCFA</strong>. "
                    "Rendez-vous dans la section <em>Mes inscriptions</em> pour procéder au paiement.",
                    inscription.formation.filiere,
                    # Pre-formate : format_html convertit chaque argument en
                    # chaine avant de l'inserer, une spec numerique ({:,.0f})
                    # y echouerait.
                    f"{total_frais * 0.75:,.0f}",
                )
            else:
                message = format_html(
                    "✅ Félicitations ! Votre dossier d'inscription à la formation "
                    "<strong>{}</strong> a été <strong>validé</strong>. "
                    "Rendez-vous dans la section <em>Mes inscriptions</em> pour procéder "
                    "au paiement (75% du montant dû).",
                    inscription.formation.filiere,
                )

        elif inscription.statut == "rejete":
            message = format_html(
                "❌ Votre dossier d'inscription à la formation "
                "<strong>{}</strong> a été <strong>rejeté</strong>. "
                "<br><strong>Motif :</strong> {}",
                inscription.formation.filiere,
                inscription.motif_rejet or "Aucun motif précisé.",
            )

        notifications.append({
            "inscription": inscription,
            "message": message,
            "is_new": inscription.id not in vues,
        })

    return render(request, 'student/notifications.html', {
        'notifications': notifications,
    })
    
    
@login_required
def notifications_count(request):
    """Retourne le nombre de notifications non vues pour la cloche."""
    from django.http import JsonResponse
    
    vues = request.session.get('notifs_vues', [])
    count = Inscription.objects.filter(
        eleve=request.user,
        statut__in=["valide", "rejete"]
    ).exclude(id__in=vues).count()
    
    return JsonResponse({'count': count})