"""Lecture d'un fichier uploadé (.xlsx ou .csv) en lignes indexées par
en-tête de colonne (le nom exact déclaré dans l'ImportSpec, pas le libellé
brut trouvé dans le fichier — la normalisation absorbe la casse/les espaces
et le suffixe " *" ajouté aux colonnes obligatoires dans le modèle généré)."""

import csv
import io


def _normalize_header(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith("*"):
        s = s[:-1].strip()
    return s.casefold()


def _column_positions(header_cells, spec):
    header_index = {}
    for idx, h in enumerate(header_cells or []):
        header_index[_normalize_header(h)] = idx
    positions = {}
    for col in spec.columns:
        pos = header_index.get(_normalize_header(col.header))
        if pos is not None:
            positions[col.header] = pos
    return positions


def _read_xlsx_rows(uploaded_file, spec):
    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(uploaded_file.read()), data_only=True, read_only=True
    )

    ws = None
    for name in wb.sheetnames:
        if name.casefold() == spec.sheet_name.casefold():
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if name.casefold() != "instructions":
                ws = wb[name]
                break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        return

    positions = _column_positions(header_cells, spec)

    row_number = 1
    for cells in rows_iter:
        row_number += 1
        if not cells or not any(c not in (None, "") for c in cells):
            continue
        row_dict = {}
        for header, pos in positions.items():
            row_dict[header] = cells[pos] if pos < len(cells) else None
        yield row_number, row_dict


def _read_csv_rows(uploaded_file, spec):
    text = uploaded_file.read().decode("utf-8-sig")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.reader(io.StringIO(text), dialect)
    rows_iter = iter(reader)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        return

    positions = _column_positions(header_cells, spec)

    row_number = 1
    for cells in rows_iter:
        row_number += 1
        if not any((c or "").strip() for c in cells):
            continue
        row_dict = {}
        for header, pos in positions.items():
            row_dict[header] = cells[pos] if pos < len(cells) else ""
        yield row_number, row_dict


def read_rows(uploaded_file, spec):
    """Retourne une liste de (numero_ligne_excel, {header: valeur_brute})."""
    name = (uploaded_file.name or "").lower()
    if name.endswith(".xlsx"):
        return list(_read_xlsx_rows(uploaded_file, spec))
    if name.endswith(".csv"):
        return list(_read_csv_rows(uploaded_file, spec))
    raise ValueError("Format de fichier non pris en charge : utilisez un fichier .xlsx ou .csv.")
