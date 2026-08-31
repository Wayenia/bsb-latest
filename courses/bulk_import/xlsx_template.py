"""Génération du fichier Excel "modèle" vierge pour un ImportSpec — reprend
la palette déjà utilisée par l'export Excel existant (courses/views.py,
fonction export_excel) pour rester visuellement cohérent."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROUGE = "C0392B"
OR = "D4A017"

_KIND_LABELS = {
    "text": "Texte",
    "int": "Nombre entier",
    "decimal": "Nombre décimal",
    "bool": "Oui / Non",
    "date": "Date (JJ/MM/AAAA)",
    "choice_static": "Choix",
    "fk_pk": "Nom exact existant",
    "fk_pk_multi": "Noms exacts existants, séparés par «,»",
    "fk_name_value": "Nom exact existant",
    "fk_name_value_multi": "Noms exacts existants, séparés par «,»",
}

_FK_KINDS = {"fk_pk", "fk_pk_multi", "fk_name_value", "fk_name_value_multi"}
_MAX_VALEURS_AFFICHEES = 40


def _valeurs_autorisees(col):
    if col.kind == "choice_static":
        return ", ".join(f"{code} ({label})" for code, label in (col.choices or []))
    if col.kind in _FK_KINDS:
        try:
            noms = list(
                col.fk_model.objects.order_by(col.fk_lookup_field)
                .values_list(col.fk_lookup_field, flat=True)
                .distinct()
            )
        except Exception:
            # Base indisponible : le modele reste generable, sans la liste des
            # valeurs existantes.
            return "(liste indisponible pour le moment — consultez la liste correspondante dans l'application)"
        if not noms:
            return "(aucun(e) existant(e) en base pour le moment)"
        tronque = noms[:_MAX_VALEURS_AFFICHEES]
        suffixe = f" … (+{len(noms) - _MAX_VALEURS_AFFICHEES} autre(s))" if len(noms) > _MAX_VALEURS_AFFICHEES else ""
        return ", ".join(str(n) for n in tronque) + suffixe
    return col.help_text or "—"


def build_template_workbook(spec):
    wb = Workbook()

    ws = wb.active
    ws.title = spec.sheet_name

    rouge_fill = PatternFill("solid", fgColor=ROUGE)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col in enumerate(spec.columns, start=1):
        header_text = col.header + (" *" if col.required else "")
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = rouge_fill
        cell.font = header_font
        cell.alignment = center_align
        width = max(16, min(40, len(header_text) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    # Lignes deja remplies : on livre l'existant plutot qu'une feuille vide, ce
    # qui evite une ressaisie integrale et montre le format attendu sur de
    # vraies donnees. La base n'est pas modifiee par ce telechargement.
    if spec.prefill_fn:
        for ligne_idx, valeurs in enumerate(spec.prefill_fn(), start=2):
            for col_idx, col in enumerate(spec.columns, start=1):
                valeur = valeurs.get(col.field_name)
                if isinstance(valeur, bool):
                    valeur = "oui" if valeur else "non"
                ws.cell(row=ligne_idx, column=col_idx, value=valeur)

    # Feuille d'exemple : montre le format sur des donnees fictives, sans
    # risque puisqu'elle n'est pas lue a l'import. Toujours presente, meme
    # quand la feuille d'import est deja remplie.
    if spec.exemples:
        exemple = wb.create_sheet("Exemple")
        note = exemple.cell(row=1, column=1,
                            value="Feuille de demonstration : elle n'est PAS importee. "
                                  "Saisissez vos donnees sur la feuille « %s »." % spec.sheet_name)
        note.font = Font(bold=True, color="9C6500")
        note.fill = PatternFill("solid", fgColor="FFF4CE")
        exemple.merge_cells(start_row=1, start_column=1,
                            end_row=1, end_column=max(1, len(spec.columns)))
        exemple.row_dimensions[1].height = 26
        for col_idx, col in enumerate(spec.columns, start=1):
            cellule = exemple.cell(row=2, column=col_idx, value=col.header)
            cellule.fill = rouge_fill
            cellule.font = header_font
            cellule.alignment = center_align
            exemple.column_dimensions[get_column_letter(col_idx)].width = \
                max(16, min(40, len(col.header) + 4))
        for ligne_idx, valeurs in enumerate(spec.exemples, start=3):
            for col_idx, col in enumerate(spec.columns, start=1):
                valeur = valeurs.get(col.field_name)
                if isinstance(valeur, bool):
                    valeur = "oui" if valeur else "non"
                exemple.cell(row=ligne_idx, column=col_idx, value=valeur)
        exemple.freeze_panes = "A3"

    instructions = wb.create_sheet("Instructions")
    or_fill = PatternFill("solid", fgColor=OR)
    entetes = ["Colonne", "Obligatoire", "Type", "Valeurs autorisées / format"]
    for col_idx, texte in enumerate(entetes, start=1):
        cell = instructions.cell(row=1, column=col_idx, value=texte)
        cell.fill = or_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = center_align

    row_idx = 2
    if spec.intro:
        instructions.cell(row=row_idx, column=1, value=spec.intro)
        instructions.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        instructions.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
        row_idx += 1

    for col in spec.columns:
        instructions.cell(row=row_idx, column=1, value=col.header)
        instructions.cell(row=row_idx, column=2, value="Oui" if col.required else "Non")
        instructions.cell(row=row_idx, column=3, value=_KIND_LABELS.get(col.kind, col.kind))
        valeurs_cell = instructions.cell(row=row_idx, column=4, value=_valeurs_autorisees(col))
        valeurs_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1

    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 26
    instructions.column_dimensions["D"].width = 70

    wb.active = wb.sheetnames.index(spec.sheet_name)
    return wb
