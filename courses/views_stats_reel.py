"""Statistiques réelles : effectifs formés (agrégés) + liste nominative des
apprenants en résidentielle, par Direction Inter-Régionale.

Fichier volontairement SÉPARÉ du module "Statistiques" existant
(`courses/views.py`) pour ne rien mélanger avec ce qui est déjà en place —
seuls `_get_scope` et `_pdf_header_lines` en sont réutilisés (import direct).

Contrairement au moteur `courses/bulk_import/` (conçu pour CRÉER des objets
neufs), ce module MET À JOUR des `Inscription`/`Eleve` déjà existants et
crée/actualise des `EffectifReel` — seuls `resolve_column` (validation d'une
cellule) et `read_rows` (lecture xlsx générique) en sont réutilisés, pas le
moteur de création `run_import`.
"""

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone

import weasyprint
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from courses.bulk_import.engine import resolve_column
from courses.bulk_import.readers import read_rows
from courses.bulk_import.specs import ColumnSpec, ImportSpec

from .models import AnneeScolaire, CentreEtFiliere, EffectifReel, Filiere, Inscription
from .permissions import require_permission
from .views import _get_scope, _pdf_header_lines, _pdf_logo_data_uri

ROUGE = "C0392B"
OR = "D4A017"

NOMINATIF_SHEET = "Liste nominative"
EFFECTIFS_SHEET = "Effectifs"


# ── Specs de colonnes (réutilisent ColumnSpec/resolve_column/read_rows du module d'import générique, sans passer par son moteur de création) ───

_NOMINATIF_COLUMNS = [
    ColumnSpec("ID (ne pas modifier)", "id", required=True, kind="int"),
    ColumnSpec("Nationalité", "nationalite", required=True, kind="text"),
    ColumnSpec("Niveau scolaire", "niveau_scolaire", required=True, kind="text"),
    ColumnSpec("Mode de formation (Résidentielle, Duale)", "mode_formation", required=True, kind="choice_static",
                choices=Inscription.MODE_FORMATION_CHOICES),
    ColumnSpec("Observations", "observations", required=False, kind="text"),
]
_NOMINATIF_SPEC = ImportSpec(
    slug="stats_reel_nominatif", verbose_name="Liste nominative", model=Inscription,
    columns=_NOMINATIF_COLUMNS, sheet_name=NOMINATIF_SHEET,
)

_EFFECTIFS_COLUMNS = [
    ColumnSpec("Site", "site", required=False, kind="text"),
    ColumnSpec("Effectif inscrits hommes", "effectif_hommes", required=True, kind="int"),
    ColumnSpec("Effectif inscrits femmes", "effectif_femmes", required=True, kind="int"),
    ColumnSpec("Effectif présents hommes", "effectif_hommes_presents", required=True, kind="int",
                help_text="Apprenants inscrits présents à l'examen de certification."),
    ColumnSpec("Effectif présents femmes", "effectif_femmes_presents", required=True, kind="int"),
    ColumnSpec("Effectif admis hommes", "effectif_hommes_admis", required=True, kind="int",
                help_text="Apprenants admis (reçus) à l'examen de certification."),
    ColumnSpec("Effectif admis femmes", "effectif_femmes_admis", required=True, kind="int"),
    # Six categories, pre-suggerees depuis le handicap declare a l'inscription
    # tant qu'aucun EffectifReel n'existe. Le DSI corrige avant televersement.
    ColumnSpec("Handicap moteur — hommes", "effectif_hommes_handicap_moteur", required=False, kind="int",
                help_text="Apprenants en situation de handicap moteur."),
    ColumnSpec("Handicap moteur — femmes", "effectif_femmes_handicap_moteur", required=False, kind="int"),
    ColumnSpec("Handicap sensoriel (visuel) — hommes", "effectif_hommes_handicap_visuel", required=False, kind="int"),
    ColumnSpec("Handicap sensoriel (visuel) — femmes", "effectif_femmes_handicap_visuel", required=False, kind="int"),
    ColumnSpec("Handicap sensoriel (auditif) — hommes", "effectif_hommes_handicap_auditif", required=False, kind="int"),
    ColumnSpec("Handicap sensoriel (auditif) — femmes", "effectif_femmes_handicap_auditif", required=False, kind="int"),
    ColumnSpec("Maladie invalidante — épilepsie (hommes)", "effectif_hommes_epilepsie", required=False, kind="int"),
    ColumnSpec("Maladie invalidante — épilepsie (femmes)", "effectif_femmes_epilepsie", required=False, kind="int"),
    ColumnSpec("Maladie invalidante — asthme (hommes)", "effectif_hommes_asthme", required=False, kind="int"),
    ColumnSpec("Maladie invalidante — asthme (femmes)", "effectif_femmes_asthme", required=False, kind="int"),
    ColumnSpec("Autres maladies — hommes", "effectif_hommes_autres_maladies", required=False, kind="int",
                help_text="Toute autre maladie invalidante non listée ci-dessus."),
    ColumnSpec("Autres maladies — femmes", "effectif_femmes_autres_maladies", required=False, kind="int"),
]
_EFFECTIFS_SPEC = ImportSpec(
    slug="stats_reel_effectifs", verbose_name="Effectifs", model=EffectifReel,
    columns=_EFFECTIFS_COLUMNS, sheet_name=EFFECTIFS_SHEET,
)

# Colonnes affichées mais NON saisies dans la feuille nominative (pré-remplies,
# lecture seule côté DSI — non revalidées à l'upload).
_NOMINATIF_READONLY_HEADERS = [
    "N°", "Nom", "Prénom(s)", "Sexe", "Centre", "Métier", "Niveau de formation",
    "Type de formation", "Date de naissance", "Téléphone", "Tuteur/parent",
]


def _formation_in_scope(user, formation):
    if user.is_superuser:
        return True
    centres_qs, _, _ = _get_scope(user)
    return centres_qs.filter(pk=formation.centre_id).exists()


def _formations_scope_qs(user):
    centres_qs, _, _ = _get_scope(user)
    return CentreEtFiliere.objects.filter(centre__in=centres_qs).select_related(
        'centre', 'centre__direction', 'filiere', 'annee_prog'
    )


def _subtotal_label(centre):
    nom = centre.nom_centre or ""
    if "Provincial" in nom:
        return "Total CPFP"
    if "Régional" in nom or "Regional" in nom:
        return "Total CRFP"
    return "Total"


# Ordre impose par le tableau des resultats aux examens : 3 sections x 4 titres
# x H/F/T, soit 36 colonnes.
TITRES_PROFESSIONNELS = ['CQP', 'BQP', 'BPT', 'BPTS']


def _build_certification_columns(titre, inscrits, presents, admis):
    """36 valeurs (section × titre × H/F/T) pour une ligne (une formation/site) —
    vides partout sauf sous le bloc du titre professionnel de cette ligne
    (une formation n'a qu'un seul titre professionnel, celui de son métier)."""
    columns = []
    for h, f in (inscrits, presents, admis):
        for t in TITRES_PROFESSIONNELS:
            if t == titre:
                columns.extend([h, f, h + f])
            else:
                columns.extend(["", "", ""])
    return columns


# Ordre impose par le modele de reference du tableau handicap.
_HANDICAP_CATEGORIES = [
    ('effectif_hommes_handicap_moteur', 'effectif_femmes_handicap_moteur'),
    ('effectif_hommes_handicap_visuel', 'effectif_femmes_handicap_visuel'),
    ('effectif_hommes_handicap_auditif', 'effectif_femmes_handicap_auditif'),
    ('effectif_hommes_epilepsie', 'effectif_femmes_epilepsie'),
    ('effectif_hommes_asthme', 'effectif_femmes_asthme'),
    ('effectif_hommes_autres_maladies', 'effectif_femmes_autres_maladies'),
]

# Pre-remplit la feuille Effectifs depuis le handicap declare ; la saisie du
# DSI prevaut ensuite sur cette suggestion.
_TYPE_HANDICAP_VERS_CHAMPS = {
    'moteur': ('effectif_hommes_handicap_moteur', 'effectif_femmes_handicap_moteur'),
    'sensoriel_visuel': ('effectif_hommes_handicap_visuel', 'effectif_femmes_handicap_visuel'),
    'sensoriel_auditif': ('effectif_hommes_handicap_auditif', 'effectif_femmes_handicap_auditif'),
    'maladie_epilepsie': ('effectif_hommes_epilepsie', 'effectif_femmes_epilepsie'),
    'maladie_asthme': ('effectif_hommes_asthme', 'effectif_femmes_asthme'),
    'autre': ('effectif_hommes_autres_maladies', 'effectif_femmes_autres_maladies'),
}


def _build_handicap_columns(sums):
    """21 valeurs (6 catégories × H/F/T, puis Total H/F/T) à partir d'un dict
    {nom_du_champ: somme} portant sur les 12 champs handicap/maladies de
    EffectifReel (une ligne = un métier, agrégé sur toutes les formations/sites
    concernés)."""
    columns = []
    total_h = total_f = 0
    for champ_h, champ_f in _HANDICAP_CATEGORIES:
        h, f = sums.get(champ_h) or 0, sums.get(champ_f) or 0
        columns.extend([h, f, h + f])
        total_h += h
        total_f += f
    columns.extend([total_h, total_f, total_h + total_f])
    return columns


def _tuteur_display(inscription):
    parts = [inscription.personne_contact_nom, inscription.personne_contact_prenom]
    nom = " ".join(p for p in parts if p)
    tel = inscription.personne_contact_tel
    if nom and tel:
        return f"{nom} ({tel})"
    return nom or tel or ""


def _filter_options(user):
    """Options des menus déroulants (année/direction/centre/filière) pour le
    dashboard et les 3 exports — toujours limitées au périmètre de l'utilisateur."""
    centres_qs, directions_qs, scope = _get_scope(user)
    return {
        'annees': AnneeScolaire.objects.order_by('-libelle_anne'),
        'directions': directions_qs.order_by('nom_direction') if scope == 'global' else None,
        'centres': centres_qs.order_by('nom_centre') if scope in ('global', 'direction') else None,
        'filieres': Filiere.objects.filter(
            centreetfiliere__centre__in=centres_qs
        ).distinct().order_by('nom_filiere'),
    }


def _annee_suffix(annee_id):
    """Suffixe de nom de fichier indiquant l'année scolaire filtrée (vide si
    aucune année n'est sélectionnée — export toutes années confondues)."""
    if not annee_id:
        return ""
    annee = AnneeScolaire.objects.filter(pk=annee_id).first()
    return f"_{annee.libelle_anne}".replace(" ", "").replace("/", "-") if annee else ""


def _pdf_header_scope(scope, centres_qs, directions_qs, direction_id, centre_id):
    """Résout (centre, direction) à afficher dans l'en-tête des PDF exportés,
    après "DIRECTION GENERALE" — jamais les deux à vide-mais-affichées
    (voir `_pdf_header_lines`, qui n'ajoute une ligne que si l'objet existe).

    Priorité : le filtre explicitement choisi sur l'écran (direction/centre),
    sinon la portée propre de l'utilisateur (sa direction pour un directeur
    inter-régional, son centre pour un gestionnaire/caissier) — rien pour une
    portée globale non filtrée (admin/DG/DEPS/agent comptable)."""
    if centre_id and scope in ('global', 'direction'):
        return centres_qs.filter(pk=centre_id).first(), None
    if direction_id and scope == 'global':
        return None, directions_qs.filter(pk=direction_id).first()
    if scope == 'centre':
        return centres_qs.first(), None
    if scope == 'direction':
        return None, directions_qs.first()
    return None, None


def _selected_filters(request):
    return {
        'annee_id': request.GET.get('annee', ''),
        'direction_id': request.GET.get('direction', ''),
        'centre_id': request.GET.get('centre', ''),
        'filiere_id': request.GET.get('filiere', ''),
    }


# ── Dashboard : liste des formations du périmètre ─────────────────────────────

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_dashboard(request):
    centres_qs, directions_qs, scope = _get_scope(request.user)
    selected = _selected_filters(request)

    formations = _formations_scope_qs(request.user)

    # Une formation est reutilisee d'une annee sur l'autre : le filtre annee se
    # base sur Inscription.annee_scolaire, pas sur CentreEtFiliere.annee_prog.
    if selected['annee_id']:
        annee_q = Q(inscription__annee_scolaire_id=selected['annee_id'])
        formations = formations.filter(inscription__annee_scolaire_id=selected['annee_id']).distinct()
    else:
        annee_q = Q()

    formations = formations.annotate(
        nb_inscrits=Count('inscription', filter=annee_q, distinct=True),
        nb_mode_renseigne=Count(
            'inscription', filter=annee_q & Q(inscription__mode_formation__isnull=False), distinct=True
        ),
        nb_effectifs=Count(
            'effectifs_reels',
            filter=Q(effectifs_reels__annee_scolaire_id=selected['annee_id']) if selected['annee_id'] else Q(),
            distinct=True,
        ),
    )

    if selected['direction_id'] and scope == 'global':
        formations = formations.filter(centre__direction_id=selected['direction_id'])
    if selected['centre_id'] and scope in ('global', 'direction'):
        formations = formations.filter(centre_id=selected['centre_id'])
    if selected['filiere_id']:
        formations = formations.filter(filiere_id=selected['filiere_id'])

    formations = formations.order_by('centre__nom_centre', 'filiere__nom_filiere')

    q = request.GET.get('q', '').strip()
    if q:
        formations = formations.filter(
            Q(centre__nom_centre__icontains=q) | Q(filiere__nom_filiere__icontains=q)
        )

    paginator = Paginator(formations, 10)
    formations_page = paginator.get_page(request.GET.get('page'))

    options = _filter_options(request.user)

    return render(request, 'member/stats_reel/dashboard.html', {
        'formations': formations_page,
        'q': q,
        'scope': scope,
        'selected': selected,
        'annees': options['annees'],
        'directions': options['directions'],
        'centres': options['centres'],
        'filieres': options['filieres'],
    })


def _resolve_annee(request, formation):
    """Année scolaire à utiliser pour le modèle/upload d'une formation.

    Une formation (CentreEtFiliere) n'étant PAS propre à une année (elle est
    réutilisée d'une année sur l'autre — voir EffectifReel plus haut), il faut
    déterminer explicitement l'année concernée : celle passée en paramètre
    (`?annee=`, reprise depuis le filtre du dashboard), ou à défaut la plus
    récente année ayant des inscriptions pour cette formation.
    """
    annee_id = request.GET.get('annee')
    if annee_id:
        return get_object_or_404(AnneeScolaire, pk=annee_id)
    return (
        AnneeScolaire.objects.filter(inscription__formation=formation)
        .order_by('-libelle_anne').first()
        or AnneeScolaire.objects.order_by('-libelle_anne').first()
    )


# ── Téléchargement du modèle Excel pour une formation ─────────────────────────

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_formation_template(request, formation_id):
    formation = get_object_or_404(
        CentreEtFiliere.objects.select_related('centre', 'filiere'), pk=formation_id
    )
    if not _formation_in_scope(request.user, formation):
        messages.error(request, "Vous n'avez pas accès à cette formation.")
        return redirect('courses:stats_reel_dashboard')

    annee = _resolve_annee(request, formation)
    if not annee:
        messages.error(request, "Aucune année de formation n'existe encore dans l'application.")
        return redirect('courses:stats_reel_dashboard')

    wb = Workbook()
    rouge_fill = PatternFill("solid", fgColor=ROUGE)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Feuille 1 : Liste nominative ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = NOMINATIF_SHEET

    headers = (
        ["ID (ne pas modifier)"] + _NOMINATIF_READONLY_HEADERS
        + [c.header + (" *" if c.required else "") for c in _NOMINATIF_COLUMNS[1:]]
    )
    for col_idx, text in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=text)
        cell.fill = rouge_fill
        cell.font = header_font
        cell.alignment = center_align
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(16, min(40, len(text) + 4))
    ws1.freeze_panes = "A2"
    # Colonne technique ID : cachée pour ne pas inciter à la modifier.
    ws1.column_dimensions["A"].hidden = True

    inscriptions = (
        formation.inscription_set
        .select_related('eleve')
        .exclude(eleve__isnull=True)
        .filter(annee_scolaire=annee)
        .order_by('eleve__nom', 'eleve__prenom')
    )
    for numero, inscription in enumerate(inscriptions, start=1):
        eleve = inscription.eleve
        ws1.append([
            inscription.id,
            numero,
            eleve.nom,
            eleve.prenom,
            eleve.get_sexe_display(),
            formation.centre.nom_centre,
            formation.filiere.nom_filiere,
            formation.filiere.titre_professionnel or "",
            formation.get_type_formation_display() if formation.type_formation else "",
            eleve.date_naissance.strftime("%d/%m/%Y") if eleve.date_naissance else "",
            eleve.tel or "",
            _tuteur_display(inscription),
            eleve.nationalite or "Burkinabè",
            eleve.niveau_scolaire or "",
            inscription.get_mode_formation_display() if inscription.mode_formation else "",
            inscription.observations_reel or "",
        ])

    # Feuille 2, effectifs editables : les inscrits sont pre-calcules, presents
    # et admis relevent de l'examen de certification et restent a saisir.
    ws2 = wb.create_sheet(EFFECTIFS_SHEET)
    eff_headers = ["Site"] + [c.header + (" *" if c.required else "") for c in _EFFECTIFS_COLUMNS[1:]]
    for col_idx, text in enumerate(eff_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=text)
        cell.fill = rouge_fill
        cell.font = header_font
        cell.alignment = center_align
        ws2.column_dimensions[get_column_letter(col_idx)].width = 24
    ws2.freeze_panes = "A2"

    existants = list(formation.effectifs_reels.filter(annee_scolaire=annee).order_by('site'))
    if existants:
        for eff in existants:
            ws2.append([eff.site] + [getattr(eff, col.field_name) for col in _EFFECTIFS_COLUMNS[1:]])
    else:
        h = inscriptions.filter(eleve__sexe='m').count()
        f = inscriptions.filter(eleve__sexe='f').count()
        # Presents et admis restent a 0 : l'app ne les connait pas. Le handicap
        # est une suggestion, corrigeable avant televersement.
        valeurs_suggerees = {'effectif_hommes': h, 'effectif_femmes': f}
        for type_code, (champ_h, champ_f) in _TYPE_HANDICAP_VERS_CHAMPS.items():
            valeurs_suggerees[champ_h] = inscriptions.filter(
                eleve__sexe='m', eleve__a_handicap=True, eleve__type_handicap=type_code
            ).count()
            valeurs_suggerees[champ_f] = inscriptions.filter(
                eleve__sexe='f', eleve__a_handicap=True, eleve__type_handicap=type_code
            ).count()
        ws2.append([""] + [valeurs_suggerees.get(col.field_name, 0) for col in _EFFECTIFS_COLUMNS[1:]])

    # ── Feuille 3 : Instructions ───────────────────────────────────────────
    ws3 = wb.create_sheet("Instructions")
    or_fill = PatternFill("solid", fgColor=OR)
    for col_idx, text in enumerate(["Feuille", "Colonne", "Obligatoire", "Aide"], start=1):
        cell = ws3.cell(row=1, column=col_idx, value=text)
        cell.fill = or_fill
        cell.font = header_font
        cell.alignment = center_align
    ws3.append([
        "—", f"Année de formation : {annee.libelle_anne}", "—",
        "Ce fichier ne couvre que cette année de formation — les élèves d'une autre année ne sont pas inclus."
    ])
    ws3.append([
        "—", "Colonnes Handicap (feuille Effectifs)", "—",
        "Pré-suggérées à partir du handicap déclaré par les apprenants à l'inscription. "
        "À vérifier et corriger si besoin avant de televerser (cette valeur n'est pré-remplie que "
        "la première fois, tant qu'aucun effectif n'a encore été enregistré pour cette formation)."
    ])
    ws3.append([NOMINATIF_SHEET, "(colonnes N° à Tuteur/parent)", "Non", "Déjà connu de l'application — pré-rempli, ne pas modifier."])
    for col in _NOMINATIF_COLUMNS[1:]:
        ws3.append([NOMINATIF_SHEET, col.header, "Oui" if col.required else "Non", col.help_text or "À saisir."])
    for col in _EFFECTIFS_COLUMNS:
        if col.help_text:
            aide = col.help_text
        elif col.field_name == "site":
            aide = "Laisser vide pour désigner le centre principal (pas un site annexe)."
        else:
            aide = "Nombre entier ≥ 0."
        ws3.append([EFFECTIFS_SHEET, col.header, "Oui" if col.required else "Non", aide])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 60

    wb.active = 0

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nom_fichier = f"stat_reelle_{formation.centre.nom_centre}_{formation.filiere.nom_filiere}_{annee.libelle_anne}".replace(" ", "_").replace("/", "-")
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}.xlsx"'
    wb.save(response)
    return response


# ── Upload du fichier rempli ──────────────────────────────────────────────────

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_formation_upload(request, formation_id):
    formation = get_object_or_404(
        CentreEtFiliere.objects.select_related('centre', 'filiere'), pk=formation_id
    )
    if not _formation_in_scope(request.user, formation):
        messages.error(request, "Vous n'avez pas accès à cette formation.")
        return redirect('courses:stats_reel_dashboard')

    annee = _resolve_annee(request, formation)
    if not annee:
        messages.error(request, "Aucune année de formation n'existe encore dans l'application.")
        return redirect('courses:stats_reel_dashboard')

    if request.method != 'POST':
        return render(request, 'member/stats_reel/upload_form.html', {'formation': formation, 'annee': annee})

    uploaded_file = request.FILES.get('fichier')
    if not uploaded_file:
        messages.error(request, "Aucun fichier n'a été sélectionné.")
        return redirect('courses:stats_reel_formation_upload', formation_id=formation.id)

    if not (uploaded_file.name or '').lower().endswith('.xlsx'):
        messages.error(request, "Le fichier doit être au format .xlsx (2 feuilles obligatoires) — .csv non pris en charge ici.")
        return redirect('courses:stats_reel_formation_upload', formation_id=formation.id)

    nominatif_result = {'updated': [], 'errors': []}
    effectifs_result = {'updated': [], 'errors': []}

    try:
        uploaded_file.seek(0)
        nominatif_rows = read_rows(uploaded_file, _NOMINATIF_SPEC)
    except ValueError as e:
        nominatif_rows = []
        nominatif_result['errors'].append((0, [str(e)]))

    try:
        uploaded_file.seek(0)
        effectifs_rows = read_rows(uploaded_file, _EFFECTIFS_SPEC)
    except ValueError as e:
        effectifs_rows = []
        effectifs_result['errors'].append((0, [str(e)]))

    for row_number, raw_row in nominatif_rows:
        resolved = {}
        errors = []
        for col in _NOMINATIF_COLUMNS:
            ok, value, err = resolve_column(col, raw_row.get(col.header))
            if ok:
                resolved[col.field_name] = value
            else:
                errors.append(err)
        if errors:
            nominatif_result['errors'].append((row_number, errors))
            continue

        try:
            with transaction.atomic():
                inscription = Inscription.objects.select_related('eleve').get(pk=resolved['id'])
                if inscription.formation_id != formation.id or inscription.annee_scolaire_id != annee.id:
                    raise Inscription.DoesNotExist
                if not inscription.eleve:
                    nominatif_result['errors'].append((row_number, ["Cette inscription n'a pas d'apprenant associé."]))
                    continue
                eleve = inscription.eleve
                eleve.nationalite = resolved['nationalite']
                eleve.niveau_scolaire = resolved['niveau_scolaire']
                eleve.save(update_fields=['nationalite', 'niveau_scolaire'])
                inscription.mode_formation = resolved['mode_formation']
                inscription.observations_reel = resolved.get('observations') or ''
                inscription.save(update_fields=['mode_formation', 'observations_reel'])
            nominatif_result['updated'].append((row_number, str(eleve)))
        except Inscription.DoesNotExist:
            nominatif_result['errors'].append((row_number, [
                f"Aucune inscription n°{resolved['id']} trouvée pour cette formation et cette année de formation "
                f"({annee.libelle_anne}) — ID incorrect, ligne d'une autre formation, ou d'une autre année."
            ]))
        except Exception as e:
            nominatif_result['errors'].append((row_number, [str(e)]))

    for row_number, raw_row in effectifs_rows:
        resolved = {}
        errors = []
        for col in _EFFECTIFS_COLUMNS:
            ok, value, err = resolve_column(col, raw_row.get(col.header))
            if ok:
                resolved[col.field_name] = value
            else:
                errors.append(err)
        if errors:
            effectifs_result['errors'].append((row_number, errors))
            continue

        try:
            with transaction.atomic():
                site = resolved.get('site') or ''
                # Colonnes handicap facultatives : cellule vide ramenee a 0, le
                # champ n'etant pas nullable.
                defaults = {
                    col.field_name: (resolved.get(col.field_name) or 0)
                    for col in _EFFECTIFS_COLUMNS if col.field_name != 'site'
                }
                obj, _created = EffectifReel.objects.update_or_create(
                    formation=formation, site=site, annee_scolaire=annee,
                    defaults=defaults,
                )
            effectifs_result['updated'].append((row_number, str(obj)))
        except Exception as e:
            effectifs_result['errors'].append((row_number, [str(e)]))

    return render(request, 'member/stats_reel/upload_result.html', {
        'formation': formation,
        'annee': annee,
        'nominatif_result': nominatif_result,
        'effectifs_result': effectifs_result,
    })


# ── Export PDF — liste nominative des apprenants en résidentielle, par DIR ────

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_export_nominatif(request):
    centres_qs, directions_qs, scope = _get_scope(request.user)

    qs = Inscription.objects.filter(
        mode_formation='residentielle',
        formation__centre__in=centres_qs,
    ).select_related('eleve', 'formation__centre__direction', 'formation__filiere', 'annee_scolaire')

    direction_id = request.GET.get('direction')
    centre_id = request.GET.get('centre')
    filiere_id = request.GET.get('filiere')
    annee_id = request.GET.get('annee')

    if direction_id and scope == 'global':
        qs = qs.filter(formation__centre__direction_id=direction_id)
    if centre_id and scope in ('global', 'direction'):
        qs = qs.filter(formation__centre_id=centre_id)
    if filiere_id:
        qs = qs.filter(formation__filiere_id=filiere_id)
    if annee_id:
        qs = qs.filter(annee_scolaire_id=annee_id)

    qs = qs.exclude(eleve__isnull=True).order_by(
        'formation__centre__direction__nom_direction',
        'formation__centre__nom_centre',
        'eleve__nom', 'eleve__prenom',
    )

    directions = {}
    for inscription in qs:
        direction = inscription.formation.centre.direction
        key = direction.nom_direction if direction else "Sans direction rattachée"
        directions.setdefault(key, []).append(inscription)

    centre_entete, direction_entete = _pdf_header_scope(scope, centres_qs, directions_qs, direction_id, centre_id)
    header_left, header_right = _pdf_header_lines(centre=centre_entete, direction=direction_entete)
    annee_obj = AnneeScolaire.objects.filter(pk=annee_id).first() if annee_id else None
    html_string = render_to_string('member/stats_reel/nominatif_pdf.html', {
        'directions': directions,
        'header_left': header_left,
        'header_right': header_right,
        'date_generation': timezone.now(),
        'annee_libelle': annee_obj.libelle_anne if annee_obj else "toutes années",
        'favicon_data_uri': _pdf_logo_data_uri(),
    })
    pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="liste_nominative_residentielle{_annee_suffix(annee_id)}.pdf"'
    return response


# ── Export Excel — situation des formés (effectifs), par DIR ──────────────────

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_export_agrege(request):
    centres_qs, directions_qs, scope = _get_scope(request.user)

    qs = EffectifReel.objects.filter(formation__centre__in=centres_qs).select_related(
        'formation__centre__direction', 'formation__filiere'
    )

    direction_id = request.GET.get('direction')
    centre_id = request.GET.get('centre')
    filiere_id = request.GET.get('filiere')
    annee_id = request.GET.get('annee')

    if direction_id and scope == 'global':
        qs = qs.filter(formation__centre__direction_id=direction_id)
    if centre_id and scope in ('global', 'direction'):
        qs = qs.filter(formation__centre_id=centre_id)
    if filiere_id:
        qs = qs.filter(formation__filiere_id=filiere_id)
    if annee_id:
        qs = qs.filter(annee_scolaire_id=annee_id)

    qs = qs.order_by(
        'formation__centre__direction__nom_direction',
        'formation__centre__nom_centre',
        'site',
        'formation__filiere__nom_filiere',
    )

    directions_presentes = {}
    for eff in qs:
        direction = eff.formation.centre.direction
        key = direction.nom_direction if direction else "Sans direction rattachée"
        directions_presentes.setdefault(key, []).append(eff)
    if not directions_presentes:
        directions_presentes = {"Aucune donnée": []}

    wb = Workbook()
    wb.remove(wb.active)

    rouge_fill = PatternFill("solid", fgColor=ROUGE)
    or_fill = PatternFill("solid", fgColor=OR)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for direction_nom, effectifs in directions_presentes.items():
        ws = wb.create_sheet(direction_nom[:31])
        headers = ["Direction inter Régionale", "Centre/site de formation", "Métier de formation", "H", "F", "Total"]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.fill = rouge_fill
            cell.font = header_font
            cell.alignment = center_align
            ws.column_dimensions[get_column_letter(col_idx)].width = 26

        row_idx = 2
        first_row_written = False
        current_centre = None
        centre_h_total = 0
        centre_f_total = 0
        dir_h_total = 0
        dir_f_total = 0

        def ecrire_soustotal_centre():
            nonlocal row_idx
            label = _subtotal_label(current_centre)
            ws.cell(row=row_idx, column=2, value=label).font = bold_font
            ws.cell(row=row_idx, column=4, value=centre_h_total).font = bold_font
            ws.cell(row=row_idx, column=5, value=centre_f_total).font = bold_font
            ws.cell(row=row_idx, column=6, value=centre_h_total + centre_f_total).font = bold_font
            row_idx += 1

        for eff in effectifs:
            centre = eff.formation.centre
            if current_centre is not None and centre.id != current_centre.id:
                ecrire_soustotal_centre()
                centre_h_total = 0
                centre_f_total = 0
            current_centre = centre
            centre_h_total += eff.effectif_hommes
            centre_f_total += eff.effectif_femmes
            dir_h_total += eff.effectif_hommes
            dir_f_total += eff.effectif_femmes

            site_label = f"{centre.nom_centre} / {eff.site}" if eff.site else centre.nom_centre
            ws.cell(row=row_idx, column=1, value=direction_nom if not first_row_written else "")
            ws.cell(row=row_idx, column=2, value=site_label)
            ws.cell(row=row_idx, column=3, value=eff.formation.filiere.nom_filiere)
            ws.cell(row=row_idx, column=4, value=eff.effectif_hommes)
            ws.cell(row=row_idx, column=5, value=eff.effectif_femmes)
            ws.cell(row=row_idx, column=6, value=eff.effectif_total)
            first_row_written = True
            row_idx += 1

        if current_centre is not None:
            ecrire_soustotal_centre()

        for col_idx, val in enumerate(
            ["", f"Total {direction_nom}", "", dir_h_total, dir_f_total, dir_h_total + dir_f_total], start=1
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = or_fill
            cell.font = header_font

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="situation_formes_agrege{_annee_suffix(annee_id)}.xlsx"'
    wb.save(response)
    return response


# ── Export PDF — RESULTATS AUX EXAMENS DE CERTIFICATION (Inscrits / Présents / Admis, par titre professionnel, par DIR) ───

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_export_certification(request):
    centres_qs, directions_qs, scope = _get_scope(request.user)

    qs = EffectifReel.objects.filter(formation__centre__in=centres_qs).select_related(
        'formation__centre__direction', 'formation__filiere'
    )

    direction_id = request.GET.get('direction')
    centre_id = request.GET.get('centre')
    filiere_id = request.GET.get('filiere')
    annee_id = request.GET.get('annee')

    if direction_id and scope == 'global':
        qs = qs.filter(formation__centre__direction_id=direction_id)
    if centre_id and scope in ('global', 'direction'):
        qs = qs.filter(formation__centre_id=centre_id)
    if filiere_id:
        qs = qs.filter(formation__filiere_id=filiere_id)
    if annee_id:
        qs = qs.filter(annee_scolaire_id=annee_id)

    qs = qs.order_by(
        'formation__centre__direction__nom_direction',
        'formation__centre__nom_centre',
        'site',
        'formation__filiere__nom_filiere',
    )

    directions = {}
    for eff in qs:
        direction = eff.formation.centre.direction
        direction_key = direction.nom_direction if direction else "Sans direction rattachée"
        direction_bucket = directions.setdefault(direction_key, {'centres': {}, 'totaux': [0] * 36})

        centre = eff.formation.centre
        centre_bucket = direction_bucket['centres'].setdefault(centre.id, {
            'nom': centre.nom_centre, 'label': _subtotal_label(centre), 'rows': [], 'totaux': [0] * 36,
        })

        titre = eff.formation.filiere.titre_professionnel
        if titre not in TITRES_PROFESSIONNELS:
            titre = TITRES_PROFESSIONNELS[0]
        columns = _build_certification_columns(
            titre,
            (eff.effectif_hommes, eff.effectif_femmes),
            (eff.effectif_hommes_presents, eff.effectif_femmes_presents),
            (eff.effectif_hommes_admis, eff.effectif_femmes_admis),
        )
        centre_bucket['rows'].append({'site': eff.site, 'metier': eff.formation.filiere.nom_filiere, 'columns': columns})

        for i, v in enumerate(columns):
            if v != "":
                centre_bucket['totaux'][i] += v
                direction_bucket['totaux'][i] += v

    total_bsb = [0] * 36
    for direction_bucket in directions.values():
        for i, v in enumerate(direction_bucket['totaux']):
            total_bsb[i] += v

    centre_entete, direction_entete = _pdf_header_scope(scope, centres_qs, directions_qs, direction_id, centre_id)
    header_left, header_right = _pdf_header_lines(centre=centre_entete, direction=direction_entete)
    annee_obj = AnneeScolaire.objects.filter(pk=annee_id).first() if annee_id else None
    html_string = render_to_string('member/stats_reel/certification_pdf.html', {
        'directions': directions,
        'total_bsb': total_bsb,
        'titres': TITRES_PROFESSIONNELS,
        'header_left': header_left,
        'header_right': header_right,
        'date_generation': timezone.now(),
        'annee_libelle': annee_obj.libelle_anne if annee_obj else "toutes années",
        'favicon_data_uri': _pdf_logo_data_uri(),
    })
    pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="resultats_examens_certification{_annee_suffix(annee_id)}.pdf"'
    return response


# ── Export PDF — RÉPARTITION DES APPRENANTS VIVANT AVEC UN HANDICAP PAR MÉTIER ET PAR SEXE ───

@require_permission('courses.gerer_statistiques_reelles')
def stats_reel_export_handicap(request):
    centres_qs, directions_qs, scope = _get_scope(request.user)

    qs = EffectifReel.objects.filter(formation__centre__in=centres_qs)

    direction_id = request.GET.get('direction')
    centre_id = request.GET.get('centre')
    filiere_id = request.GET.get('filiere')
    annee_id = request.GET.get('annee')

    if direction_id and scope == 'global':
        qs = qs.filter(formation__centre__direction_id=direction_id)
    if centre_id and scope in ('global', 'direction'):
        qs = qs.filter(formation__centre_id=centre_id)
    if filiere_id:
        qs = qs.filter(formation__filiere_id=filiere_id)
    if annee_id:
        qs = qs.filter(annee_scolaire_id=annee_id)

    # Une ligne par metier, agregee sur tout le perimetre : ce tableau ne se
    # decompose pas par direction, contrairement aux deux autres exports.
    champs_sommes = [c for paire in _HANDICAP_CATEGORIES for c in paire]
    par_metier = (
        qs.values('formation__filiere__nom_filiere')
        .annotate(**{c: Sum(c) for c in champs_sommes})
        .order_by('formation__filiere__nom_filiere')
    )

    metiers = []
    total_sums = {c: 0 for c in champs_sommes}
    for row in par_metier:
        for c in champs_sommes:
            total_sums[c] += row[c] or 0
        metiers.append({
            'nom': row['formation__filiere__nom_filiere'],
            'columns': _build_handicap_columns(row),
        })

    centre_entete, direction_entete = _pdf_header_scope(scope, centres_qs, directions_qs, direction_id, centre_id)
    header_left, header_right = _pdf_header_lines(centre=centre_entete, direction=direction_entete)
    annee_obj = AnneeScolaire.objects.filter(pk=annee_id).first() if annee_id else None
    html_string = render_to_string('member/stats_reel/handicap_pdf.html', {
        'metiers': metiers,
        'total_columns': _build_handicap_columns(total_sums),
        'header_left': header_left,
        'header_right': header_right,
        'date_generation': timezone.now(),
        'annee_libelle': annee_obj.libelle_anne if annee_obj else "toutes années",
        'favicon_data_uri': _pdf_logo_data_uri(),
    })
    pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="repartition_handicap{_annee_suffix(annee_id)}.pdf"'
    return response
